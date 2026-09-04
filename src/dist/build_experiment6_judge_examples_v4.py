#!/usr/bin/env python3
"""Build the immutable 26-row Experiment 6 v4 judge reference bundle."""

from __future__ import annotations

import argparse
import ast
import hashlib
import json
import re
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping

import openpyxl

from experiment6_paths import PATHS
WORKSPACE_ROOT = PATHS.workspace

REPO_ROOT = PATHS.repo
FIELDS = ("ObjectName", "DataName", "Position", "Trend", "Num", "Text")
SEMANTIC_FIELDS = ("ObjectName", "Trend", "Text")
REPAIR_VERSION = "experiment6-judge-examples-canonical-v4"


class ProtocolError(RuntimeError):
    """Raised when the frozen reference workbook violates the v4 contract."""


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace(
        "+00:00", "Z"
    )


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def workspace_path(raw: str | Mapping[str, Any]) -> Path:
    if isinstance(raw, Mapping):
        if "root" in raw:
            return PATHS.resolve_locator(raw)
        raw = str(raw["path"])
    path = Path(raw)
    return path if path.is_absolute() else WORKSPACE_ROOT / path


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rendered = json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(rendered, encoding="utf-8")
    temporary.replace(path)


def write_jsonl(path: Path, values: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rendered = "".join(
        json.dumps(value, ensure_ascii=False, sort_keys=True) + "\n"
        for value in values
    )
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(rendered, encoding="utf-8")
    temporary.replace(path)


def parse_json_or_literal(raw: str) -> tuple[list[dict[str, Any]], str]:
    errors: list[str] = []
    for name, parser in (("json", json.loads), ("python-literal", ast.literal_eval)):
        try:
            value = parser(raw)
        except (ValueError, SyntaxError) as error:
            errors.append(f"{name}: {error}")
            continue
        if not isinstance(value, list) or not all(
            isinstance(item, dict) for item in value
        ):
            raise ProtocolError("result must decode to an array of binding objects")
        return [dict(item) for item in value], name
    raise ProtocolError("; ".join(errors))


def split_multi_anchor_binding(
    binding: Mapping[str, Any], data_names: list[str]
) -> list[dict[str, Any]]:
    objects = binding.get("ObjectName")
    positions = binding.get("Position")
    numbers = binding.get("Num")
    if not (
        isinstance(objects, list)
        and isinstance(positions, list)
        and isinstance(numbers, list)
        and len(objects) == len(positions) == len(numbers) == len(data_names)
    ):
        raise ProtocolError("Excel row 4 multi-anchor fields cannot be split one-to-one")
    result: list[dict[str, Any]] = []
    for index, data_name in enumerate(data_names):
        item = deepcopy(dict(binding))
        item["ObjectName"] = [objects[index]]
        item["DataName"] = data_name
        item["Position"] = [positions[index]]
        item["Num"] = [numbers[index]]
        result.append(item)
    return result


def repair_result(excel_row: int, raw: str) -> tuple[list[dict[str, Any]], list[dict[str, str]], str]:
    repairs: list[dict[str, str]] = []
    parser = "json"

    if excel_row == 4:
        bindings, parser = parse_json_or_literal(raw)
        bindings = (
            bindings[:2]
            + split_multi_anchor_binding(
                bindings[2], ["Compact size", "Midsize to large"]
            )
            + split_multi_anchor_binding(
                bindings[3], ["Compact size", "Midsize to large"]
            )
        )
        repairs.append({
            "type": "schema-adapter",
            "detail": "split two multi-DataName bindings into four fixed single-anchor bindings; correct the 2023 second series from Mini- and subcompact size to Midsize to large using data/reason",
        })
        return bindings, repairs, parser

    if excel_row == 5:
        bindings, parser = parse_json_or_literal(raw)
        bindings[2]["ObjectName"] = ["energy percentage"]
        repairs.append({
            "type": "coreference-adapter",
            "detail": "replace absent ObjectName with the explicit phrase energy percentage from text/reason",
        })
        return bindings, repairs, parser

    repaired = raw
    if excel_row == 10:
        repaired = repaired.replace(
            '"Text":"The "double bottom" pattern',
            '"Text":"The \\"double bottom\\" pattern',
        )
        repairs.append({"type": "json-syntax", "detail": "escape nested double-bottom quotation marks"})
    elif excel_row == 13:
        bindings, parser = parse_json_or_literal(raw)
        for binding in bindings:
            begin = binding.pop("BeginIndex")
            end = binding.pop("EndIndex")
            binding["Position"] = [{"Begin": begin, "End": end}]
        bindings[1]["Num"] = [31.0]
        repairs.extend([
            {"type": "schema-adapter", "detail": "convert BeginIndex/EndIndex to Position[]"},
            {"type": "numeric-adapter", "detail": "convert 31% string to finite JSON number array [31.0]"},
        ])
        return bindings, repairs, parser
    elif excel_row == 14:
        repaired = repaired.replace(
            '"Text":"The "triple top" pattern',
            '"Text":"The \\"triple top\\" pattern',
        )
        repairs.append({"type": "json-syntax", "detail": "escape nested triple-top quotation marks"})
    elif excel_row == 15:
        repaired = repaired.replace("[62.8%]", "[62.8]")
        repairs.append({"type": "numeric-adapter", "detail": "convert bare 62.8% token to finite JSON number 62.8"})
    elif excel_row == 18:
        repaired = repaired.replace(
            '"Num":"None,"Text":', '"Num":"None","Text":'
        )
        if not repaired.rstrip().endswith("]"):
            repaired = repaired.rstrip() + "]"
        repairs.append({"type": "json-syntax", "detail": "close the None string and outer binding array"})
    elif excel_row == 24:
        repaired = repaired.replace("[70.64%]", "[70.64]")
        repairs.append({"type": "numeric-adapter", "detail": "convert bare 70.64% token to finite JSON number 70.64"})
    elif excel_row == 27:
        repaired = repaired.replace("”", '"').replace("“", '"')
        repairs.append({"type": "json-syntax", "detail": "normalize typographic JSON delimiters to ASCII quotes"})

    bindings, parser = parse_json_or_literal(repaired)
    if excel_row == 20:
        bindings[1]["ObjectName"] = ["Food prices"]
        repairs.append({
            "type": "coreference-adapter",
            "detail": "resolve the implicit subject of reversing to Food prices from the preceding clause and reason",
        })
    if excel_row == 24:
        bindings[2]["ObjectName"] = ["sales revenue"]
        repairs.append({
            "type": "coreference-adapter",
            "detail": "resolve growth in sales to sales revenue from text/reason",
        })
    return bindings, repairs, parser


def validate_binding(binding: Any, where: str) -> list[str]:
    errors: list[str] = []
    if not isinstance(binding, dict):
        return [f"{where} must be an object"]
    missing = set(FIELDS) - set(binding)
    extras = set(binding) - set(FIELDS)
    if missing:
        errors.append(f"{where} missing fields {sorted(missing)}")
    if extras:
        errors.append(f"{where} has extra fields {sorted(extras)}")
    if errors:
        return errors
    object_name = binding["ObjectName"]
    if not (
        isinstance(object_name, list)
        and object_name
        and all(isinstance(item, str) and item.strip() for item in object_name)
    ):
        errors.append(f"{where}.ObjectName must be a non-empty string array")
    if not isinstance(binding["DataName"], str):
        errors.append(f"{where}.DataName must be a string")
    position = binding["Position"]
    if not isinstance(position, list) or not all(
        isinstance(item, dict) for item in position
    ):
        errors.append(f"{where}.Position must be an array of objects")
    trend = binding["Trend"]
    if trend is not None and not isinstance(trend, str):
        errors.append(f"{where}.Trend must be a string or null")
    num = binding["Num"]

    def is_absent(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            return value.strip().lower() in {"", "none"}
        if isinstance(value, list):
            return not value or all(is_absent(item) for item in value)
        return False

    num_absent = is_absent(num)
    if not num_absent and not (
        isinstance(num, list)
        and all(
            not isinstance(item, bool)
            and isinstance(item, (int, float))
            and float("-inf") < float(item) < float("inf")
            for item in num
        )
    ):
        errors.append(f"{where}.Num must be absent or a finite JSON number array")
    if not isinstance(binding["Text"], str):
        errors.append(f"{where}.Text must be a string")
    return errors


def prompt_prefix(field: str, examples: list[Mapping[str, Any]]) -> str:
    projected: list[dict[str, Any]] = []
    for example in examples:
        bindings = []
        for binding in example["canonicalBindings"]:
            bindings.append({
                "DataName": binding["DataName"],
                "Position": binding["Position"],
                field: binding[field],
                "resultText": binding["Text"],
            })
        projected.append({
            "excelRow": example["excelRow"],
            "pattern": example["pattern"],
            "type": example["type"],
            "sourceText": example["text"],
            "referenceBindings": bindings,
            "reason": example["reason"],
        })
    header = (
        f"Canonical annotation references for {field}. These 26 examples come from "
        "1_full_used_data.xlsx after the audited v4 schema adapter. They illustrate "
        "annotation intent, not permission to repair A or B. resultText is always "
        "included so complete-proposition semantics remain visible.\n"
    )
    return header + json.dumps(
        {"field": field, "examples": projected},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ) + "\n"


def build(config_path: Path, output_dir: Path) -> dict[str, Any]:
    config = json.loads(config_path.read_text(encoding="utf-8"))
    spec = config["judgeExamples"]
    workbook_path = workspace_path(spec)
    actual_sha = sha256_file(workbook_path)
    if actual_sha != spec["sha256"]:
        raise ProtocolError(f"judge workbook SHA-256 mismatch: {actual_sha}")
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_name = str(spec["sheet"])
    if sheet_name not in workbook.sheetnames:
        if len(workbook.sheetnames) == 1:
            sheet_name = workbook.sheetnames[0]
        else:
            raise ProtocolError(f"judge workbook sheet missing: {sheet_name}")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [str(value or "") for value in rows[0]]
    required = {"pattern", "type", "file", "data", "text", "result", "reason"}
    missing = required - set(headers)
    if missing:
        raise ProtocolError(f"judge workbook columns missing: {sorted(missing)}")
    columns = {name: headers.index(name) for name in required}
    source_rows = rows[1:]
    if len(source_rows) != int(spec["expectedRows"]):
        raise ProtocolError(
            f"judge workbook rows={len(source_rows)} expected={spec['expectedRows']}"
        )

    examples: list[dict[str, Any]] = []
    repair_count = 0
    for offset, cells in enumerate(source_rows, start=2):
        raw_result = str(cells[columns["result"]] or "")
        bindings, repairs, parser = repair_result(offset, raw_result)
        validation_errors = [
            error
            for index, binding in enumerate(bindings)
            for error in validate_binding(binding, f"row_{offset}.binding_{index}")
        ]
        if validation_errors:
            raise ProtocolError("; ".join(validation_errors))
        repair_count += int(bool(repairs))
        examples.append({
            "exampleId": f"judge_example_{offset:02d}",
            "excelRow": offset,
            "file": cells[columns["file"]],
            "pattern": str(cells[columns["pattern"]] or ""),
            "type": str(cells[columns["type"]] or ""),
            "data": str(cells[columns["data"]] or ""),
            "text": str(cells[columns["text"]] or ""),
            "rawResult": raw_result,
            "reason": str(cells[columns["reason"]] or ""),
            "canonicalBindings": bindings,
            "rawParser": parser,
            "repairs": repairs,
            "repairVersion": REPAIR_VERSION,
        })

    output_dir.mkdir(parents=True, exist_ok=True)
    canonical_path = output_dir / "canonical_examples.jsonl"
    write_jsonl(canonical_path, examples)
    prefix_files: dict[str, dict[str, Any]] = {}
    for field in SEMANTIC_FIELDS:
        rendered = prompt_prefix(field, examples)
        path = output_dir / f"judge_prompt_prefix_{field}.txt"
        path.write_text(rendered, encoding="utf-8")
        prefix_files[field] = {
            "path": str(path),
            "sha256": sha256_file(path),
            "characters": len(rendered),
            "lexicalTokenEstimate": len(re.findall(r"\w+|[^\w\s]", rendered)),
        }
    diff_path = output_dir / "repair_manifest.jsonl"
    write_jsonl(diff_path, [
        {
            "exampleId": item["exampleId"],
            "excelRow": item["excelRow"],
            "rawResultSha256": sha256_text(item["rawResult"]),
            "canonicalBindingsSha256": sha256_text(json.dumps(
                item["canonicalBindings"], ensure_ascii=False, sort_keys=True
            )),
            "repairs": item["repairs"],
        }
        for item in examples
    ])
    manifest = {
        "protocol": REPAIR_VERSION,
        "createdAt": utc_now(),
        "source": {
            "path": str(workbook_path),
            "sheet": sheet_name,
            "sha256": actual_sha,
            "rows": len(examples),
        },
        "validation": {
            "status": "passed",
            "canonicalRows": len(examples),
            "canonicalBindings": sum(len(item["canonicalBindings"]) for item in examples),
            "rowsWithRepairs": repair_count,
            "schemaFields": list(FIELDS),
        },
        "files": {
            "canonicalExamples": {
                "path": str(canonical_path),
                "sha256": sha256_file(canonical_path),
            },
            "repairManifest": {
                "path": str(diff_path),
                "sha256": sha256_file(diff_path),
            },
            "promptPrefixes": prefix_files,
        },
    }
    write_json(output_dir / "manifest.json", manifest)
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--config",
        type=Path,
        default=REPO_ROOT / "config" / "experiment6_narrative2_hybrid_v4.json",
    )
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    manifest = build(args.config.resolve(), args.output_dir.resolve())
    print(json.dumps(manifest, ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
