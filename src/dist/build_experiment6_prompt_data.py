#!/usr/bin/env python3
"""Build Experiment 6 narrative prompt CSVs and gold JSONL files.

Experiment 6 evaluates data binding.  The prompt input is built only from the
narrative ``data`` and ``text`` fields, plus prompt-mode examples when the mode
requires them.  ``Binding_Result`` is copied from ``narratives1.xlsx`` as the
gold/reference label; it is not a model prediction.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import random
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl

from binding_extraction import extract_result_items, item_dict


REPO_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = REPO_ROOT.parent
DEFAULT_SOURCE_XLSX = WORKSPACE_ROOT / "data" / "src" / "narratives" / "narratives1.xlsx"
DEFAULT_EXAMPLE_CSV = WORKSPACE_ROOT / "data" / "src" / "full_example.csv"
PROMPT_OUTPUTS = {
    "original": WORKSPACE_ROOT / "data" / "finqa_original" / "narratives1_rel_fact_instruction.csv",
    "zero-shot": WORKSPACE_ROOT / "data" / "finqa_zero_shot" / "narratives1_rel_fact_instruction.csv",
    "many-shot": WORKSPACE_ROOT / "data" / "finqa_many_shot" / "narratives1_rel_fact_instruction.csv",
    "dynamic-shot": WORKSPACE_ROOT / "data" / "finqa_dynamic_shot" / "narratives1_rel_fact_instruction.csv",
}
GOLD_OUTPUTS = {
    "original": WORKSPACE_ROOT / "data" / "finqa_original" / "narratives_gold.jsonl",
    "zero-shot": WORKSPACE_ROOT / "data" / "finqa_zero_shot" / "narratives_gold.jsonl",
    "many-shot": WORKSPACE_ROOT / "data" / "finqa_many_shot" / "narratives_gold.jsonl",
    "dynamic-shot": WORKSPACE_ROOT / "data" / "finqa_dynamic_shot" / "narratives_gold.jsonl",
}

OUTPUT_COLUMNS = [
    "Unnamed: 0",
    "Sentence",
    "input",
    "Question",
    "GT_Answer",
    "GT_Program",
    "Pre_Text",
    "Post_Text",
    "Tables",
    "Table_Text",
    "Rel_Fact",
    "Source",
    "Narrative_Data",
    "Narrative_Text",
    "Binding_Result",
    "Binding_Reason",
    "Prompt_Mode",
    "Generator_Model",
]

NEW_PROMPT_HEADER = """# Retriever + Data Binding

## Instruction
First read the task description and think as a financial analyst.
There could be multiple sentences associated with real facts for the given context and the question.

## Task
Extract relevant facts from given context and provided question. Return the relevant fact and binding result and reason.
"""

OUTPUT_FORMAT = """## Output Format
Return a valid JSON object:
{
  "RetFact":"",
  "Binding":[{"ObjectName":[],"DataName":"","Position":[{"Begin":[],"End":[]}],"Trend":"None","Num":[],"Text":""}],
  "Reason":""
}
"""

ORIGINAL_SYSTEM_PROMPT = (
    "You are a financial data binding assistant. Given chart data and narrative text, "
    "extract every data-text binding. Return only: result: "
    "[{\"ObjectName\":[],\"DataName\":\"\",\"Position\":[{\"Begin\":[],\"End\":[]}],"
    "\"Trend\":\"None\",\"Num\":[],\"Text\":\"\"}] reason: \"\""
)
QUESTION = "What data-text bindings are described in the narrative?"


def default_binding() -> dict[str, Any]:
    return {
        "ObjectName": [],
        "DataName": "",
        "Position": [{"Begin": [], "End": []}],
        "Trend": "None",
        "Num": [],
        "Text": "",
    }


def schema_result_json(ret_fact: str, reason: str = "") -> str:
    """Return a RetFact-only schema example; Binding and Reason are placeholders."""
    payload = {
        "RetFact": str(ret_fact or "").strip(),
        "Binding": [default_binding()],
        "Reason": str(reason or "").strip(),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)


@dataclass(frozen=True)
class NarrativeRow:
    index: int
    source: str
    data: str
    text: str
    result: str
    reason: str


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def stable_case_id(raw_value: Any, row_index: int) -> str:
    text = str(raw_value or "").strip()
    if not text:
        return f"row_{row_index}"
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("_") or f"row_{row_index}"


def find_column(headers: dict[str, int], name: str) -> int:
    key = normalize_header(name)
    if key not in headers:
        raise ValueError(f"Cannot find column {name!r}; available={sorted(headers)}")
    return headers[key]


def read_narratives(args: argparse.Namespace) -> list[NarrativeRow]:
    workbook = openpyxl.load_workbook(args.source_xlsx, data_only=True)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet {args.sheet!r} not found; available={workbook.sheetnames}")
    worksheet = workbook[args.sheet]
    headers = {normalize_header(worksheet.cell(1, col).value): col for col in range(1, worksheet.max_column + 1)}
    source_col = find_column(headers, "Source")
    data_col = find_column(headers, "data")
    text_col = find_column(headers, "text")
    result_col = find_column(headers, "result")
    reason_col = headers.get(normalize_header("reason"))

    rows: list[NarrativeRow] = []
    for excel_row in range(2, worksheet.max_row + 1):
        data = str(worksheet.cell(excel_row, data_col).value or "").strip()
        text = str(worksheet.cell(excel_row, text_col).value or "").strip()
        result = str(worksheet.cell(excel_row, result_col).value or "").strip()
        if not data or not text or not result:
            continue
        source = stable_case_id(worksheet.cell(excel_row, source_col).value, excel_row)
        reason = str(worksheet.cell(excel_row, reason_col).value or "").strip() if reason_col else ""
        rows.append(NarrativeRow(len(rows), source, data, text, result, reason))
        if args.limit and len(rows) >= args.limit:
            break
    if len(rows) < args.min_rows:
        raise RuntimeError(f"Only {len(rows)} usable rows found; min_rows={args.min_rows}")
    return rows


def read_examples(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def tokens(text: str) -> Counter[str]:
    return Counter(re.findall(r"[A-Za-z0-9_.%-]+", text.lower()))


def cosine(left: Counter[str], right: Counter[str]) -> float:
    if not left or not right:
        return 0.0
    dot = sum(count * right.get(term, 0) for term, count in left.items())
    left_norm = math.sqrt(sum(count * count for count in left.values()))
    right_norm = math.sqrt(sum(count * count for count in right.values()))
    return dot / (left_norm * right_norm) if left_norm and right_norm else 0.0


def example_payload(example: dict[str, str]) -> str:
    # Experiment 6 examples are RetFact-only; Binding and Reason stay as schema placeholders.
    context_parts = [
        str(example.get("Tables") or "").strip(),
        str(example.get("Sentence") or "").strip(),
    ]
    table_text = str(example.get("Table_Text") or "").strip()
    if table_text:
        context_parts.append(table_text)
    question = str(example.get("Question") or "").strip()
    context = "\n".join(part for part in context_parts if part)
    if question:
        context = f"{context}; question:{question}"
    result = schema_result_json(str(example.get("Rel_Fact") or "").strip())
    return (
        "### Example\n"
        "Example Type: RetFact-only schema example; Binding and Reason are placeholders.\n"
        f"Context: {context}\n"
        f"Result: {result}\n"
        "###"
    )


def select_examples(
    mode: str,
    row: NarrativeRow,
    examples: list[dict[str, str]],
    args: argparse.Namespace,
) -> list[dict[str, str]]:
    if mode == "zero-shot":
        return []
    if mode == "many-shot":
        rng = random.Random(args.random_seed)
        if len(examples) <= args.many_shot_count:
            return list(examples)
        return rng.sample(examples, args.many_shot_count)
    if mode == "dynamic-shot":
        query = tokens(f"{row.data} {row.text}")
        ranked = sorted(
            examples,
            key=lambda example: cosine(query, tokens(f"{example.get('Sentence', '')} {example.get('input', '')}")),
            reverse=True,
        )
        return ranked[: args.dynamic_shot_count]
    return []


def build_schema_prompt(row: NarrativeRow, examples: list[dict[str, str]]) -> str:
    context = f"## Context\n{row.data}\n{row.text}; question:{QUESTION}"
    sections = [NEW_PROMPT_HEADER, context, OUTPUT_FORMAT]
    if examples:
        sections.append("\n---\n".join(example_payload(example) for example in examples))
    return "\n\n".join(section.strip() for section in sections if section.strip()) + "\n"


def build_original_prompt(row: NarrativeRow) -> str:
    messages = [
        {"role": "system", "content": ORIGINAL_SYSTEM_PROMPT},
        {"role": "user", "content": f"data:{row.data}\ntext:[{row.text}]"},
    ]
    return json.dumps(messages, ensure_ascii=False)


def csv_row(mode: str, row: NarrativeRow, prompt: str) -> dict[str, str]:
    return {
        "Unnamed: 0": str(row.index),
        "Sentence": row.text,
        "input": prompt,
        "Question": QUESTION,
        "GT_Answer": "",
        "GT_Program": "",
        "Pre_Text": row.text,
        "Post_Text": "",
        "Tables": row.data,
        "Table_Text": "",
        "Rel_Fact": row.result,
        "Source": row.source,
        "Narrative_Data": row.data,
        "Narrative_Text": row.text,
        "Binding_Result": row.result,
        "Binding_Reason": row.reason,
        "Prompt_Mode": mode,
        "Generator_Model": "gold_from_narratives1_xlsx",
    }


def write_csv(path: Path, rows: Iterable[dict[str, str]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
            count += 1
    return count


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")
            count += 1
    return count


def gold_rows(narratives: list[NarrativeRow]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    reports: list[dict[str, Any]] = []
    for row in narratives:
        items, _records, report = extract_result_items(
            {"case_id": row.source, "result": row.result},
            fallback_case_id=row.source,
            row_number=row.index + 2,
            strict=True,
        )
        rows.append({"case_id": row.source, "items": [item_dict(item) for item in items]})
        reports.append(report)
    return rows, reports


def build(args: argparse.Namespace) -> dict[str, Any]:
    narratives = read_narratives(args)
    examples = read_examples(args.example_csv)
    gold, reports = gold_rows(narratives)
    outputs: list[dict[str, Any]] = []

    for mode, path in PROMPT_OUTPUTS.items():
        csv_rows = []
        for row in narratives:
            if mode == "original":
                prompt = build_original_prompt(row)
            else:
                prompt = build_schema_prompt(row, select_examples(mode, row, examples, args))
            csv_rows.append(csv_row(mode, row, prompt))
        csv_count = write_csv(path, csv_rows)
        gold_count = write_jsonl(GOLD_OUTPUTS[mode], gold)
        outputs.append(
            {
                "prompt_mode": mode,
                "csv": str(path),
                "gold_jsonl": str(GOLD_OUTPUTS[mode]),
                "rows": csv_count,
                "gold_rows": gold_count,
            }
        )

    report = {
        "source_xlsx": str(args.source_xlsx),
        "example_csv": str(args.example_csv),
        "rows": len(narratives),
        "many_shot_count": args.many_shot_count,
        "dynamic_shot_count": args.dynamic_shot_count,
        "outputs": outputs,
        "gold_extraction_reports": reports,
    }
    if args.report_json:
        args.report_json.parent.mkdir(parents=True, exist_ok=True)
        args.report_json.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-xlsx", type=Path, default=DEFAULT_SOURCE_XLSX)
    parser.add_argument("--sheet", default="label")
    parser.add_argument("--example-csv", type=Path, default=DEFAULT_EXAMPLE_CSV)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--min-rows", type=int, default=2)
    parser.add_argument("--many-shot-count", type=int, default=26)
    parser.add_argument("--dynamic-shot-count", type=int, default=10)
    parser.add_argument("--random-seed", type=int, default=20260612)
    parser.add_argument("--report-json", type=Path)
    return parser.parse_args()


def main() -> None:
    print(json.dumps(build(parse_args()), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
