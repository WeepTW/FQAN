#!/usr/bin/env python3
"""Resumable full-matrix runner for Experiment 6 narrative2 v2.

Gold targets are deliberately unavailable to this process. Retriever-family
models use their native RetFact formatter and a constrained GPT-5.5 converter;
all other configured models produce bindings directly.
"""

from __future__ import annotations

import argparse
import csv
import fcntl
import hashlib
import io
import json
import math
import os
import random
import re
import shutil
import subprocess
import sys
import time
import traceback
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, replace
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Iterable, Mapping, Sequence
from unicodedata import normalize as unicode_normalize

import openpyxl
from openai import OpenAI
from transformers import AutoTokenizer

SCRIPTS_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS_ROOT))

from experiment6_paths import PATHS  # noqa: E402
from experiment6_registry import (  # noqa: E402
    SourceRegistry,
    load_registry,
)
import run_experiment6_binding_generation as legacy  # noqa: E402

REPO_ROOT = PATHS.repo
WORKSPACE_ROOT = PATHS.workspace
sys.path.insert(0, str(REPO_ROOT))

from narrative.finflier_prompt import (  # noqa: E402
    build_prompt as build_finflier_prompt,
    load_prompt_asset as load_finflier_prompt_asset,
)
ACTIVE_SOURCE_REGISTRY: SourceRegistry | None = None


POSITION_INDEX_CONTRACT = """Binding coordinate contract:
Position coordinates are [row_index, data_column_index], both zero-based.
The compact CSV column named __row__ is synthetic row metadata: use its values as
row_index, but never count __row__ as a data column. data_column_index 0 is the first
original chart column after __row__. DataName must exactly and case-sensitively equal
the original chart header at data_column_index. Do not use a narrative concept as
DataName when it is not that exact chart header. When a narrative asserts a measured
series for an object or category, anchor DataName and Position to the measured series
value cell; do not substitute the axis/category label cell merely because it names
ObjectName. This remains true when the narrative omits the numeric value; keep Num empty."""


CONVERTER_SYSTEM_PROMPT = """You are a constrained RetFact-to-binding converter.
Return exactly one strict JSON object with exactly two top-level keys: result and reason.
result is an array. Every item has exactly ObjectName, DataName, Position, Trend, Num, Text.
ObjectName is a non-empty string array. DataName and Text are strings. Trend is a string.
Position is an array of objects with exactly Begin and End, each a two-integer array.
Num is an array of finite JSON numbers.
The candidate RetFact is the sole authority for whether a binding exists. Chart data and
narrative text may only ground names, coordinates, values, and exact supporting spans for
facts already present in the candidate. Never add a fact, number, trend, entity, or binding
that is absent from the candidate. If the candidate is blank, malformed, or insufficient,
return {"result":[],"reason":"candidate insufficient"}. No markdown or commentary."""
CONVERTER_SYSTEM_PROMPT += "\n" + POSITION_INDEX_CONTRACT

DIRECT_OUTPUT_CONTRACT = """You are a financial data-binding model.
Return exactly one strict JSON object with exactly two top-level keys: result and reason.
result is an array. Every item has exactly ObjectName, DataName, Position, Trend, Num, Text.
ObjectName is a non-empty string array. DataName and Text are strings. Trend is a string.
Position is an array of objects with exactly Begin and End, each a two-integer array.
Num is an array of finite JSON numbers. Use only supplied chart data and narrative text.
Return no markdown, hidden reasoning, prose, Python, or additional keys."""
DIRECT_SYSTEM_PROMPT = DIRECT_OUTPUT_CONTRACT + "\n" + POSITION_INDEX_CONTRACT

RETRIEVER_INSTRUCTION = """# Financial RetFact retrieval
Extract only the chart-backed fact or facts asserted by the supplied narrative.
Do not infer facts absent from the chart or narrative."""

DIRECT_INSTRUCTION = """# Financial data-text binding
Extract every chart-backed binding asserted by the supplied narrative.
DataName and Position must identify the exact chart value. Preserve the full proposition."""

PROMPT_MODE_LABELS = {
    "original": "original",
    "zero-shot": "zero-shot",
    "many-shot": "many-shot",
    "dynamic-shot": "dynamic-shot",
}
BUNDLE_MODE_DIRECTORIES = {
    "original": "original",
    "zero-shot": "zero",
    "many-shot": "many",
    "dynamic-shot": "dynamic",
}
MODE_SUFFIX = {"zero-shot": "z", "many-shot": "m", "dynamic-shot": "d"}
RETRIEVER_MODEL_IDS = {
    "flan": "google/flan-t5-large",
    "mistral": "mistralai/Mistral-7B-Instruct-v0.3",
    "t5gemma2": "google/t5gemma-2-1b-1b",
}
BASE_ROUTE_MODES = ("formal", "historical", "direct-diagnostic")
BARE_PERCENT_VALUE = re.compile(
    r"(?P<number>[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)%"
    r"(?=\s*[,}\]])"
)
CSV_FIELDS = [
    "Unnamed: 0", "Sentence", "input", "Question", "GT_Answer", "GT_Program",
    "Pre_Text", "Post_Text", "Tables", "Table_Text", "Rel_Fact", "Source",
    "Narrative_Data", "Narrative_Text", "Binding_Result", "Binding_Reason",
    "Prompt_Mode", "Generator_Model",
]


class ProtocolError(RuntimeError):
    """Raised when input, routing, or output violates the frozen protocol."""


@dataclass(frozen=True)
class Example:
    index: int
    retfact: str
    search_text: str


@dataclass(frozen=True)
class InputRow:
    index: int
    number: str
    source: str
    data_raw: str
    data_compact: str
    text: str
    retriever_prompts: Mapping[str, str]
    direct_prompts: Mapping[str, str]
    shot_ids: Mapping[str, tuple[int, ...]]


@dataclass(frozen=True)
class MatrixCase:
    output_id: str
    source_id: str
    prompt_mode: str
    route: str
    part: int
    official: bool

    @property
    def legacy_case(self) -> legacy.MatrixCase:
        return legacy.MatrixCase(
            experiment_id=self.output_id,
            source_id=self.source_id,
            narrative_route=f"narrative_{self.prompt_mode.replace('-', '_')}",
        )


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def gpu_execution_identity(device: str | None) -> dict[str, Any] | None:
    """Capture execution-only GPU provenance without changing the fingerprint."""

    if device is None:
        return None
    identity: dict[str, Any] = {
        "cudaVisibleDevices": str(device),
        "status": "unavailable",
    }
    if device == "cpu":
        identity.update({
            "status": "resolved",
            "deviceType": "cpu",
            "cudaVisibleDevices": "",
        })
        return identity
    try:
        completed = subprocess.run(
            [
                "nvidia-smi",
                f"--id={device}",
                "--query-gpu=name,uuid,driver_version",
                "--format=csv,noheader,nounits",
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=15,
        )
        if completed.returncode == 0 and completed.stdout.strip():
            name, uuid, driver = [
                value.strip()
                for value in completed.stdout.splitlines()[0].split(",", 2)
            ]
            identity.update({
                "status": "resolved",
                "name": name,
                "uuid": uuid,
                "driverVersion": driver,
            })
        summary = subprocess.run(
            ["nvidia-smi"],
            check=False,
            capture_output=True,
            text=True,
            timeout=15,
        )
        match = re.search(r"CUDA Version:\s*([0-9.]+)", summary.stdout)
        if match:
            identity["cudaVersion"] = match.group(1)
    except (OSError, subprocess.SubprocessError, ValueError):
        pass
    return identity


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_text(value: str) -> str:
    return sha256_bytes(value.encode("utf-8"))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ProtocolError(f"cannot load JSON {path}: {exc}") from exc


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(
        path.name + f".tmp.{os.getpid()}.{time.monotonic_ns()}"
    )
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, allow_nan=False) + "\n",
        encoding="utf-8",
    )
    os.replace(temporary, path)


def write_jsonl(path: Path, values: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(
        path.name + f".tmp.{os.getpid()}.{time.monotonic_ns()}"
    )
    with temporary.open("w", encoding="utf-8") as handle:
        for value in values:
            handle.write(json.dumps(value, ensure_ascii=False, allow_nan=False) + "\n")
    os.replace(temporary, path)


def write_text_atomic(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(
        path.name + f".tmp.{os.getpid()}.{time.monotonic_ns()}"
    )
    temporary.write_text(value, encoding="utf-8")
    os.replace(temporary, path)


def freeze_generation_snapshot(
    output_root: Path, config: Mapping[str, Any]
) -> Path:
    """Write the generation snapshot once; later workers only verify it."""

    snapshot_path = output_root / "generation_config.snapshot.json"
    if snapshot_path.is_file():
        if read_json(snapshot_path) != config:
            raise ProtocolError(
                "frozen generation config snapshot differs from current config: "
                f"{snapshot_path}"
            )
    else:
        write_json(snapshot_path, config)
    return snapshot_path


def freeze_compatibility_snapshot(
    output_root: Path, compatibility: Mapping[str, Any]
) -> Path:
    snapshot_path = output_root / "compatibility_fingerprint.json"
    if snapshot_path.is_file():
        existing = read_json(snapshot_path)
        if existing.get("sha256") != compatibility.get("sha256"):
            raise ProtocolError(
                "output root compatibility fingerprint differs; use a fresh root"
            )
    else:
        write_json(snapshot_path, compatibility)
    return snapshot_path


def workspace_path(raw: str | Mapping[str, Any]) -> Path:
    if isinstance(raw, Mapping):
        if "root" in raw:
            return PATHS.resolve_locator(raw)
        raw = str(raw["path"])
    path = Path(raw)
    return path if path.is_absolute() else WORKSPACE_ROOT / path


def model_reference(raw: object) -> str:
    """Resolve a local logical locator while preserving Hugging Face model IDs."""
    if isinstance(raw, Mapping):
        return str(PATHS.resolve_locator(raw))
    return str(raw)


def load_config(path: Path) -> dict[str, Any]:
    global ACTIVE_SOURCE_REGISTRY
    config = read_json(path)
    supported_protocols = {
        "experiment6-narrative2-full-v2",
        "experiment6-narrative2-generation-v4-38case",
        "experiment6-narrative2-generation-v5-corrected12",
        "experiment6-narrative2-generation-v6-finflier-no-adaptor",
        "experiment6-narrative2-generation-v6-flan-finflier-long-context",
    }
    if not isinstance(config, dict) or config.get("protocol") not in supported_protocols:
        raise ProtocolError(
            "generation config protocol must be one of "
            + ", ".join(sorted(supported_protocols))
        )
    registry_spec = config.get("sourceRegistry")
    if registry_spec is None:
        ACTIVE_SOURCE_REGISTRY = load_registry()
    elif not isinstance(registry_spec, Mapping):
        raise ProtocolError("sourceRegistry must be a logical-root locator")
    else:
        registry_path = PATHS.resolve_locator(registry_spec)
        ACTIVE_SOURCE_REGISTRY = load_registry(registry_path)
        expected_registry_sha = registry_spec.get("sha256")
        if (
            expected_registry_sha is not None
            and ACTIVE_SOURCE_REGISTRY.file_sha256 != expected_registry_sha
        ):
            raise ProtocolError(
                "source registry SHA-256 mismatch: "
                f"{ACTIVE_SOURCE_REGISTRY.file_sha256} != {expected_registry_sha}"
            )
    runtime_routes = config.get("runtimeRoutes", {})
    if not isinstance(runtime_routes, Mapping):
        raise ProtocolError("runtimeRoutes must be an object")
    for route_name, route in runtime_routes.items():
        if not isinstance(route, Mapping) or route.get("responseFormat") != "json_schema":
            continue
        schema_path = Path(str(route.get("responseSchemaPath") or ""))
        if not schema_path.is_absolute():
            schema_path = REPO_ROOT / schema_path
        if not schema_path.is_file():
            raise ProtocolError(
                f"response schema missing for {route_name}: {schema_path}"
            )
        actual_sha = sha256_file(schema_path)
        if actual_sha != route.get("responseSchemaSha256"):
            raise ProtocolError(
                f"response schema SHA-256 mismatch for {route_name}: {actual_sha}"
            )
        schema_document = read_json(schema_path)
        if (
            not isinstance(schema_document, dict)
            or not isinstance(schema_document.get("name"), str)
            or not isinstance(schema_document.get("schema"), dict)
        ):
            raise ProtocolError(
                f"response schema has invalid envelope for {route_name}"
            )
    return config


def verify_bundle(bundle: Path) -> dict[str, Any]:
    manifest_path = bundle / "bundle_manifest.json"
    manifest = read_json(manifest_path)
    entries = manifest.get("files") if isinstance(manifest, dict) else None
    if not isinstance(entries, list):
        raise ProtocolError("generation bundle manifest has no files[]")
    checked: list[dict[str, Any]] = []
    for entry in entries:
        relative = Path(str(entry.get("path") or ""))
        if relative.is_absolute() or ".." in relative.parts:
            raise ProtocolError(f"unsafe bundle path: {relative}")
        target = bundle / relative
        if not target.is_file():
            raise ProtocolError(f"bundle file missing: {relative}")
        actual_sha = sha256_file(target)
        actual_bytes = target.stat().st_size
        if actual_sha != entry.get("sha256") or actual_bytes != entry.get("bytes"):
            raise ProtocolError(f"bundle manifest mismatch: {relative}")
        checked.append({"path": relative.as_posix(), "sha256": actual_sha, "bytes": actual_bytes})
    return {
        "status": "passed",
        "manifest": str(manifest_path),
        "manifestSha256": sha256_file(manifest_path),
        "filesChecked": len(checked),
        "files": checked,
    }


def materialize_prompt_bundles(
    output_root: Path,
    rows: Sequence[InputRow],
    config: Mapping[str, Any],
) -> tuple[list[InputRow], dict[str, Any]]:
    """Freeze route-aware prompt bytes and make inference read those bytes."""

    bundle_root = output_root / "input_bundles"
    bundle_root.mkdir(parents=True, exist_ok=True)
    reports: dict[str, Any] = {}

    for mode, directory_name in BUNDLE_MODE_DIRECTORIES.items():
        mode_root = bundle_root / directory_name
        expected_text: dict[Path, str] = {}
        model_inputs: list[dict[str, Any]] = []
        for row in rows:
            retriever_relative = (
                Path("messages") / "retriever" / f"{row.source}.txt"
            )
            direct_relative = Path("messages") / "direct" / f"{row.source}.txt"
            retriever_prompt = row.retriever_prompts[mode]
            direct_prompt = row.direct_prompts[mode]
            expected_text[retriever_relative] = retriever_prompt
            expected_text[direct_relative] = direct_prompt
            model_inputs.append(
                {
                    "index": row.index,
                    "number": row.number,
                    "source": row.source,
                    "promptMode": mode,
                    "inputType": str(config.get("inputType") or "type"),
                    "shotIds": list(row.shot_ids[mode]),
                    "dataSha256": sha256_text(row.data_raw),
                    "textSha256": sha256_text(row.text),
                    "retrieverMessagePath": retriever_relative.as_posix(),
                    "retrieverPromptSha256": sha256_text(retriever_prompt),
                    "directMessagePath": direct_relative.as_posix(),
                    "directPromptSha256": sha256_text(direct_prompt),
                }
            )

        model_inputs_text = "".join(
            json.dumps(item, ensure_ascii=False, sort_keys=True, allow_nan=False)
            + "\n"
            for item in model_inputs
        )
        expected_text[Path("model_inputs.jsonl")] = model_inputs_text
        for relative, content in expected_text.items():
            target = mode_root / relative
            if target.is_file():
                if target.read_text(encoding="utf-8") != content:
                    raise ProtocolError(
                        "frozen prompt bundle differs from regenerated bytes: "
                        f"{target}"
                    )
            else:
                write_text_atomic(target, content)

        file_entries = []
        for relative in sorted(expected_text, key=lambda item: item.as_posix()):
            target = mode_root / relative
            file_entries.append(
                {
                    "path": relative.as_posix(),
                    "bytes": target.stat().st_size,
                    "sha256": sha256_file(target),
                }
            )
        manifest = {
            "schemaVersion": 1,
            "protocol": "experiment6-narrative2-prompt-bundle-v1",
            "promptMode": mode,
            "inputType": str(config.get("inputType") or "type"),
            "rowCount": len(rows),
            "routes": ["retriever", "direct"],
            "sourceWorkbookSha256": config["inputWorkbook"]["sha256"],
            "promptPolicyVersion": config["promptBuilder"]["promptPolicyVersion"],
            "files": file_entries,
        }
        manifest_path = mode_root / "bundle_manifest.json"
        if manifest_path.is_file():
            if read_json(manifest_path) != manifest:
                raise ProtocolError(
                    "frozen prompt manifest differs from regenerated manifest: "
                    f"{manifest_path}"
                )
        else:
            write_json(manifest_path, manifest)
        reports[mode] = {
            "directory": str(mode_root),
            "manifest": str(manifest_path),
            "manifestSha256": sha256_file(manifest_path),
            "rows": len(model_inputs),
            "inputType": str(config.get("inputType") or "type"),
            "messageFiles": len(rows) * 2,
            "routes": ["retriever", "direct"],
        }

    bound_rows: list[InputRow] = []
    for row in rows:
        retriever_prompts: dict[str, str] = {}
        direct_prompts: dict[str, str] = {}
        for mode, directory_name in BUNDLE_MODE_DIRECTORIES.items():
            mode_root = bundle_root / directory_name
            retriever_prompts[mode] = (
                mode_root / "messages" / "retriever" / f"{row.source}.txt"
            ).read_text(encoding="utf-8")
            direct_prompts[mode] = (
                mode_root / "messages" / "direct" / f"{row.source}.txt"
            ).read_text(encoding="utf-8")
        bound_rows.append(
            replace(
                row,
                retriever_prompts=retriever_prompts,
                direct_prompts=direct_prompts,
            )
        )
    return bound_rows, {
        "protocol": "experiment6-narrative2-prompt-bundle-v1",
        "root": str(bundle_root),
        "modes": reports,
        "rowCount": len(bound_rows),
        "inputType": str(config.get("inputType") or "type"),
        "generationReadsFrozenMessages": True,
    }


def stable_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def repair_bare_percentage_json_values(text: str) -> tuple[str, int]:
    """Quote numeric percentage values outside JSON strings, preserving their lexeme."""

    output: list[str] = []
    index = 0
    in_string = False
    escaped = False
    repairs = 0
    while index < len(text):
        character = text[index]
        if in_string:
            output.append(character)
            if escaped:
                escaped = False
            elif character == "\\":
                escaped = True
            elif character == '"':
                in_string = False
            index += 1
            continue
        if character == '"':
            in_string = True
            output.append(character)
            index += 1
            continue
        output.append(character)
        index += 1
        if character != ":":
            continue
        whitespace_start = index
        while index < len(text) and text[index].isspace():
            index += 1
        output.append(text[whitespace_start:index])
        match = BARE_PERCENT_VALUE.match(text, index)
        if match is None:
            continue
        output.append(json.dumps(match.group("number") + "%", ensure_ascii=False))
        index = match.end()
        repairs += 1
    return "".join(output), repairs


def compact_chart_data_with_audit(raw: str) -> tuple[str, dict[str, Any]]:
    text = str(raw or "").strip()
    audit: dict[str, Any] = {
        "inputStrictJson": True,
        "repairRule": None,
        "repairCount": 0,
    }
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as original_error:
        repaired, repair_count = repair_bare_percentage_json_values(text)
        if repair_count == 0:
            raise ProtocolError(
                f"chart data is invalid JSON and has no approved lossless repair: {original_error}"
            ) from original_error
        try:
            payload = json.loads(repaired)
        except json.JSONDecodeError as repaired_error:
            raise ProtocolError(
                "chart data remains invalid after approved bare-percentage repair: "
                f"{repaired_error}"
            ) from repaired_error
        audit = {
            "inputStrictJson": False,
            "repairRule": "quote-bare-percentage-json-values-v1",
            "repairCount": repair_count,
            "repairedJsonSha256": sha256_text(repaired),
        }

    if isinstance(payload, dict) and payload and all(isinstance(value, list) for value in payload.values()):
        lengths = {len(value) for value in payload.values()}
        if len(lengths) == 1:
            columns = list(payload)
            stream = io.StringIO()
            writer = csv.writer(stream, lineterminator="\n")
            writer.writerow(["__row__", *columns])
            for row_index in range(next(iter(lengths))):
                writer.writerow([
                    row_index,
                    *[
                        json.dumps(payload[column][row_index], ensure_ascii=False, separators=(",", ":"))
                        for column in columns
                    ],
                ])
            return stream.getvalue().rstrip(), audit
    if isinstance(payload, list) and payload and all(isinstance(item, dict) for item in payload):
        columns: list[str] = []
        for item in payload:
            for key in item:
                if key not in columns:
                    columns.append(key)
        stream = io.StringIO()
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(["__row__", *columns])
        for row_index, item in enumerate(payload):
            writer.writerow([
                row_index,
                *[
                    json.dumps(item.get(column), ensure_ascii=False, separators=(",", ":"))
                    for column in columns
                ],
            ])
        return stream.getvalue().rstrip(), audit
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":")), audit


def compact_chart_data(raw: str) -> str:
    return compact_chart_data_with_audit(raw)[0]


def tokens(text: str) -> Counter[str]:
    return Counter(re.findall(r"[A-Za-z0-9_.%+-]+", unicode_normalize("NFKC", text).lower()))


def cosine(left: Counter[str], right: Counter[str]) -> float:
    if not left or not right:
        return 0.0
    dot = sum(count * right.get(term, 0) for term, count in left.items())
    left_norm = math.sqrt(sum(count * count for count in left.values()))
    right_norm = math.sqrt(sum(count * count for count in right.values()))
    return dot / (left_norm * right_norm) if left_norm and right_norm else 0.0


def load_examples(path: Path) -> list[Example]:
    if not path.is_file():
        raise ProtocolError(f"few-shot source missing: {path}")
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    examples: list[Example] = []
    for index, row in enumerate(rows):
        retfact = str(row.get("Rel_Fact") or "").strip()
        if not retfact:
            continue
        search_text = " ".join(
            str(row.get(key) or "") for key in ("Sentence", "Question", "input", "Tables", "Table_Text")
        )
        examples.append(Example(index=index, retfact=retfact, search_text=search_text))
    if len(examples) < 26:
        raise ProtocolError(f"few-shot source has only {len(examples)} usable RetFacts")
    return examples


def example_block(selected: Sequence[Example]) -> str:
    if not selected:
        return ""
    lines = ["## RetFact examples"]
    for ordinal, example in enumerate(selected, start=1):
        lines.append(f"[EXAMPLE {ordinal:02d}] {example.retfact}")
    return "\n".join(lines)


def build_prompt(
    *,
    row_source: str,
    data_compact: str,
    narrative: str,
    mode: str,
    selected: Sequence[Example],
    direct: bool,
) -> str:
    instruction = DIRECT_INSTRUCTION if direct else RETRIEVER_INSTRUCTION
    if direct:
        output_contract = (
            'Return exactly {"result":[{"ObjectName":["..."],"DataName":"...",'
            '"Position":[{"Begin":[0,0],"End":[0,0]}],"Trend":"None","Num":[],"Text":"..."}],'
            '"reason":"..."}'
        )
    else:
        output_contract = "Return the candidate RetFact in the model family's canonical structured output."
    sections = [
        instruction,
        f"Prompt mode: {mode}",
        f"Source: {row_source}",
        "## Binding coordinate contract\n" + POSITION_INDEX_CONTRACT,
        "## Chart data (lossless compact form)\n" + data_compact,
        "## Narrative\n" + narrative,
    ]
    examples_text = example_block(selected)
    if examples_text:
        sections.append(examples_text)
    sections.append("## Output contract\n" + output_contract)
    return "\n\n".join(section.strip() for section in sections if section.strip()) + "\n"


def read_input_rows(
    config: Mapping[str, Any],
    limit: int,
    row_source: str | None = None,
) -> tuple[list[InputRow], dict[str, Any]]:
    workbook_spec = config["inputWorkbook"]
    workbook_path = workspace_path(workbook_spec)
    actual_sha = sha256_file(workbook_path)
    if actual_sha != workbook_spec["sha256"]:
        raise ProtocolError(f"workbook SHA-256 mismatch: {actual_sha}")
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_name = str(workbook_spec["sheet"])
    if sheet_name not in workbook.sheetnames:
        raise ProtocolError(f"workbook sheet missing: {sheet_name}")
    worksheet = workbook[sheet_name]
    header_values = [stable_scalar(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    required = {"Number", "Source", "data", "text"}
    missing = required - set(header_values)
    if missing:
        raise ProtocolError(f"workbook columns missing: {sorted(missing)}")
    column = {name: header_values.index(name) for name in required}

    examples = load_examples(workspace_path(config["exampleCsv"]))
    builder = config["promptBuilder"]
    many_count = int(builder["manyShotCount"])
    dynamic_count = int(builder["dynamicShotCount"])
    fixed_rng = random.Random(int(builder["dynamicSeed"]))
    many_examples = fixed_rng.sample(examples, many_count)
    input_type = str(config.get("inputType") or "type")
    finflier_spec = config.get("finflierPrompt")
    finflier_asset = None
    if finflier_spec is not None:
        if input_type != "FinFlier" or not isinstance(finflier_spec, Mapping):
            raise ProtocolError(
                "finflierPrompt requires inputType=FinFlier and an object spec"
            )
        asset_spec = finflier_spec.get("asset")
        if not isinstance(asset_spec, Mapping):
            raise ProtocolError("finflierPrompt.asset must be a logical locator")
        finflier_asset = load_finflier_prompt_asset(
            workspace_path(asset_spec), str(asset_spec.get("sha256") or "")
        )
        if finflier_asset.policy_version != builder["promptPolicyVersion"]:
            raise ProtocolError("FinFlier prompt policy version mismatch")
    elif input_type != "type":
        raise ProtocolError(f"unsupported inputType without prompt asset: {input_type}")

    source_records: list[tuple[str, str, str, str]] = []
    for cells in worksheet.iter_rows(min_row=2, values_only=True):
        number = stable_scalar(cells[column["Number"]])
        source = stable_scalar(cells[column["Source"]])
        data = stable_scalar(cells[column["data"]])
        narrative = stable_scalar(cells[column["text"]])
        if not source or not data or not narrative:
            raise ProtocolError(f"blank required workbook field at Excel row {len(source_records) + 2}")
        source_records.append((number, source, data, narrative))
    full_count = len(source_records)
    if full_count != int(config["expectedRows"]):
        raise ProtocolError(f"workbook rows={full_count} expected={config['expectedRows']}")
    if len({record[1] for record in source_records}) != full_count:
        raise ProtocolError("workbook Source values are not unique")
    if row_source is not None:
        selected = [record for record in source_records if record[1] == row_source]
        if len(selected) != 1:
            raise ProtocolError(
                f"--row-source must select exactly one Source, got {len(selected)}: {row_source}"
            )
        source_records = selected
    if limit > 0:
        source_records = source_records[:limit]

    rows: list[InputRow] = []
    prompt_audit: list[dict[str, Any]] = []
    for index, (number, source, data, narrative) in enumerate(source_records):
        try:
            compact_data, chart_normalization = compact_chart_data_with_audit(data)
        except ProtocolError as error:
            raise ProtocolError(f"{source} chart data normalization failed: {error}") from error
        query = tokens(compact_data + " " + narrative)
        dynamic_examples = sorted(
            examples,
            key=lambda example: (-cosine(query, tokens(example.search_text)), example.index),
        )[:dynamic_count]
        selected_by_mode: dict[str, Sequence[Example]] = {
            "original": (),
            "zero-shot": (),
            "many-shot": many_examples,
            "dynamic-shot": dynamic_examples,
        }
        retriever_prompts: dict[str, str] = {}
        direct_prompts: dict[str, str] = {}
        shot_ids: dict[str, tuple[int, ...]] = {}
        for mode, selected in selected_by_mode.items():
            retriever_prompts[mode] = build_prompt(
                row_source=source,
                data_compact=compact_data,
                narrative=narrative,
                mode=mode,
                selected=selected,
                direct=False,
            )
            direct_prompts[mode] = build_prompt(
                row_source=source,
                data_compact=compact_data,
                narrative=narrative,
                mode=mode,
                selected=selected,
                direct=True,
            )
            shot_ids[mode] = tuple(example.index for example in selected)
            expected_shots = many_count if mode == "many-shot" else dynamic_count if mode == "dynamic-shot" else 0
            if retriever_prompts[mode].count("[EXAMPLE ") != expected_shots:
                raise ProtocolError(f"{source} {mode} shot count mismatch")
            forbidden = ('"targetBindings"', '"Binding_Result"', '"gold_targets"')
            if any(marker in retriever_prompts[mode] or marker in direct_prompts[mode] for marker in forbidden):
                raise ProtocolError(f"{source} {mode} prompt contains a gold marker")
        finflier_prompt_audit = None
        if finflier_asset is not None:
            direct_prompts["original"], finflier_prompt_audit = build_finflier_prompt(
                finflier_asset,
                chart_data=data,
                narrative=narrative,
                coordinate_contract=POSITION_INDEX_CONTRACT,
                output_contract=DIRECT_OUTPUT_CONTRACT,
            )
            forbidden = ('"targetBindings"', '"Binding_Result"', '"gold_targets"')
            if any(marker in direct_prompts["original"] for marker in forbidden):
                raise ProtocolError(f"{source} FinFlier prompt contains a gold marker")
            if not finflier_prompt_audit["finflierPromptApplied"]:
                raise ProtocolError(f"{source} FinFlier prompt was not applied")
        rows.append(InputRow(
            index=index,
            number=number,
            source=source,
            data_raw=data,
            data_compact=compact_data,
            text=narrative,
            retriever_prompts=retriever_prompts,
            direct_prompts=direct_prompts,
            shot_ids=shot_ids,
        ))
        prompt_audit.append({
            "index": index,
            "source": source,
            "dataRawSha256": sha256_text(data),
            "dataCompactSha256": sha256_text(compact_data),
            "chartNormalization": chart_normalization,
            "textSha256": sha256_text(narrative),
            "retrieverPromptSha256": {mode: sha256_text(value) for mode, value in retriever_prompts.items()},
            "directPromptSha256": {mode: sha256_text(value) for mode, value in direct_prompts.items()},
            "shotIds": {mode: list(value) for mode, value in shot_ids.items()},
            "inputType": input_type,
            "finflierPrompt": finflier_prompt_audit,
        })
    example_path = workspace_path(config["exampleCsv"])
    return rows, {
        "status": "passed",
        "path": str(workbook_path),
        "sha256": actual_sha,
        "sheet": sheet_name,
        "rows": len(rows),
        "fullWorkbookRows": full_count,
        "columns": header_values,
        "exampleCsv": str(example_path),
        "exampleCsvSha256": sha256_file(example_path),
        "usableExamples": len(examples),
        "inputType": input_type,
        "finflierPrompt": (
            {
                "asset": dict(finflier_spec["asset"]),
                "policyVersion": finflier_asset.policy_version,
                "generalExampleCount": finflier_asset.general_example_count,
            }
            if finflier_asset is not None else None
        ),
        "chartNormalization": {
            "policyVersion": "strict-json-plus-quoted-bare-percentage-v1",
            "repairRows": sum(
                item["chartNormalization"]["repairCount"] > 0
                for item in prompt_audit
            ),
            "repairedSources": [
                item["source"] for item in prompt_audit
                if item["chartNormalization"]["repairCount"] > 0
            ],
        },
        "promptAudit": prompt_audit,
    }


def expand_matrix(config: Mapping[str, Any]) -> list[MatrixCase]:
    cases: list[MatrixCase] = []
    for part_spec in config["parts"]:
        part = int(part_spec["part"])
        route = str(part_spec["route"])
        if "models" in part_spec:
            for model in part_spec["models"]:
                for mode in part_spec["promptModes"]:
                    suffix = MODE_SUFFIX[str(mode)]
                    cases.append(MatrixCase(
                        output_id=f"{model['outputStem']}_{suffix}",
                        source_id=str(model["sourceId"]),
                        prompt_mode=str(mode),
                        route=route,
                        part=part,
                        official=True,
                    ))
        else:
            for item in part_spec["cases"]:
                cases.append(MatrixCase(
                    output_id=str(item["outputId"]),
                    source_id=str(item["sourceId"]),
                    prompt_mode=str(item.get("promptMode") or part_spec.get("promptMode")),
                    route=str(item.get("route") or route),
                    part=part,
                    official=True,
                ))
    controls = [
        MatrixCase(
            output_id=str(item["outputId"]),
            source_id=str(item["sourceId"]),
            prompt_mode=str(item["promptMode"]),
            route=str(item["route"]),
            part=0,
            official=False,
        )
        for item in config["controls"]
    ]
    if len(cases) != int(config["expectedOfficialCases"]):
        raise ProtocolError(f"official matrix has {len(cases)} cases, expected {config['expectedOfficialCases']}")
    if len(controls) != int(config["expectedDiagnosticCases"]):
        raise ProtocolError(f"control matrix has {len(controls)} cases, expected {config['expectedDiagnosticCases']}")
    all_cases = cases + controls
    if len({case.output_id for case in all_cases}) != len(all_cases):
        raise ProtocolError("matrix outputId values are not unique")
    if any(case.prompt_mode not in PROMPT_MODE_LABELS for case in all_cases):
        raise ProtocolError("matrix has an invalid prompt mode")
    counts = Counter(case.part for case in cases)
    raw_expected_counts = config.get(
        "expectedPartCounts", {"1": 9, "2": 24, "3": 17, "4": 4}
    )
    expected_counts = Counter(
        {int(part): int(count) for part, count in raw_expected_counts.items()}
    )
    if counts != expected_counts:
        raise ProtocolError(
            f"part counts mismatch: actual={dict(counts)} expected={dict(expected_counts)}"
        )
    return all_cases


def source_record(source_id: str) -> Mapping[str, Any] | None:
    if ACTIVE_SOURCE_REGISTRY is None:
        return None
    try:
        return ACTIVE_SOURCE_REGISTRY.source(source_id)
    except Exception as exc:
        raise ProtocolError(f"unknown registered source_id: {source_id}") from exc


def family_for_source(source_id: str) -> str | None:
    source = source_record(source_id)
    if source is None:
        return legacy.family_from_source_id(source_id)
    if source.get("kind") not in {"base", "adapter"}:
        return None
    family = source.get("family")
    return str(family) if family not in (None, "none") else None


def source_kind(source_id: str) -> str | None:
    source = source_record(source_id)
    return (
        str(source.get("kind"))
        if source is not None and source.get("kind") is not None
        else None
    )


def validate_case_route(case: MatrixCase) -> None:
    if ACTIVE_SOURCE_REGISTRY is None:
        return
    try:
        ACTIVE_SOURCE_REGISTRY.resolve_source(case.source_id, case.route)
    except Exception as exc:
        raise ProtocolError(str(exc)) from exc


def uses_direct_diagnostic_route(
    case: MatrixCase, base_route_mode: str
) -> bool:
    if base_route_mode not in BASE_ROUTE_MODES:
        raise ProtocolError(f"unsupported base route mode: {base_route_mode}")
    if base_route_mode != "direct-diagnostic":
        return False
    family = family_for_source(case.source_id)
    if family is None or source_kind(case.source_id) != "base":
        raise ProtocolError(
            "direct-diagnostic is restricted to native base retriever source IDs; "
            f"got {case.output_id}/{case.source_id}"
        )
    return True


def effective_route(case: MatrixCase, base_route_mode: str) -> str:
    if base_route_mode not in BASE_ROUTE_MODES:
        raise ProtocolError(f"unsupported base route mode: {base_route_mode}")
    if base_route_mode == "formal":
        validate_case_route(case)
        return case.route
    if uses_direct_diagnostic_route(case, base_route_mode):
        return "direct-diagnostic-native"
    return (
        "retriever-converter"
        if family_for_source(case.source_id) is not None
        else case.route
    )


def token_preflight(
    rows: Sequence[InputRow],
    cases: Sequence[MatrixCase],
    config: Mapping[str, Any],
    base_route_mode: str = "historical",
) -> dict[str, Any]:
    retriever = config["retriever"]
    maximum = int(retriever["maxInputTokens"])
    execution_routes = {
        case.output_id: effective_route(case, base_route_mode) for case in cases
    }
    retriever_cases = [
        case for case in cases
        if execution_routes[case.output_id]
        in {"adapter-converter", "retriever-converter"}
    ]
    native_direct_cases = [
        case for case in cases
        if execution_routes[case.output_id] == "direct-diagnostic-native"
        or (
            execution_routes[case.output_id] == "direct-binding"
            and source_kind(case.source_id) == "base"
        )
    ]
    generic_direct_cases = [
        case for case in cases
        if execution_routes[case.output_id] == "direct-binding"
        and source_kind(case.source_id) != "base"
    ]
    families = sorted({
        family_for_source(case.source_id)
        for case in retriever_cases
        if family_for_source(case.source_id) is not None
    })
    reports: dict[str, Any] = {}
    for family in families:
        assert family is not None
        model_name = str(retriever["tokenizers"][family])
        tokenizer = AutoTokenizer.from_pretrained(model_name, local_files_only=True, trust_remote_code=True)
        lengths: list[dict[str, Any]] = []
        for case in retriever_cases:
            if family_for_source(case.source_id) != family:
                continue
            for row in rows:
                prompt = row.retriever_prompts[case.prompt_mode]
                if family == "mistral":
                    prompt += "##Label Descriptions:"
                token_count = len(tokenizer(prompt, add_special_tokens=True, truncation=False)["input_ids"])
                lengths.append({
                    "outputId": case.output_id,
                    "source": row.source,
                    "promptMode": case.prompt_mode,
                    "tokens": token_count,
                })
        over = [item for item in lengths if item["tokens"] > maximum]
        if over:
            raise ProtocolError(f"{family} has {len(over)} prompts over {maximum} tokens; first={over[0]}")
        reports[family] = {
            "tokenizer": model_name,
            "maxAllowed": maximum,
            "maxObserved": max(item["tokens"] for item in lengths),
            "minObserved": min(item["tokens"] for item in lengths),
            "measurements": len(lengths),
            "truncationAllowed": False,
        }
    direct = config["directBinding"]
    direct_maximum = int(direct["maxInputTokens"])
    context_window = int(direct["localContextTokens"])
    direct_reports: dict[str, Any] = {}
    selected_source_ids = {case.source_id for case in generic_direct_cases}
    for source_id, model_name in direct.get("tokenizers", {}).items():
        if source_id not in selected_source_ids:
            continue
        tokenizer = AutoTokenizer.from_pretrained(
            model_reference(model_name), local_files_only=True, trust_remote_code=True,
        )
        lengths: list[dict[str, Any]] = []
        for case in generic_direct_cases:
            if case.source_id != source_id:
                continue
            for row in rows:
                prompt = row.direct_prompts[case.prompt_mode]
                token_count = len(tokenizer(
                    prompt, add_special_tokens=True, truncation=False,
                )["input_ids"])
                lengths.append({
                    "outputId": case.output_id,
                    "source": row.source,
                    "promptMode": case.prompt_mode,
                    "tokens": token_count,
                })
        over_input = [item for item in lengths if item["tokens"] > direct_maximum]
        over_context = [
            item for item in lengths
            if item["tokens"] + int(direct["maxNewTokens"]) > context_window
        ]
        if over_input:
            raise ProtocolError(
                f"{source_id} has {len(over_input)} direct prompts over {direct_maximum} tokens; "
                f"first={over_input[0]}"
            )
        if over_context:
            raise ProtocolError(
                f"{source_id} has {len(over_context)} prompt+completion budgets over "
                f"{context_window} tokens; first={over_context[0]}"
            )
        direct_reports[source_id] = {
            "tokenizer": model_reference(model_name),
            "tokenizerRole": str(
                direct.get("tokenizerRoles", {}).get(source_id) or "model-native"
            ),
            "maxInputAllowed": direct_maximum,
            "contextWindow": context_window,
            "maxNewTokens": int(direct["maxNewTokens"]),
            "maxObserved": max(item["tokens"] for item in lengths),
            "minObserved": min(item["tokens"] for item in lengths),
            "maxPromptPlusCompletion": (
                max(item["tokens"] for item in lengths) + int(direct["maxNewTokens"])
            ),
            "measurements": len(lengths),
            "truncationAllowed": False,
        }
    for case in native_direct_cases:
        family = family_for_source(case.source_id)
        assert family is not None
        model_name = str(retriever["tokenizers"][family])
        family_direct_maximum = int(
            direct.get("familyMaxInputTokens", {}).get(family, direct_maximum)
        )
        family_context_window = int(
            direct.get("familyContextTokens", {}).get(family, context_window)
        )
        tokenizer = AutoTokenizer.from_pretrained(
            model_name, local_files_only=True, trust_remote_code=True,
        )
        lengths: list[dict[str, Any]] = []
        for row in rows:
            prompt = row.direct_prompts[case.prompt_mode]
            token_count = len(tokenizer(
                prompt, add_special_tokens=True, truncation=False,
            )["input_ids"])
            lengths.append({
                "outputId": case.output_id,
                "source": row.source,
                "promptMode": case.prompt_mode,
                "tokens": token_count,
            })
        over_input = [
            item for item in lengths if item["tokens"] > family_direct_maximum
        ]
        over_context = [
            item for item in lengths
            if item["tokens"] + int(direct["maxNewTokens"]) > family_context_window
        ]
        if over_input:
            raise ProtocolError(
                f"{case.output_id} has {len(over_input)} native direct prompts over "
                f"{family_direct_maximum} tokens; first={over_input[0]}"
            )
        if over_context:
            raise ProtocolError(
                f"{case.output_id} has {len(over_context)} native direct prompt+completion "
                f"budgets over {family_context_window} tokens; first={over_context[0]}"
            )
        direct_reports[case.output_id] = {
            "sourceId": case.source_id,
            "family": family,
            "route": execution_routes[case.output_id],
            "tokenizer": model_name,
            "tokenizerRole": "model-native",
            "maxInputAllowed": family_direct_maximum,
            "contextWindow": family_context_window,
            "maxNewTokens": int(direct["maxNewTokens"]),
            "maxObserved": max(item["tokens"] for item in lengths),
            "minObserved": min(item["tokens"] for item in lengths),
            "maxPromptPlusCompletion": (
                max(item["tokens"] for item in lengths)
                + int(direct["maxNewTokens"])
            ),
            "measurements": len(lengths),
            "truncationAllowed": False,
            "structuredOutput": "off",
            "adapter": None,
            "converter": None,
            "historicalMistralLabelSuffix": (
                False if family == "mistral" else None
            ),
        }
    return {
        "status": "passed",
        "baseRouteMode": base_route_mode,
        "families": reports,
        "directBinding": direct_reports,
    }


def csv_record(row: InputRow, prompt: str, mode: str) -> dict[str, str]:
    return {
        "Unnamed: 0": str(row.index),
        "Sentence": row.text,
        "input": prompt,
        "Question": "What chart-backed fact is asserted by the narrative?",
        "GT_Answer": "",
        "GT_Program": "",
        "Pre_Text": row.text,
        "Post_Text": "",
        "Tables": row.data_compact,
        "Table_Text": "",
        "Rel_Fact": "__BLINDED__",
        "Source": row.source,
        "Narrative_Data": row.data_compact,
        "Narrative_Text": row.text,
        "Binding_Result": "",
        "Binding_Reason": "",
        "Prompt_Mode": mode,
        "Generator_Model": "blind_inference",
    }


def write_prompt_csv(path: Path, rows: Sequence[InputRow], mode: str, *, direct: bool) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for row in rows:
            prompt = row.direct_prompts[mode] if direct else row.retriever_prompts[mode]
            writer.writerow(csv_record(row, prompt, mode))


def retriever_namespace(
    run_dir: Path,
    config: Mapping[str, Any],
    limit: int,
    family: str | None = None,
) -> SimpleNamespace:
    retriever = config["retriever"]
    family_batches = retriever.get("familyBatchSizes", {})
    family_devices = retriever.get("familyCudaVisibleDevices", {})
    cuda_visible_devices = family_devices.get(
        family, retriever["cudaVisibleDevices"]
    )
    return SimpleNamespace(
        pred_dir=run_dir,
        limit=limit,
        batch_size=int(family_batches.get(family, retriever["batchSize"])),
        max_tokens=int(retriever["maxNewTokens"]),
        max_input_tokens=int(retriever["maxInputTokens"]),
        structured_output=str(retriever["structuredOutput"]),
        case_timeout_seconds=int(retriever["caseTimeoutSeconds"]),
        cuda_visible_devices=str(cuda_visible_devices),
        t5gemma_cache_safe_input_tokens=int(
            retriever.get("generationCache", {})
            .get("t5gemma2", {})
            .get("disableAboveInputTokens", 0)
        ),
        attention_query_chunk_size=int(
            retriever.get("attentionQueryChunkSize", 0)
        ),
    )


def direct_namespace(run_dir: Path, config: Mapping[str, Any], limit: int) -> SimpleNamespace:
    direct = config["directBinding"]
    retriever = config["retriever"]
    return SimpleNamespace(
        pred_dir=run_dir,
        limit=limit,
        batch_size=int(retriever["batchSize"]),
        max_tokens=int(direct["maxNewTokens"]),
        max_input_tokens=int(direct["maxInputTokens"]),
        structured_output=str(retriever["structuredOutput"]),
        case_timeout_seconds=int(retriever["caseTimeoutSeconds"]),
        cuda_visible_devices=str(retriever["cudaVisibleDevices"]),
        row_timeout_seconds=int(direct["rowTimeoutSeconds"]),
        binding_generator_parallelism=int(direct["parallelism"]),
        binding_generator_total_timeout_seconds=0,
    )



def native_direct_namespace(
    run_dir: Path,
    config: Mapping[str, Any],
    limit: int,
    family: str,
) -> SimpleNamespace:
    direct = config["directBinding"]
    retriever = config["retriever"]
    family_batches = retriever.get("familyBatchSizes", {})
    family_devices = retriever.get("familyCudaVisibleDevices", {})
    return SimpleNamespace(
        pred_dir=run_dir,
        limit=limit,
        batch_size=int(family_batches.get(family, retriever["batchSize"])),
        max_tokens=int(direct["maxNewTokens"]),
        max_input_tokens=int(
            direct.get("familyMaxInputTokens", {}).get(
                family, direct["maxInputTokens"]
            )
        ),
        structured_output="off",
        case_timeout_seconds=int(retriever["caseTimeoutSeconds"]),
        cuda_visible_devices=str(
            family_devices.get(family, retriever["cudaVisibleDevices"])
        ),
        t5gemma_cache_safe_input_tokens=int(
            retriever.get("generationCache", {})
            .get("t5gemma2", {})
            .get("disableAboveInputTokens", 0)
        ),
        append_label_descriptions=False,
    )

def extract_candidate(raw: str) -> str:
    text = legacy.extract_pred_text(str(raw or "")).strip()
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return text
    if isinstance(payload, dict):
        for key in ("RetFact", "retfact", "Rel_Fact"):
            value = payload.get(key)
            if value not in (None, ""):
                return str(value).strip()
    return text


def validate_binding(item: Any) -> tuple[bool, str]:
    required = {"ObjectName", "DataName", "Position", "Trend", "Num", "Text"}
    if not isinstance(item, dict):
        return False, "binding_not_object"
    if set(item) != required:
        return False, f"binding_keys={sorted(item)}"
    if not isinstance(item["ObjectName"], list) or not item["ObjectName"] or not all(
        isinstance(value, str) and value.strip() for value in item["ObjectName"]
    ):
        return False, "ObjectName_not_nonempty_string_array"
    if not isinstance(item["DataName"], str):
        return False, "DataName_not_string"
    if not isinstance(item["Trend"], str):
        return False, "Trend_not_string"
    if not isinstance(item["Text"], str):
        return False, "Text_not_string"
    if not isinstance(item["Num"], list) or not all(
        isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))
        for value in item["Num"]
    ):
        return False, "Num_not_finite_number_array"
    if not isinstance(item["Position"], list):
        return False, "Position_not_array"
    for position in item["Position"]:
        if not isinstance(position, dict) or set(position) != {"Begin", "End"}:
            return False, "Position_item_invalid_keys"
        for key in ("Begin", "End"):
            coordinate = position[key]
            if (
                not isinstance(coordinate, list)
                or len(coordinate) != 2
                or not all(isinstance(value, int) and not isinstance(value, bool) for value in coordinate)
            ):
                return False, f"Position_{key}_not_two_integer_array"
    return True, "valid"


def strict_parse_output(raw: str, *, converter: bool) -> tuple[list[dict[str, Any]], str, dict[str, Any]]:
    text = str(raw or "").strip()
    report: dict[str, Any] = {"strict": True, "rawSha256": sha256_text(text)}
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        report.update({"valid": False, "error": "json_decode", "detail": str(exc)})
        return [], "", report
    reason = ""
    if converter:
        if not isinstance(payload, dict) or set(payload) != {"result", "reason"}:
            report.update({"valid": False, "error": "converter_top_level_contract"})
            return [], "", report
        result = payload["result"]
        reason = payload["reason"]
        if not isinstance(reason, str):
            report.update({"valid": False, "error": "reason_not_string"})
            return [], "", report
    elif isinstance(payload, dict) and set(payload) == {"result", "reason"}:
        result = payload["result"]
        reason = payload["reason"]
        if not isinstance(reason, str):
            report.update({"valid": False, "error": "reason_not_string"})
            return [], "", report
    elif isinstance(payload, dict) and set(payload) == {"Binding"}:
        result = payload["Binding"]
    elif isinstance(payload, list):
        result = payload
    else:
        report.update({"valid": False, "error": "direct_top_level_contract"})
        return [], "", report
    if not isinstance(result, list):
        report.update({"valid": False, "error": "result_not_array"})
        return [], reason, report
    for index, item in enumerate(result):
        valid, detail = validate_binding(item)
        if not valid:
            report.update({"valid": False, "error": detail, "bindingIndex": index})
            return [], reason, report
    report.update({"valid": True, "bindings": len(result)})
    return result, reason, report


def reject_nonstandard_json_constant(value: str) -> Any:
    raise ValueError(f"non-standard JSON constant: {value}")


def nonformal_repair(raw: str) -> dict[str, Any]:
    text = str(raw or "").strip()
    fenced = re.sub(r"^\s*'''(?:json)?\s*|\s*'''\s*$", "", text, flags=re.IGNORECASE)
    starts = [position for position in (fenced.find("{"), fenced.find("[")) if position >= 0]
    if not starts:
        return {"available": False}
    start = min(starts)
    for end in range(len(fenced), start, -1):
        try:
            payload = json.loads(
                fenced[start:end],
                parse_constant=reject_nonstandard_json_constant,
            )
        except (json.JSONDecodeError, ValueError):
            continue
        return {"available": True, "payload": payload, "method": "fence-strip-balanced-json"}
    return {"available": False}


def converter_prompt(row: InputRow, candidate: str, case: MatrixCase) -> str:
    return "\n\n".join([
        f"Source: {row.source}",
        f"Upstream model: {case.source_id}",
        f"Prompt mode: {case.prompt_mode}",
        "Binding coordinate contract:\n" + POSITION_INDEX_CONTRACT,
        "Chart data (grounding only):\n" + row.data_compact,
        "Narrative text (grounding only):\n" + row.text,
        "Candidate RetFact (sole authority for fact existence):\n" + (candidate or "[BLANK]"),
    ])


def load_checkpoint(path: Path, expected_rows: int) -> list[dict[str, Any] | None]:
    rows: list[dict[str, Any] | None] = [None] * expected_rows
    if not path.is_file():
        return rows
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            item = json.loads(line)
            index = int(item["index"])
        except (json.JSONDecodeError, KeyError, TypeError, ValueError):
            continue
        if 0 <= index < expected_rows and str(item.get("status", "")).startswith("completed"):
            rows[index] = item
    return rows


def flush_checkpoint(path: Path, rows: Sequence[dict[str, Any] | None]) -> None:
    write_jsonl(path, [row for row in rows if row is not None])


def response_usage(response: Any) -> dict[str, Any]:
    usage = getattr(response, "usage", None)
    if usage is None:
        return {}
    if hasattr(usage, "model_dump"):
        return usage.model_dump()
    return {
        "prompt_tokens": getattr(usage, "prompt_tokens", None),
        "completion_tokens": getattr(usage, "completion_tokens", None),
        "total_tokens": getattr(usage, "total_tokens", None),
    }


def run_converter(
    *,
    case: MatrixCase,
    rows: Sequence[InputRow],
    candidates: Sequence[str],
    run_dir: Path,
    run_seed: int,
    config: Mapping[str, Any],
) -> tuple[list[str], dict[str, Any]]:
    converter = config["converter"]
    api_key = (
        os.environ.get(str(converter["apiKeyEnvironment"]))
        or str(converter["defaultApiKey"])
    )
    client = OpenAI(
        base_url=str(converter["baseUrl"]),
        api_key=api_key,
        timeout=float(converter["requestTimeoutSeconds"]),
        max_retries=0,
    )
    checkpoint_path = run_dir / "converter_raw_responses.jsonl"
    checkpoints = load_checkpoint(checkpoint_path, len(rows))
    outputs = [""] * len(rows)
    attempts_limit = int(converter["maxAttempts"])
    delays = [float(value) for value in converter["retryDelaysSeconds"]]
    parallelism = int(converter.get("parallelism", 1))
    if parallelism < 1:
        raise ProtocolError("converter.parallelism must be >= 1")

    def convert_one(index: int) -> tuple[int, dict[str, Any], str]:
        row = rows[index]
        candidate = candidates[index] if index < len(candidates) else ""
        prompt = converter_prompt(row, candidate, case)
        transport_errors: list[dict[str, Any]] = []
        started = time.monotonic()
        response = None
        for attempt in range(1, attempts_limit + 1):
            try:
                response = client.chat.completions.create(
                    model=str(converter["requestedModel"]),
                    messages=[
                        {"role": "system", "content": CONVERTER_SYSTEM_PROMPT},
                        {"role": "user", "content": prompt},
                    ],
                    reasoning_effort=str(converter["reasoningEffort"]),
                    max_tokens=int(converter["maxTokens"]),
                    temperature=0,
                    seed=run_seed + index,
                )
                break
            except Exception as exc:
                transport_errors.append({
                    "attempt": attempt,
                    "type": exc.__class__.__name__,
                    "message": str(exc),
                    "tracebackTail": traceback.format_exc()[-2000:],
                })
                if attempt < attempts_limit:
                    delay = delays[min(attempt - 1, len(delays) - 1)] if delays else 0
                    time.sleep(delay)
        if response is None:
            return index, {
                "index": index,
                "source": row.source,
                "status": "runtime_blocked",
                "transportErrors": transport_errors,
                "runtimeSeconds": time.monotonic() - started,
                "prompt": prompt,
                "promptSha256": sha256_text(prompt),
                "candidate": candidate,
                "candidateSha256": sha256_text(candidate),
            }, ""

        choice = response.choices[0]
        raw = choice.message.content or ""
        actual_model = str(getattr(response, "model", "") or "")
        if actual_model != str(converter["actualModelRequired"]):
            return index, {
                "index": index,
                "source": row.source,
                "status": "runtime_blocked_model_identity",
                "requestedModel": converter["requestedModel"],
                "actualModel": actual_model,
                "rawResponse": raw,
                "prompt": prompt,
                "promptSha256": sha256_text(prompt),
                "candidate": candidate,
                "candidateSha256": sha256_text(candidate),
                "transportErrors": transport_errors,
                "runtimeSeconds": time.monotonic() - started,
            }, ""

        result, reason, format_report = strict_parse_output(raw, converter=True)
        status = (
            "completed_format_valid"
            if format_report["valid"]
            else "completed_format_invalid"
        )
        record = {
            "index": index,
            "source": row.source,
            "status": status,
            "requestedModel": converter["requestedModel"],
            "actualModel": actual_model,
            "reasoningEffort": converter["reasoningEffort"],
            "seed": run_seed + index,
            "prompt": prompt,
            "promptSha256": sha256_text(prompt),
            "candidate": candidate,
            "candidateSha256": sha256_text(candidate),
            "rawResponse": raw,
            "rawResponseSha256": sha256_text(raw),
            "stopReason": getattr(choice, "finish_reason", None),
            "usage": response_usage(response),
            "format": format_report,
            "formalResult": result,
            "formalReason": reason,
            "nonformalRepair": (
                nonformal_repair(raw)
                if not format_report["valid"]
                else {"available": False}
            ),
            "transportErrors": transport_errors,
            "runtimeSeconds": time.monotonic() - started,
        }
        return index, record, raw

    pending: list[int] = []
    for index, existing in enumerate(checkpoints):
        if existing is not None:
            outputs[index] = str(existing.get("rawResponse") or "")
        else:
            pending.append(index)

    if parallelism == 1:
        completed = (convert_one(index) for index in pending)
        for index, record, raw in completed:
            checkpoints[index] = record
            outputs[index] = raw
            flush_checkpoint(checkpoint_path, checkpoints)
    else:
        with ThreadPoolExecutor(
            max_workers=min(parallelism, max(1, len(pending))),
            thread_name_prefix="experiment6-converter",
        ) as pool:
            futures = [pool.submit(convert_one, index) for index in pending]
            for future in as_completed(futures):
                index, record, raw = future.result()
                checkpoints[index] = record
                outputs[index] = raw
                flush_checkpoint(checkpoint_path, checkpoints)

    return outputs, {
        "stage": "retfact_to_binding_converter",
        "requestedModel": converter["requestedModel"],
        "requiredActualModel": converter["actualModelRequired"],
        "reasoningEffort": converter["reasoningEffort"],
        "baseUrl": converter["baseUrl"],
        "maxTokens": converter["maxTokens"],
        "maxAttempts": attempts_limit,
        "parallelism": parallelism,
        "checkpoint": str(checkpoint_path),
        "rowsRecorded": sum(item is not None for item in checkpoints),
    }


def normalize_predictions(
    *,
    case: MatrixCase,
    rows: Sequence[InputRow],
    raw_predictions: Sequence[str],
    run_number: int,
    run_seed: int,
    runtime: Mapping[str, Any],
    converter: bool,
    input_type: str = "type",
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    predictions: list[dict[str, Any]] = []
    repairs: list[dict[str, Any]] = []
    reports: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        raw = raw_predictions[index] if index < len(raw_predictions) else ""
        result, reason, report = strict_parse_output(raw, converter=converter)
        formal_result = result if report["valid"] else []
        requested_model = runtime.get("requestedModel") or case.source_id
        actual_model = (
            runtime.get("actualModel")
            or runtime.get("requiredActualModel")
            or case.source_id
        )
        model = {
            "requestedModel": requested_model,
            "actualModel": actual_model,
            "adapter": runtime.get("adapter"),
            "quantization": runtime.get("quantization"),
            "runtimeProfile": runtime.get("runtimeProfile"),
            "thinkingEnabled": runtime.get("thinkingEnabled"),
            "responseFormat": runtime.get("responseFormat"),
            "responseSchemaPath": runtime.get("responseSchemaPath"),
        }
        prompt = (
            row.retriever_prompts[case.prompt_mode]
            if converter
            else row.direct_prompts[case.prompt_mode]
        )
        predictions.append({
            "index": index,
            "number": row.number,
            "source": row.source,
            "outputId": case.output_id,
            "sourceId": case.source_id,
            "part": case.part,
            "official": case.official,
            "route": case.route,
            "promptMode": case.prompt_mode,
            "inputType": input_type,
            "run": run_number,
            "seed": run_seed,
            "model": model,
            "promptSha256": sha256_text(prompt),
            "requestedModel": requested_model,
            "actualModel": actual_model,
            "adapter": runtime.get("adapter"),
            "quantization": runtime.get("quantization"),
            "runtimeProfile": runtime.get("runtimeProfile"),
            "thinkingEnabled": runtime.get("thinkingEnabled"),
            "responseFormat": runtime.get("responseFormat"),
            "responseSchemaPath": runtime.get("responseSchemaPath"),
            "inputData": row.data_compact,
            "inputText": row.text,
            "result": formal_result,
            "reason": reason if report["valid"] else "",
            "rawResponse": raw,
            "rawResponseSha256": sha256_text(raw),
            "formatValid": bool(report["valid"]),
            "parserDiagnostic": report,
        })
        reports.append({"index": index, "source": row.source, **report})
        if not report["valid"]:
            repairs.append({
                "index": index,
                "source": row.source,
                "official": False,
                "excludedFromScores": True,
                "repair": nonformal_repair(raw),
            })
    valid_rows = sum(1 for report in reports if report["valid"])
    return predictions, repairs, {
        "rows": len(rows),
        "validRows": valid_rows,
        "invalidRows": len(rows) - valid_rows,
        "formatComplianceRate": valid_rows / len(rows) if rows else 0.0,
        "reports": reports,
    }


def canonical_raw_response_records(
    predictions: Sequence[Mapping[str, Any]],
    run_dir: Path,
    converter: bool,
) -> list[dict[str, Any]]:
    diagnostics: dict[int, Mapping[str, Any]] = {}
    if converter:
        path = run_dir / "converter_raw_responses.jsonl"
        if path.is_file():
            diagnostics = {
                int(item["index"]): item
                for item in (
                    json.loads(line)
                    for line in path.read_text(encoding="utf-8").splitlines()
                    if line.strip()
                )
                if isinstance(item, dict) and isinstance(item.get("index"), int)
            }
    else:
        for path in sorted((run_dir / "raw").glob("*.jsonl")):
            for line in path.read_text(encoding="utf-8").splitlines():
                if not line.strip():
                    continue
                item = json.loads(line)
                if (
                    isinstance(item, dict)
                    and isinstance(item.get("index"), int)
                    and isinstance(item.get("response"), dict)
                ):
                    diagnostics[int(item["index"])] = item

    records: list[dict[str, Any]] = []
    for prediction in predictions:
        index = int(prediction["index"])
        diagnostic = diagnostics.get(index, {})
        response = (
            diagnostic.get("response")
            if isinstance(diagnostic.get("response"), dict)
            else {}
        )
        finish_reasons = response.get("finishReasons")
        stop_reason = diagnostic.get("stopReason")
        if stop_reason is None and isinstance(finish_reasons, list):
            stop_reason = finish_reasons[0] if finish_reasons else None
        records.append({
            "index": index,
            "source": prediction["source"],
            "run": prediction["run"],
            "model": prediction["model"],
            "promptSha256": prediction["promptSha256"],
            "status": diagnostic.get("status"),
            "rawResponse": prediction["rawResponse"],
            "rawResponseSha256": prediction["rawResponseSha256"],
            "stopReason": stop_reason,
            "tokenUsage": diagnostic.get("usage") or response.get("usage"),
            "transportErrors": diagnostic.get("transportErrors", []),
            "formatValid": prediction["formatValid"],
            "parserDiagnostic": prediction["parserDiagnostic"],
        })
    return records


def public_model_runtime(case: MatrixCase, raw_runtime: Mapping[str, Any]) -> dict[str, Any]:
    engine_config = raw_runtime.get("config") if isinstance(raw_runtime.get("config"), dict) else {}
    actual = (
        engine_config.get("actual_model")
        or engine_config.get("model")
        or RETRIEVER_MODEL_IDS.get(str(raw_runtime.get("family") or ""))
        or raw_runtime.get("actual_engine")
        or case.source_id
    )
    route = engine_config.get("route") or raw_runtime.get("stage") or case.route
    engine = str(engine_config.get("engine") or "")
    runtime_profile = engine_config.get("runtime_profile") or raw_runtime.get("runtime_profile")
    quantization = None
    if raw_runtime.get("family") == "mistral":
        quantization = "4bit-nf4"
    elif route == "local_vllm_openai_compatible":
        runtime_profile = os.environ.get("VLLM_RUNTIME_PROFILE") or runtime_profile
        quantization = os.environ.get("VLLM_QUANTIZATION")
        actual_lower = str(actual).lower()
        if not quantization and engine == "qwen3_6" and "fp8" in actual_lower:
            quantization = "fp8"
        if not quantization and engine == "llama4" and "w4a16" in actual_lower:
            quantization = "w4a16"
    elif route == "local_llama_cpp_openai_compatible":
        quantization = os.environ.get("LLAMA_CPP_QUANT")

    family = raw_runtime.get("family")
    if family and runtime_profile is None:
        runtime_profile = (
            f"{family}-canonical-batch{int(raw_runtime.get('batch_size', 1))}"
        )

    thinking_enabled: bool | None = None
    response_format = os.environ.get("GENERATOR_RESPONSE_FORMAT") or None
    response_schema_path = os.environ.get("GENERATOR_RESPONSE_SCHEMA_PATH") or None
    if engine == "qwen3_6":
        thinking_enabled = str(
            os.environ.get("QWEN3_6_ENABLE_THINKING", "1")
        ).strip().lower() not in {"", "0", "false", "no", "off"}

    return {
        "requestedModel": case.source_id,
        "actualModel": actual,
        "route": route,
        "backend": engine_config.get("backend"),
        "runtimeProfile": runtime_profile,
        "thinkingEnabled": thinking_enabled,
        "responseFormat": response_format,
        "responseSchemaPath": response_schema_path,
        "adapter": raw_runtime.get("adapter_dir"),
        "quantization": quantization,
        "formalModel": engine_config.get("formal_model"),
        "raw": dict(raw_runtime),
    }


def failure_model_provenance(
    case: MatrixCase,
    config: Mapping[str, Any],
    runtime_stages: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    observed = next(
        (
            dict(stage)
            for stage in runtime_stages
            if isinstance(stage, Mapping)
            and any(stage.get(key) is not None for key in ("actualModel", "adapter", "runtimeProfile", "responseFormat", "responseSchemaPath"))
        ),
        {},
    )
    family = family_for_source(case.source_id)
    adapter = observed.get("adapter")
    quantization = observed.get("quantization")
    runtime_profile = observed.get("runtimeProfile")
    thinking_enabled = observed.get("thinkingEnabled")
    response_format = observed.get("responseFormat")
    response_schema_path = observed.get("responseSchemaPath")
    intended_model: str | None = None
    if family is not None:
        intended_model = RETRIEVER_MODEL_IDS[family]
        if adapter is None and legacy.is_finetuned_retriever_source(case.source_id):
            adapter = str(legacy.adapter_dir_for(case.source_id))
        if quantization is None and family == "mistral":
            quantization = "4bit-nf4"
        runtime_profile = runtime_profile or f"{family}-canonical"
    else:
        route_config = config.get("runtimeRoutes", {}).get(case.source_id, {})
        if isinstance(route_config, Mapping):
            intended_model = (
                model_reference(route_config.get("modelPath"))
                if route_config.get("modelPath") is not None
                else str(route_config.get("servedModel"))
                if route_config.get("servedModel") is not None
                else None
            )
            quantization = quantization or route_config.get("quantization")
            runtime_profile = runtime_profile or route_config.get("runtimeProfile")
            if thinking_enabled is None:
                thinking_enabled = route_config.get("enableThinking")
            response_format = response_format or route_config.get("responseFormat")
            response_schema_path = response_schema_path or route_config.get("responseSchemaPath")
    return {
        "requestedModel": case.source_id,
        "actualModel": observed.get("actualModel"),
        "intendedModel": intended_model or case.source_id,
        "adapter": adapter,
        "quantization": quantization,
        "runtimeProfile": runtime_profile,
        "thinkingEnabled": thinking_enabled,
        "responseFormat": response_format,
        "responseSchemaPath": response_schema_path,
    }


def runtime_blocked_row_records(
    case: MatrixCase,
    rows: Sequence[InputRow],
    run_number: int,
    run_seed: int,
    provenance: Mapping[str, Any],
    error: Mapping[str, Any],
) -> list[dict[str, Any]]:
    return [
        {
            "index": row.index,
            "number": row.number,
            "source": row.source,
            "outputId": case.output_id,
            "run": run_number,
            "seed": run_seed,
            "status": "runtime_blocked",
            "requestedModel": provenance.get("requestedModel"),
            "actualModel": provenance.get("actualModel"),
            "intendedModel": provenance.get("intendedModel"),
            "adapter": provenance.get("adapter"),
            "quantization": provenance.get("quantization"),
            "runtimeProfile": provenance.get("runtimeProfile"),
            "thinkingEnabled": provenance.get("thinkingEnabled"),
            "responseFormat": provenance.get("responseFormat"),
            "responseSchemaPath": provenance.get("responseSchemaPath"),
            "error": {
                "type": error.get("type"),
                "message": error.get("message"),
                "detailReference": "status.json#error",
            },
        }
        for row in rows
    ]


def load_retriever_candidate_checkpoint(
    path: Path,
    rows: Sequence[InputRow],
    run_number: int,
    run_seed: int,
) -> tuple[list[str], list[str]] | None:
    if not path.is_file():
        return None
    try:
        records = [
            json.loads(line)
            for line in path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
    except (OSError, json.JSONDecodeError):
        return None
    if len(records) != len(rows):
        return None
    raw_predictions: list[str] = []
    candidates: list[str] = []
    for index, (record, row) in enumerate(zip(records, rows)):
        if (
            not isinstance(record, dict)
            or record.get("index") != index
            or record.get("source") != row.source
            or record.get("run") != run_number
            or record.get("seed") != run_seed
            or not isinstance(record.get("raw"), str)
            or not isinstance(record.get("candidate"), str)
            or record.get("candidateSha256") != sha256_text(record["candidate"])
        ):
            return None
        raw_predictions.append(record["raw"])
        candidates.append(record["candidate"])
    return raw_predictions, candidates


def resumed_retriever_runtime(
    case: MatrixCase,
    family: str,
    config: Mapping[str, Any],
    run_seed: int,
    candidate_path: Path,
) -> dict[str, Any]:
    use_adapter = legacy.is_finetuned_retriever_source(case.source_id)
    adapter = str(legacy.adapter_dir_for(case.source_id)) if use_adapter else None
    retriever = config["retriever"]
    batch_size = int(retriever["batchSize"])
    cuda_device = retriever.get("familyCudaVisibleDevices", {}).get(
        family, retriever["cudaVisibleDevices"]
    )
    return {
        "requestedModel": case.source_id,
        "actualModel": RETRIEVER_MODEL_IDS[family],
        "route": case.route,
        "backend": "candidate-checkpoint",
        "runtimeProfile": f"{family}-canonical-batch{batch_size}",
        "thinkingEnabled": None,
        "responseFormat": None,
        "responseSchemaPath": None,
        "adapter": adapter,
        "quantization": "4bit-nf4" if family == "mistral" else None,
        "formalModel": None,
        "raw": {
            "family": family,
            "candidate_checkpoint": str(candidate_path),
            "candidate_checkpoint_reused": True,
            "cuda_visible_devices": str(cuda_device),
            "structured_output": str(retriever["structuredOutput"]),
            "max_input_tokens": int(retriever["maxInputTokens"]),
            "max_new_tokens": int(retriever["maxNewTokens"]),
            "attention_query_chunk_size": int(
                retriever.get("attentionQueryChunkSize", 0)
            ),
            "batch_size": batch_size,
            "run_seed": run_seed,
        },
    }


def archive_no_resume_artifacts(
    *,
    output_root: Path,
    run_dir: Path,
    manifest_path: Path,
    output_id: str,
    run_number: int,
) -> Path | None:
    if not run_dir.exists() and not manifest_path.exists():
        return None
    archive_dir = (
        output_root
        / "no_resume_archives"
        / output_id
        / f"run_{run_number:02d}"
        / f"{utc_now().replace(':', '').replace('-', '')}_{time.time_ns()}"
    )
    archive_dir.mkdir(parents=True, exist_ok=False)
    if run_dir.exists():
        os.replace(run_dir, archive_dir / "run_dir")
    if manifest_path.exists():
        os.replace(manifest_path, archive_dir / "manifest.json")
    return archive_dir


def retry_operation(
    operation,
    *,
    max_attempts: int,
    retry_delays_seconds: Sequence[float],
    on_error=None,
) -> tuple[Any, int, list[dict[str, Any]]]:
    """Run a transient operation with bounded, auditable retries."""
    if max_attempts < 1:
        raise ProtocolError("retry max_attempts must be at least one")
    errors: list[dict[str, Any]] = []
    for attempt in range(1, max_attempts + 1):
        try:
            return operation(), attempt, errors
        except Exception as exc:
            error = {
                "attempt": attempt,
                "errorType": type(exc).__name__,
                "message": str(exc),
            }
            errors.append(error)
            if on_error is not None:
                on_error(attempt, exc, error)
            if attempt >= max_attempts:
                raise
            delay_index = min(attempt - 1, len(retry_delays_seconds) - 1)
            delay = (
                float(retry_delays_seconds[delay_index])
                if retry_delays_seconds
                else 0.0
            )
            if delay > 0:
                time.sleep(delay)
    raise AssertionError("retry loop terminated unexpectedly")


def preserve_retriever_attempt(
    run_dir: Path,
    attempt: int,
    exc: Exception,
    error: Mapping[str, Any],
) -> None:
    """Preserve partial raw artifacts and traceback for a failed attempt."""
    attempt_dir = run_dir / "retriever_attempts" / f"attempt_{attempt:02d}"
    attempt_dir.mkdir(parents=True, exist_ok=True)
    copied: list[str] = []
    raw_dir = run_dir / "raw"
    if raw_dir.is_dir():
        for source in sorted(raw_dir.iterdir()):
            if source.is_file():
                destination = attempt_dir / source.name
                shutil.copy2(source, destination)
                copied.append(str(destination.relative_to(run_dir)))
    write_json(
        attempt_dir / "error.json",
        {
            **dict(error),
            "exception": repr(exc),
            "traceback": traceback.format_exc(),
            "copiedArtifacts": copied,
            "recordedAt": utc_now(),
        },
    )


def run_case_once(
    *,
    case: MatrixCase,
    rows: Sequence[InputRow],
    run_number: int,
    run_seed: int,
    output_root: Path,
    config: Mapping[str, Any],
    resume: bool,
    base_route_mode: str = "historical",
    run_identity: Mapping[str, Any] | None = None,
    cuda_visible_devices_override: str | None = None,
) -> dict[str, Any]:
    run_dir = output_root / "cases" / case.output_id / f"run_{run_number:02d}"
    manifest_path = output_root / "manifests" / f"{case.output_id}__run_{run_number:02d}.json"
    if resume and manifest_path.is_file():
        existing = read_json(manifest_path)
        expected_fingerprint = (
            run_identity.get("compatibilityFingerprint") if run_identity else None
        )
        if expected_fingerprint and existing.get("compatibilityFingerprint") != expected_fingerprint:
            raise ProtocolError(
                "resume compatibility fingerprint mismatch; start a fresh root"
            )
        if existing.get("status") in {"completed", "completed_with_format_errors"}:
            return existing
    no_resume_archive = (
        archive_no_resume_artifacts(
            output_root=output_root,
            run_dir=run_dir,
            manifest_path=manifest_path,
            output_id=case.output_id,
            run_number=run_number,
        )
        if not resume
        else None
    )
    run_dir.mkdir(parents=True, exist_ok=True)
    started_at = utc_now()
    started = time.monotonic()
    os.environ["EXPERIMENT6_RUN_SEED"] = str(run_seed)
    os.environ["PYTHONHASHSEED"] = str(run_seed)
    os.environ["GPT5_5_REASONING_EFFORT"] = "medium"
    os.environ["GENERATOR_TRANSIENT_MAX_ATTEMPTS"] = str(config["directBinding"]["maxAttempts"])
    os.environ["GENERATOR_TRANSIENT_RETRY_DELAYS"] = "5,15"
    prompt_rows = [
        csv_record(row, row.direct_prompts[case.prompt_mode], case.prompt_mode)
        for row in rows
    ]
    prompt_records = [{
        "index": row.index,
        "source": row.source,
        "promptMode": case.prompt_mode,
        "inputType": str(config.get("inputType") or "type"),
        "retrieverPrompt": row.retriever_prompts[case.prompt_mode],
        "retrieverPromptSha256": sha256_text(row.retriever_prompts[case.prompt_mode]),
        "directPrompt": row.direct_prompts[case.prompt_mode],
        "directPromptSha256": sha256_text(row.direct_prompts[case.prompt_mode]),
        "shotIds": list(row.shot_ids[case.prompt_mode]),
    } for row in rows]
    write_jsonl(run_dir / "prompts.jsonl", prompt_records)

    runtime_stages: list[dict[str, Any]] = []
    raw_predictions: list[str] = []
    converter_used = False
    runtime_blocked_rows = 0
    try:
        family = family_for_source(case.source_id)
        kind = source_kind(case.source_id)
        execution_route = effective_route(case, base_route_mode)
        direct_diagnostic = uses_direct_diagnostic_route(case, base_route_mode)
        native_direct = direct_diagnostic or (
            execution_route == "direct-binding" and kind == "base"
        )
        if case.route == "converter-control":
            candidates = [""] * len(rows)
            raw_predictions, converter_runtime = run_converter(
                case=case,
                rows=rows,
                candidates=candidates,
                run_dir=run_dir,
                run_seed=run_seed,
                config=config,
            )
            runtime_stages.append(converter_runtime)
            converter_used = True
        elif native_direct:
            assert family is not None
            csv_path = run_dir / "direct_input.csv"
            write_prompt_csv(csv_path, rows, case.prompt_mode, direct=True)
            direct_args = native_direct_namespace(
                run_dir, config, len(rows), family
            )
            if cuda_visible_devices_override is not None:
                direct_args.cuda_visible_devices = cuda_visible_devices_override
            scheduler = config["retriever"]["scheduler"]
            lock_namespace = str(scheduler["gpuLockNamespace"])
            safe_device = re.sub(
                r"[^A-Za-z0-9_.-]", "_", str(direct_args.cuda_visible_devices)
            )
            gpu_lock_path = (
                output_root / "runtime" / f"{lock_namespace}_{safe_device}.lock"
            )
            gpu_lock_path.parent.mkdir(parents=True, exist_ok=True)
            wait_started = time.monotonic()
            with gpu_lock_path.open("a+", encoding="utf-8") as gpu_lock:
                fcntl.flock(gpu_lock.fileno(), fcntl.LOCK_EX)
                lock_wait_seconds = time.monotonic() - wait_started
                try:
                    retry = config["retriever"].get("retry", {})
                    direct_result, direct_attempts, direct_errors = retry_operation(
                        lambda: legacy.run_retriever_case(
                            case.legacy_case,
                            csv_path,
                            case.prompt_mode,
                            direct_args,
                            use_adapter=False,
                            family_override=family,
                            raw_suffix=(
                                ".direct_diagnostic"
                                if direct_diagnostic
                                else ".direct_binding"
                            ),
                        ),
                        max_attempts=int(retry.get("maxAttempts", 1)),
                        retry_delays_seconds=tuple(
                            float(value)
                            for value in retry.get("retryDelaysSeconds", [])
                        ),
                        on_error=lambda attempt, exc, error: preserve_retriever_attempt(
                            run_dir, attempt, exc, error
                        ),
                    )
                    raw_predictions, direct_runtime = direct_result
                finally:
                    fcntl.flock(gpu_lock.fileno(), fcntl.LOCK_UN)
            if len(raw_predictions) != len(rows):
                raise ProtocolError(
                    f"native direct model produced {len(raw_predictions)}/{len(rows)} rows"
                )
            direct_runtime = dict(direct_runtime)
            direct_runtime.update({
                "attempts": direct_attempts,
                "retryErrors": direct_errors,
                "batch_size": int(direct_args.batch_size),
                "runtime_profile": (
                    f"{family}-"
                    + (
                        "direct-diagnostic" if direct_diagnostic else "direct-binding"
                    )
                    + f"-batch{int(direct_args.batch_size)}"
                ),
                "stage": (
                    "direct-diagnostic-native"
                    if direct_diagnostic
                    else "direct-binding-native"
                ),
                "prediction_contract": "binding-strict-json-unconstrained",
                "prompt_route": "direct",
                "converter_used": False,
                "generation_cache_used": False,
                "scheduler": {
                    "policy": str(scheduler["policy"]),
                    "gpu_lock_path": str(gpu_lock_path),
                    "gpu_lock_wait_seconds": lock_wait_seconds,
                },
            })
            runtime_stages.append(public_model_runtime(case, direct_runtime))
        elif execution_route in {"adapter-converter", "retriever-converter"}:
            if family is None:
                raise ProtocolError(f"registered retriever family missing: {case.source_id}")
            use_adapter = kind == "adapter"
            retriever_overrides: dict[str, Any] = {}
            if use_adapter:
                if ACTIVE_SOURCE_REGISTRY is None:
                    raise ProtocolError("formal adapter route requires source registry")
                resolved_adapter_source = ACTIVE_SOURCE_REGISTRY.resolve_source(
                    case.source_id, case.route
                )
                retriever_overrides = {
                    "family_override": family,
                    "adapter_dir_override": Path(
                        resolved_adapter_source["adapter"]["resolvedPath"]
                    ),
                }
            csv_path = run_dir / "retriever_input.csv"
            # FinFlier's long-context prompt is only built into direct_prompts["original"]
            # (see build_finflier_prompt above); retriever_prompts never receive it. Without
            # this branch, an adapter-converter case declaring inputType=FinFlier would
            # silently run the registered short prompt instead of the FinFlier prompt.
            if str(config.get("inputType") or "type") == "FinFlier":
                write_prompt_csv(csv_path, rows, "original", direct=True)
            else:
                write_prompt_csv(csv_path, rows, case.prompt_mode, direct=False)
            candidate_path = run_dir / "retriever_candidates.jsonl"
            candidate_checkpoint = (
                load_retriever_candidate_checkpoint(
                    candidate_path, rows, run_number, run_seed
                )
                if resume else None
            )
            if candidate_checkpoint is not None:
                retriever_predictions, candidates = candidate_checkpoint
                runtime_stages.append(
                    resumed_retriever_runtime(
                        case, family, config, run_seed, candidate_path
                    )
                )
            else:
                retriever_args = retriever_namespace(
                    run_dir, config, len(rows), family=family
                )
                if cuda_visible_devices_override is not None:
                    retriever_args.cuda_visible_devices = cuda_visible_devices_override
                scheduler = config["retriever"]["scheduler"]
                lock_namespace = str(scheduler["gpuLockNamespace"])
                safe_device = re.sub(
                    r"[^A-Za-z0-9_.-]", "_", str(retriever_args.cuda_visible_devices)
                )
                gpu_lock_path = (
                    output_root / "runtime" / f"{lock_namespace}_{safe_device}.lock"
                )
                gpu_lock_path.parent.mkdir(parents=True, exist_ok=True)
                wait_started = time.monotonic()
                with gpu_lock_path.open("a+", encoding="utf-8") as gpu_lock:
                    fcntl.flock(gpu_lock.fileno(), fcntl.LOCK_EX)
                    lock_wait_seconds = time.monotonic() - wait_started
                    try:
                        retry = config["retriever"].get("retry", {})
                        retriever_result, retriever_attempts, retriever_errors = retry_operation(
                            lambda: legacy.run_retriever_case(
                                case.legacy_case,
                                csv_path,
                                case.prompt_mode,
                                retriever_args,
                                use_adapter=use_adapter,
                                **retriever_overrides,
                            ),
                            max_attempts=int(retry.get("maxAttempts", 1)),
                            retry_delays_seconds=tuple(
                                float(value)
                                for value in retry.get("retryDelaysSeconds", [])
                            ),
                            on_error=lambda attempt, exc, error: preserve_retriever_attempt(
                                run_dir, attempt, exc, error
                            ),
                        )
                        retriever_predictions, retriever_runtime = retriever_result
                    finally:
                        fcntl.flock(gpu_lock.fileno(), fcntl.LOCK_UN)
                retriever_runtime = dict(retriever_runtime)
                retriever_runtime["attempts"] = retriever_attempts
                retriever_runtime["retryErrors"] = retriever_errors
                retriever_runtime["batch_size"] = int(retriever_args.batch_size)
                retriever_runtime["runtime_profile"] = (
                    f"{family}-canonical-batch{int(retriever_args.batch_size)}"
                )
                retriever_runtime["scheduler"] = {
                    "policy": str(scheduler["policy"]),
                    "gpu_lock_path": str(gpu_lock_path),
                    "gpu_lock_wait_seconds": lock_wait_seconds,
                    "converter_overlap": bool(scheduler["converterOverlap"]),
                }
                if len(retriever_predictions) != len(rows):
                    raise ProtocolError(
                        f"retriever produced {len(retriever_predictions)}/{len(rows)} rows"
                    )
                candidates = [
                    extract_candidate(value) for value in retriever_predictions
                ]
                write_jsonl(candidate_path, [
                    {
                        "index": index,
                        "source": rows[index].source,
                        "run": run_number,
                        "seed": run_seed,
                        "batchSize": int(retriever_args.batch_size),
                        "raw": retriever_predictions[index],
                        "candidate": candidates[index],
                        "candidateSha256": sha256_text(candidates[index]),
                    }
                    for index in range(len(rows))
                ])
                runtime_stages.append(
                    public_model_runtime(case, retriever_runtime)
                )
            raw_predictions, converter_runtime = run_converter(
                case=case,
                rows=rows,
                candidates=candidates,
                run_dir=run_dir,
                run_seed=run_seed,
                config=config,
            )
            runtime_stages.append(converter_runtime)
            converter_used = True
        else:
            csv_path = run_dir / "direct_input.csv"
            write_prompt_csv(csv_path, rows, case.prompt_mode, direct=True)
            direct_args = direct_namespace(run_dir, config, len(rows))
            first_error: Exception | None = None
            if execution_route != "direct-binding":
                raise ProtocolError(
                    f"no execution branch for route {execution_route}: {case.output_id}"
                )
            direct_runtime: dict[str, Any] = {}
            for attempt in range(1, 2):
                try:
                    raw_predictions, direct_runtime = legacy.run_no_adapter_case(
                        case.legacy_case,
                        csv_path,
                        case.prompt_mode,
                        prompt_rows,
                        direct_args,
                    )
                    break
                except Exception as exc:
                    first_error = first_error or exc
                    write_json(run_dir / f"direct_attempt_{attempt:02d}_error.json", {
                        "attempt": attempt,
                        "type": exc.__class__.__name__,
                        "message": str(exc),
                        "tracebackTail": traceback.format_exc()[-4000:],
                    })
            if not raw_predictions:
                raise first_error or RuntimeError("direct generation returned no predictions")
            if len(raw_predictions) != len(rows):
                raise ProtocolError(f"direct model produced {len(raw_predictions)}/{len(rows)} rows")
            runtime_stages.append(public_model_runtime(case, direct_runtime))

        checkpoint = run_dir / "converter_raw_responses.jsonl"
        if converter_used and checkpoint.is_file():
            checkpoint_rows = [
                json.loads(line)
                for line in checkpoint.read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            runtime_blocked_rows = sum(
                1 for item in checkpoint_rows if str(item.get("status", "")).startswith("runtime_blocked")
            )
        elif not converter_used:
            direct_raw_paths = sorted((run_dir / "raw").glob("*.jsonl"))
            for direct_raw_path in direct_raw_paths:
                direct_rows = [
                    json.loads(line)
                    for line in direct_raw_path.read_text(encoding="utf-8").splitlines()
                    if line.strip()
                ]
                runtime_blocked_rows += sum(
                    1
                    for item in direct_rows
                    if str(item.get("status", "")).startswith("runtime_blocked")
                )
        model_runtime = runtime_stages[0] if runtime_stages else {}
        if converter_used:
            model_runtime = {
                **model_runtime,
                "requestedModel": case.source_id,
                "actualModel": model_runtime.get("actualModel") or case.source_id,
            }
        predictions, repairs, format_report = normalize_predictions(
            case=case,
            rows=rows,
            raw_predictions=raw_predictions,
            run_number=run_number,
            run_seed=run_seed,
            runtime=model_runtime,
            converter=converter_used,
            input_type=str(config.get("inputType") or "type"),
        )
        write_jsonl(run_dir / "predictions.jsonl", predictions)
        write_jsonl(
            run_dir / "raw_response.jsonl",
            canonical_raw_response_records(predictions, run_dir, converter_used),
        )
        write_jsonl(run_dir / "repair_predictions.nonformal.jsonl", repairs)
        write_json(run_dir / "format_report.json", format_report)
        write_json(run_dir / "runtime.json", {
            "startedAt": started_at,
            "finishedAt": utc_now(),
            "runtimeSeconds": time.monotonic() - started,
            "seed": run_seed,
            "seedSupport": (
                "standard-chat-completions-or-subprocess-recorded"
                if family is None
                else "transformers-set-seed-sampled"
                if family == "t5gemma2"
                else "retriever-greedy-deterministic-seed-recorded"
            ),
            "stages": runtime_stages,
        })
        status = (
            "runtime_blocked"
            if runtime_blocked_rows
            else "completed_with_format_errors"
            if format_report["invalidRows"]
            else "completed"
        )
        artifact_files = {
            "predictions": str(run_dir / "predictions.jsonl"),
            "rawResponse": str(run_dir / "raw_response.jsonl"),
            "prompts": str(run_dir / "prompts.jsonl"),
            "runtime": str(run_dir / "runtime.json"),
            "formatReport": str(run_dir / "format_report.json"),
            "nonformalRepair": str(run_dir / "repair_predictions.nonformal.jsonl"),
        }
        if checkpoint.is_file():
            artifact_files["converterRawResponses"] = str(checkpoint)
        candidate_path = run_dir / "retriever_candidates.jsonl"
        if candidate_path.is_file():
            artifact_files["retrieverCandidates"] = str(candidate_path)
        for stage_index, stage_runtime in enumerate(runtime_stages, start=1):
            raw_runtime = stage_runtime.get("raw") if isinstance(stage_runtime, dict) else None
            raw_output = raw_runtime.get("raw_output") if isinstance(raw_runtime, dict) else None
            if raw_output and Path(str(raw_output)).is_file():
                artifact_files[f"stage{stage_index}Raw"] = str(raw_output)
        artifact_hashes = {
            name: sha256_file(Path(path))
            for name, path in artifact_files.items()
            if Path(path).is_file()
        }
        manifest = {
            "protocol": config["protocol"],
            **dict(run_identity or {}),
            "outputId": case.output_id,
            "sourceId": case.source_id,
            "part": case.part,
            "official": case.official,
            "route": case.route,
            "declaredRoute": case.route,
            "effectiveRoute": effective_route(case, base_route_mode),
            "baseRouteMode": base_route_mode,
            "promptMode": case.prompt_mode,
            "inputType": str(config.get("inputType") or "type"),
            "run": run_number,
            "seed": run_seed,
            "expectedRows": len(rows),
            "acceptedRows": format_report["validRows"],
            "rejectedRows": format_report["invalidRows"],
            "runtimeBlockedRows": runtime_blocked_rows,
            "formatComplianceRate": format_report["formatComplianceRate"],
            "status": status,
            "startedAt": started_at,
            "finishedAt": utc_now(),
            "runtimeSeconds": time.monotonic() - started,
            "noResumeArchive": (
                str(no_resume_archive) if no_resume_archive is not None else None
            ),
            "requestedModel": case.source_id,
            "actualModel": model_runtime.get("actualModel") or case.source_id,
            "adapter": model_runtime.get("adapter"),
            "quantization": model_runtime.get("quantization"),
            "runtimeProfile": model_runtime.get("runtimeProfile"),
            "thinkingEnabled": model_runtime.get("thinkingEnabled"),
            "responseFormat": model_runtime.get("responseFormat"),
            "responseSchemaPath": model_runtime.get("responseSchemaPath"),
            "converterModel": config["converter"]["requestedModel"] if converter_used else None,
            "reasoningEffort": (
                config["converter"]["reasoningEffort"]
                if converter_used or case.source_id == "gpt5_5"
                else None
            ),
            "files": artifact_files,
            "hashes": artifact_hashes,
        }
        write_json(run_dir / "status.json", manifest)
        write_json(manifest_path, manifest)
        return manifest
    except Exception as exc:
        error_record = {
            "type": exc.__class__.__name__,
            "message": str(exc),
            "tracebackTail": traceback.format_exc()[-6000:],
        }
        provenance = failure_model_provenance(case, config, runtime_stages)
        blocked_rows_path = run_dir / "runtime_blocked_rows.jsonl"
        write_jsonl(
            blocked_rows_path,
            runtime_blocked_row_records(
                case, rows, run_number, run_seed, provenance, error_record
            ),
        )
        failure_files = {
            "prompts": str(run_dir / "prompts.jsonl"),
            "runtimeBlockedRows": str(blocked_rows_path),
        }
        for raw_index, raw_path in enumerate(sorted((run_dir / "raw").glob("*")), start=1):
            if raw_path.is_file():
                failure_files[f"raw{raw_index}"] = str(raw_path)
        for error_index, error_path in enumerate(sorted(run_dir.glob("*_error.json")), start=1):
            if error_path.is_file():
                failure_files[f"attemptError{error_index}"] = str(error_path)
        for name, path in (
            ("converterRawResponses", run_dir / "converter_raw_responses.jsonl"),
            ("retrieverCandidates", run_dir / "retriever_candidates.jsonl"),
        ):
            if path.is_file():
                failure_files[name] = str(path)
        failure = {
            "protocol": config["protocol"],
            "outputId": case.output_id,
            **dict(run_identity or {}),
            "sourceId": case.source_id,
            "part": case.part,
            "declaredRoute": case.route,
            "official": case.official,
            "route": case.route,
            "effectiveRoute": effective_route(case, base_route_mode),
            "baseRouteMode": base_route_mode,
            "promptMode": case.prompt_mode,
            "inputType": str(config.get("inputType") or "type"),
            "run": run_number,
            "seed": run_seed,
            "expectedRows": len(rows),
            "acceptedRows": 0,
            "rejectedRows": 0,
            "runtimeBlockedRows": len(rows),
            "status": "runtime_blocked",
            "startedAt": started_at,
            "finishedAt": utc_now(),
            "runtimeSeconds": time.monotonic() - started,
            "noResumeArchive": (
                str(no_resume_archive) if no_resume_archive is not None else None
            ),
            "requestedModel": provenance["requestedModel"],
            "actualModel": provenance["actualModel"],
            "intendedModel": provenance["intendedModel"],
            "adapter": provenance["adapter"],
            "quantization": provenance["quantization"],
            "runtimeProfile": provenance["runtimeProfile"],
            "thinkingEnabled": provenance["thinkingEnabled"],
            "responseFormat": provenance["responseFormat"],
            "responseSchemaPath": provenance["responseSchemaPath"],
            "converterModel": config["converter"]["requestedModel"] if converter_used else None,
            "reasoningEffort": (
                config["converter"]["reasoningEffort"]
                if converter_used or case.source_id == "gpt5_5"
                else None
            ),
            "formatComplianceRate": None,
            "error": error_record,
            "files": failure_files,
            "hashes": {
                name: sha256_file(Path(path))
                for name, path in failure_files.items()
                if Path(path).is_file()
            },
        }
        write_json(run_dir / "status.json", failure)
        write_json(manifest_path, failure)
        return failure


def git_state() -> dict[str, Any]:
    def run(*args: str) -> str:
        completed = subprocess.run(
            ["git", *args],
            cwd=REPO_ROOT,
            check=False,
            capture_output=True,
            text=True,
        )
        return completed.stdout.strip()
    return {
        "head": run("rev-parse", "HEAD"),
        "branch": run("rev-parse", "--abbrev-ref", "HEAD"),
        "statusPorcelain": run("status", "--short"),
    }


def select_smoke_cases(cases: Sequence[MatrixCase]) -> list[MatrixCase]:
    selected: list[MatrixCase] = []
    seen: set[tuple[str, str]] = set()
    for case in cases:
        key = (case.source_id, "control" if not case.official else "model")
        if key in seen:
            continue
        seen.add(key)
        selected.append(case)
    return selected


def prediction_count_summary(
    all_cases: Sequence[MatrixCase],
    selected_cases: Sequence[MatrixCase],
    configured_runs: Sequence[int],
    selected_runs: Sequence[int],
    expected_rows: int,
    selected_rows: int,
) -> dict[str, int]:
    """Report full-protocol and invocation-scoped prediction counts."""
    return {
        "expectedFormalPredictions": (
            sum(case.official for case in all_cases)
            * len(configured_runs)
            * expected_rows
        ),
        "expectedControlPredictions": (
            sum(not case.official for case in all_cases)
            * len(configured_runs)
            * expected_rows
        ),
        "selectedFormalPredictions": (
            sum(case.official for case in selected_cases)
            * len(selected_runs)
            * selected_rows
        ),
        "selectedControlPredictions": (
            sum(not case.official for case in selected_cases)
            * len(selected_runs)
            * selected_rows
        ),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--config",
        type=Path,
        default=REPO_ROOT / "config" / "experiment6_narrative2_generation.json",
    )
    parser.add_argument("--output-root", type=Path, required=True)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--preflight-only", action="store_true")
    parser.add_argument("--smoke-only", action="store_true")
    parser.add_argument("--case", action="append", default=[])
    parser.add_argument("--source-id", action="append", default=[])
    parser.add_argument("--run", type=int, action="append", default=[])
    parser.add_argument(
        "--base-route-mode",
        choices=BASE_ROUTE_MODES,
        default="formal",
        help=(
            "formal obeys case.route; historical explicitly replays the old "
            "base-to-RetFact-to-converter diagnostic; "
            "direct-diagnostic uses native base binding generation without "
            "adapter, structured decoding, or converter"
        ),
    )
    parser.add_argument(
        "--cuda-visible-devices",
        choices=("0", "1", "0,1", "cpu"),
        help=(
            "execution-only GPU override recorded outside the scientific fingerprint; "
            "0,1 permits Transformers device_map sharding for long-context FLAN"
        ),
    )
    parser.add_argument(
        "--row-source",
        help="exact Source selector for an isolated smoke; does not alter the source workbook",
    )
    parser.add_argument("--no-resume", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    config = load_config(args.config.resolve())
    output_root = args.output_root.resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    bundle = workspace_path(config["generationBundle"])
    all_cases = expand_matrix(config)
    selected_cases = all_cases
    if args.case:
        requested = set(args.case)
        selected_cases = [case for case in selected_cases if case.output_id in requested]
        missing = requested - {case.output_id for case in selected_cases}
        if missing:
            raise ProtocolError(f"unknown --case values: {sorted(missing)}")
    if args.source_id:
        requested_sources = set(args.source_id)
        known_sources = {case.source_id for case in all_cases}
        missing_sources = requested_sources - known_sources
        if missing_sources:
            raise ProtocolError(f"unknown --source-id values: {sorted(missing_sources)}")
        selected_cases = [
            case for case in selected_cases if case.source_id in requested_sources
        ]
        if not selected_cases:
            raise ProtocolError("--case and --source-id filters selected no cases")
    if args.smoke_only:
        selected_cases = select_smoke_cases(selected_cases)
    if args.base_route_mode == "direct-diagnostic":
        for case in selected_cases:
            uses_direct_diagnostic_route(case, args.base_route_mode)
    effective_limit = 1 if args.smoke_only and args.limit <= 0 else args.limit
    if ACTIVE_SOURCE_REGISTRY is None:
        raise ProtocolError("source registry was not loaded")
    config_sha256 = sha256_file(args.config.resolve())
    implementation_files = {
        "narrativeRunner": Path(__file__).resolve(),
        "bindingGenerationRunner": Path(legacy.__file__).resolve(),
        "sourceRegistryLoader": SCRIPTS_ROOT / "experiment6_registry.py",
        "pathResolver": SCRIPTS_ROOT / "experiment6_paths.py",
        "seq2seqInference": legacy.RETRIEVER_ROOT / "seq2seq_retriever.py",
        "mistralInference": (
            legacy.RETRIEVER_ROOT / "Mistral" / "mistral_inference.py"
        ),
    }
    if config.get("finflierPrompt") is not None:
        implementation_files["finflierPromptHelper"] = (
            REPO_ROOT / "narrative" / "finflier_prompt.py"
        )
    implementation_sha256s = {
        name: sha256_file(path) for name, path in implementation_files.items()
    }
    compatibility = ACTIVE_SOURCE_REGISTRY.compatibility_snapshot(
        [
            {
                "outputId": case.output_id,
                "sourceId": case.source_id,
                "promptMode": case.prompt_mode,
                "inputType": str(config.get("inputType") or "type"),
                "route": case.route,
            }
            for case in all_cases
        ],
        execution_mode=(
            f"{args.base_route_mode};rows="
            f"{args.row_source or effective_limit or int(config['expectedRows'])}"
        ),
        generation_config_sha256=config_sha256,
        prompt_policy_version=str(config["promptBuilder"]["promptPolicyVersion"]),
        input_workbook_sha256=str(config["inputWorkbook"]["sha256"]),
        implementation_sha256s=implementation_sha256s,
    )
    generation_assets = ACTIVE_SOURCE_REGISTRY.generation_assets()
    resolved_sources = {
        case.output_id: ACTIVE_SOURCE_REGISTRY.resolve_source(
            case.source_id, case.route
        )
        for case in all_cases
    }
    bundle_report = verify_bundle(bundle)
    freeze_compatibility_snapshot(output_root, compatibility)
    rows, input_report = read_input_rows(config, effective_limit, args.row_source)
    rows, prompt_bundle_report = materialize_prompt_bundles(
        output_root, rows, config
    )
    token_report = token_preflight(
        rows, selected_cases, config, args.base_route_mode
    )
    runs = [int(value) for value in config["runs"]]
    if args.run:
        runs = args.run
    if args.smoke_only:
        runs = [1]

    prediction_counts = prediction_count_summary(
        all_cases=all_cases,
        selected_cases=selected_cases,
        configured_runs=[int(value) for value in config["runs"]],
        selected_runs=runs,
        expected_rows=int(config["expectedRows"]),
        selected_rows=len(rows),
    )

    preflight = {
        "protocol": config["protocol"],
        "time": utc_now(),
        "configPath": str(args.config.resolve()),
        "configSha256": config_sha256,
        "compatibility": compatibility,
        "sourceRegistry": {
            "path": str(ACTIVE_SOURCE_REGISTRY.path),
            "sha256": ACTIVE_SOURCE_REGISTRY.file_sha256,
        },
        "resolvedGenerationAssets": generation_assets,
        "resolvedSources": resolved_sources,
        "bundle": bundle_report,
        "promptBundles": prompt_bundle_report,
        "input": input_report,
        "matrix": {
            "officialCases": sum(case.official for case in all_cases),
            "diagnosticCases": sum(not case.official for case in all_cases),
            "partCounts": dict(Counter(case.part for case in all_cases if case.official)),
            "selectedCases": [case.__dict__ for case in selected_cases],
            "runs": runs,
            "baseRouteMode": args.base_route_mode,
            "cudaVisibleDevicesOverride": args.cuda_visible_devices,
            "rowSource": args.row_source,
            **prediction_counts,
        },
        "tokens": token_report,
        "git": git_state(),
        "status": "passed",
    }
    invocation_preflight_dir = output_root / "preflight_invocations"
    invocation_preflight_dir.mkdir(parents=True, exist_ok=True)
    invocation_id = (
        utc_now().replace(":", "").replace("-", "")
        + "__"
        + sha256_text(json.dumps({
            "cases": [case.output_id for case in selected_cases],
            "runs": runs,
            "limit": effective_limit,
            "rowSource": args.row_source,
            "baseRouteMode": args.base_route_mode,
        }, sort_keys=True))[:12]
    )
    write_json(invocation_preflight_dir / f"{invocation_id}.json", preflight)
    root_preflight_path = output_root / "preflight.json"
    if len(selected_cases) == len(all_cases) or not root_preflight_path.is_file():
        write_json(root_preflight_path, preflight)
    freeze_generation_snapshot(output_root, config)
    if args.preflight_only:
        print(json.dumps({
            "status": "preflight_passed",
            "outputRoot": str(output_root),
            "officialCases": preflight["matrix"]["officialCases"],
            "diagnosticCases": preflight["matrix"]["diagnosticCases"],
            "rows": len(rows),
            "selectedFormalPredictions": preflight["matrix"]["selectedFormalPredictions"],
            "selectedControlPredictions": preflight["matrix"]["selectedControlPredictions"],
        }, ensure_ascii=False))
        return

    report_path = output_root / "generation_report.json"
    report_lock_path = output_root / "generation_report.lock"

    def manifest_snapshot() -> list[dict[str, Any]]:
        snapshot: list[dict[str, Any]] = []
        for path in sorted((output_root / "manifests").glob("*.json")):
            try:
                item = read_json(path)
            except (OSError, json.JSONDecodeError):
                continue
            if item.get("protocol") == config["protocol"]:
                snapshot.append(item)
        return snapshot

    manifest_by_key = {
        (str(item.get("outputId")), int(item.get("run"))): item
        for item in manifest_snapshot()
    }
    invocation_manifests: list[dict[str, Any]] = []
    for case in selected_cases:
        for run_number in runs:
            seed = int(config["seedBase"]) + run_number
            print(json.dumps({
                "time": utc_now(),
                "event": "case_run_started",
                "outputId": case.output_id,
                "run": run_number,
                "seed": seed,
            }, ensure_ascii=False), flush=True)
            manifest = run_case_once(
                case=case,
                rows=rows,
                run_number=run_number,
                run_seed=seed,
                output_root=output_root,
                config=config,
                resume=not args.no_resume,
                base_route_mode=args.base_route_mode,
                cuda_visible_devices_override=args.cuda_visible_devices,
                run_identity={
                    "compatibilityFingerprint": compatibility["sha256"],
                    "sourceRegistry": {
                        "path": str(ACTIVE_SOURCE_REGISTRY.path),
                        "sha256": ACTIVE_SOURCE_REGISTRY.file_sha256,
                    },
                    "resolvedGenerationAssets": generation_assets,
                    "resolvedSource": resolved_sources[case.output_id],
                    "executionTopology": gpu_execution_identity(
                        args.cuda_visible_devices
                    ),
                },
            )
            invocation_manifests.append(manifest)
            manifest_by_key[(case.output_id, run_number)] = manifest
            for disk_manifest in manifest_snapshot():
                manifest_by_key[(
                    str(disk_manifest["outputId"]), int(disk_manifest["run"]),
                )] = disk_manifest
            merged_manifests = [
                manifest_by_key[key] for key in sorted(manifest_by_key)
            ]
            print(json.dumps({
                "time": utc_now(),
                "event": "case_run_finished",
                "outputId": case.output_id,
                "run": run_number,
                "status": manifest["status"],
                "acceptedRows": manifest.get("acceptedRows"),
                "rejectedRows": manifest.get("rejectedRows"),
                "runtimeBlockedRows": manifest.get("runtimeBlockedRows"),
            }, ensure_ascii=False), flush=True)
            with report_lock_path.open("a+", encoding="utf-8") as lock_handle:
                fcntl.flock(lock_handle.fileno(), fcntl.LOCK_EX)
                current_manifests = manifest_snapshot()
                write_json(report_path, {
                    "protocol": config["protocol"],
                    "updatedAt": utc_now(),
                    "lastInvocationSelectedCases": len(selected_cases),
                    "lastInvocationSelectedRuns": len(runs),
                    "manifests": current_manifests,
                    "statusCounts": dict(
                        Counter(item["status"] for item in current_manifests)
                    ),
                    "complete": (
                        len(current_manifests)
                        == (
                            int(config["expectedOfficialCases"])
                            + int(config["expectedDiagnosticCases"])
                        ) * len(config["runs"])
                        and all(
                            item["status"]
                            in {"completed", "completed_with_format_errors"}
                            for item in current_manifests
                        )
                    ),
                })
                fcntl.flock(lock_handle.fileno(), fcntl.LOCK_UN)
    blocked = [item for item in invocation_manifests if item["status"] == "runtime_blocked"]
    print(json.dumps({
        "status": "runtime_blocked" if blocked else "completed",
        "outputRoot": str(output_root),
        "caseRuns": len(invocation_manifests),
        "blockedCaseRuns": len(blocked),
    }, ensure_ascii=False))
    if blocked:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
