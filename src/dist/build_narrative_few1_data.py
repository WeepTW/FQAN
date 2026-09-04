#!/usr/bin/env python3
"""Build the Experiment 6 few_1_epoch narrative binding dataset.

The script converts ``data/src/narratives/narratives1.xlsx`` into the JSONL
shape consumed by ``dist/evaluate_data_binding.py``. Controlled prediction
JSONL is an explicit smoke option only; formal runs must supply real prediction
JSONL separately.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl

from binding_extraction import extract_result_items, item_dict


REPO_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = REPO_ROOT.parent
DEFAULT_SOURCE_XLSX = WORKSPACE_ROOT / "data" / "src" / "narratives" / "narratives1.xlsx"
DEFAULT_OUTPUT_ROOT = WORKSPACE_ROOT / "data" / "financial_narratives" / "few_1_epoch"
DEFAULT_GOLD_DIR = WORKSPACE_ROOT / "data" / "financial_narratives" / "gold"
DEFAULT_TESTING_GOLD_JSONL = WORKSPACE_ROOT / "data" / "testing" / "narratives_gold.jsonl"
DEFAULT_EXPT_ID = "experiment_6_narrative_few_1_epoch"
DEFAULT_MATRIX = (
    "6_flan_z:finqa_flan_z:narrative_zero_shot "
    "6_flan_m:finqa_flan_m:narrative_many_shot "
    "6_flan_d:finqa_flan_d:narrative_dynamic_shot "
    "6_mistral_z:finqa_mistral_z:narrative_zero_shot "
    "6_mistral_m:finqa_mistral_m:narrative_many_shot "
    "6_mistral_d:finqa_mistral_d:narrative_dynamic_shot "
    "6_t5gemma2_z:finqa_t5gemma2_z:narrative_zero_shot "
    "6_t5gemma2_m:finqa_t5gemma2_m:narrative_many_shot "
    "6_t5gemma2_d:finqa_t5gemma2_d:narrative_dynamic_shot "
    "6_untrain_z:untrained_models:narrative_zero_shot "
    "6_untrain_m:untrained_models:narrative_many_shot "
    "6_untrain_d:untrained_models:narrative_dynamic_shot "
    "original_no_gpt41:retriever_models_no_gpt41:narrative_original "
    "original_with_gpt41:retriever_models_with_gpt41:narrative_original"
)
PROMPT_TYPE_GOLD_PATHS = {
    "narrative_original": WORKSPACE_ROOT / "data" / "finqa_original" / "narratives_gold.jsonl",
    "narrative_zero_shot": WORKSPACE_ROOT / "data" / "finqa_zero_shot" / "narratives_gold.jsonl",
    "narrative_many_shot": WORKSPACE_ROOT / "data" / "finqa_many_shot" / "narratives_gold.jsonl",
    "narrative_dynamic_shot": WORKSPACE_ROOT / "data" / "finqa_dynamic_shot" / "narratives_gold.jsonl",
}


@dataclass(frozen=True)
class MatrixCase:
    experiment_id: str
    source_id: str
    narrative_route: str


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def find_column(headers: dict[str, int], preferred: str, aliases: Iterable[str]) -> int:
    for key in (preferred, *aliases):
        normalized = normalize_header(key)
        if normalized in headers:
            return headers[normalized]
    raise ValueError(f"Cannot find column {preferred!r}; available={sorted(headers)}")


def stable_case_id(raw_value: Any, row_index: int) -> str:
    text = str(raw_value or "").strip()
    if not text:
        return f"row_{row_index}"
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("_") or f"row_{row_index}"


def parse_matrix(raw: str) -> list[MatrixCase]:
    cases: list[MatrixCase] = []
    for item in raw.split():
        parts = item.split(":")
        if len(parts) != 3 or not all(parts):
            raise ValueError(f"Invalid matrix item: {item}")
        cases.append(MatrixCase(parts[0], parts[1], parts[2]))
    return cases


def read_rows(args: argparse.Namespace) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(args.source_xlsx, data_only=True)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet {args.sheet!r} not found; available={workbook.sheetnames}")
    worksheet = workbook[args.sheet]
    headers = {
        normalize_header(worksheet.cell(1, col).value): col
        for col in range(1, worksheet.max_column + 1)
    }
    result_col = find_column(headers, args.result_column, ("binding", "output", "prediction", "final"))
    source_col = headers.get(normalize_header(args.source_column))

    rows: list[dict[str, Any]] = []
    blocked: list[dict[str, Any]] = []
    extraction_reports: list[dict[str, Any]] = []
    for excel_row in range(2, worksheet.max_row + 1):
        source_value = worksheet.cell(excel_row, source_col).value if source_col else None
        result_value = worksheet.cell(excel_row, result_col).value
        if result_value in (None, ""):
            if source_value not in (None, ""):
                blocked.append({"excel_row": excel_row, "case_id": stable_case_id(source_value, excel_row), "reason": "missing_result"})
            continue
        case_id = stable_case_id(source_value, excel_row)
        try:
            items, extracted_records, report = extract_result_items(
                {"case_id": case_id, "result": result_value},
                fallback_case_id=case_id,
                row_number=excel_row,
                strict=True,
            )
        except Exception as exc:
            blocked.append({"excel_row": excel_row, "case_id": case_id, "reason": "invalid_result", "error": str(exc)})
            continue
        row = {
            "case_id": case_id,
            "items": [item_dict(item) for item in items],
            "source_excel_row": excel_row,
            "extracted_records": extracted_records,
        }
        rows.append(row)
        extraction_reports.append(report)
        if args.limit and len(rows) >= args.limit:
            break

    if len(rows) < args.min_rows:
        raise RuntimeError(f"Only {len(rows)} usable rows found; min_rows={args.min_rows}")
    return rows, blocked, extraction_reports


def prediction_rows_from_gold(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    predictions: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        items = list(row["items"])
        if index == 1 and len(items) > 1:
            removable = next((i for i, item in enumerate(items) if item["type"] == "numerical"), len(items) - 1)
            items = [item for i, item in enumerate(items) if i != removable]
        predictions.append({"case_id": row["case_id"], "items": items})
    return predictions


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            clean_row = {"case_id": row["case_id"], "items": row["items"]}
            handle.write(json.dumps(clean_row, ensure_ascii=False, sort_keys=True) + "\n")


def write_metadata(path: Path, payload: dict[str, Any]) -> None:
    metadata_path = path.with_suffix(path.suffix + ".metadata.json")
    metadata_path.parent.mkdir(parents=True, exist_ok=True)
    metadata_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def build(args: argparse.Namespace) -> dict[str, Any]:
    rows, blocked, extraction_reports = read_rows(args)
    matrix_cases = parse_matrix(args.matrix)
    pred_dir = REPO_ROOT / "Experiment" / args.expt_id / "binding_eval_predictions"

    base_gold = args.output_root / "gold.jsonl"
    base_pred = args.output_root / "controlled_predictions.jsonl"
    write_jsonl(base_gold, rows)
    write_jsonl(args.testing_gold_jsonl, rows)
    controlled_metadata = {
        "controlled_smoke": True,
        "source": "gold_scaffold",
        "prediction_strategy": "controlled_smoke_second_row_missing_one_numerical_item",
        "formal_result": False,
    }
    if args.write_controlled_predictions:
        prediction_rows = prediction_rows_from_gold(rows)
        write_jsonl(base_pred, prediction_rows)
        write_metadata(base_pred, controlled_metadata)

    prompt_type_gold_outputs = []
    local_prompt_gold_dir = args.output_root / "prompt_type_gold"
    for narrative_route, path in PROMPT_TYPE_GOLD_PATHS.items():
        local_path = local_prompt_gold_dir / f"{narrative_route}.jsonl"
        write_jsonl(local_path, rows)
        shared_path = None
        if not args.skip_shared_prompt_gold:
            write_jsonl(path, rows)
            shared_path = path
        prompt_type_gold_outputs.append(
            {
                "narrative_route": narrative_route,
                "gold_jsonl": str(local_path if args.skip_shared_prompt_gold else path),
                "local_gold_jsonl": str(local_path),
                "shared_gold_jsonl": str(shared_path) if shared_path else None,
            }
        )

    case_outputs = []
    for case in matrix_cases:
        gold_path = args.gold_dir / f"{case.experiment_id}.jsonl"
        pred_path = pred_dir / f"{case.experiment_id}.jsonl"
        write_jsonl(gold_path, rows)
        if args.write_controlled_predictions:
            write_jsonl(pred_path, prediction_rows)
            write_metadata(pred_path, controlled_metadata)
        case_outputs.append(
            {
                "experiment_id": case.experiment_id,
                "source_id": case.source_id,
                "narrative_route": case.narrative_route,
                "gold_jsonl": str(gold_path),
                "pred_jsonl": str(pred_path) if args.write_controlled_predictions else None,
            }
        )

    report = {
        "source_xlsx": str(args.source_xlsx),
        "sheet": args.sheet,
        "dataset_id": args.dataset_id,
        "rows_written": len(rows),
        "blocked_rows_seen_before_limit": blocked,
        "base_gold_jsonl": str(base_gold),
        "testing_gold_jsonl": str(args.testing_gold_jsonl),
        "prompt_type_gold_jsonl": prompt_type_gold_outputs,
        "base_prediction_jsonl": str(base_pred) if args.write_controlled_predictions else None,
        "controlled_smoke": bool(args.write_controlled_predictions),
        "prediction_strategy": (
            "controlled_smoke_second_row_missing_one_numerical_item"
            if args.write_controlled_predictions
            else "none_formal_gold_only"
        ),
        "formal_result": not args.write_controlled_predictions,
        "experiment_id": args.expt_id,
        "matrix_cases": case_outputs,
        "extraction": {
            "row_reports": extraction_reports,
            "metric_items": sum(len(row["items"]) for row in rows),
            "extraction_records": sum(len(row["extracted_records"]) for row in rows),
        },
    }
    report_path = args.output_root / "build_report.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-xlsx", type=Path, default=DEFAULT_SOURCE_XLSX)
    parser.add_argument("--sheet", default="label")
    parser.add_argument("--result-column", default="result")
    parser.add_argument("--source-column", default="Source")
    parser.add_argument("--dataset-id", default="few_1_epoch")
    parser.add_argument("--limit", type=int, default=0, help="Maximum usable rows to write; 0 means no limit.")
    parser.add_argument("--min-rows", type=int, default=2)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--gold-dir", type=Path, default=DEFAULT_GOLD_DIR)
    parser.add_argument("--testing-gold-jsonl", type=Path, default=DEFAULT_TESTING_GOLD_JSONL)
    parser.add_argument("--expt-id", default=DEFAULT_EXPT_ID)
    parser.add_argument("--matrix", default=DEFAULT_MATRIX)
    parser.add_argument("--write-controlled-predictions", action="store_true")
    parser.add_argument("--skip-shared-prompt-gold", action="store_true")
    return parser.parse_args()


def main() -> None:
    report = build(parse_args())
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
