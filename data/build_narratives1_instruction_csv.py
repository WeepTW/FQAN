#!/usr/bin/env python3
from __future__ import annotations

import argparse
import ast
import csv
from concurrent.futures import ThreadPoolExecutor, as_completed
import importlib.util
import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

csv.field_size_limit(10**9)

WORKSPACE_ROOT = Path(__file__).resolve().parents[1]
REQUIRED_WORKBOOK_COLUMNS = ("Source", "data", "text", "result", "reason")
ORIGINAL_EXAMPLE_WORKBOOK_COLUMNS = ("data", "text", "result", "reason")
CORE_COLUMNS = [
    "",
    "Sentence",
    "input",
    "Question",
    "GT_Answer",
    "GT_Program",
    "Pre_Text",
    "Post_Text",
    "Tables",
    "Table_Text",
    "Rel_Fact",
]
EXTRA_COLUMNS = [
    "Source",
    "Narrative_Data",
    "Narrative_Text",
    "Binding_Result",
    "Binding_Reason",
    "Prompt_Mode",
    "Generator_Model",
]
FIELDNAMES = CORE_COLUMNS + EXTRA_COLUMNS
PROMPT_MODES = ("original", "zero_shot", "many_shot", "dynamic_shot")
PLACEHOLDERS = ("{Sentences}", "{Question}", "{examples_by_prompt_type}", "`prompt-type`")
BARE_PERCENT_VALUE_RE = re.compile(r"(:\s*)([-+]?\d+(?:\.\d+)?)%(?=\s*[,}\]])")

FINFLIER_SYSTEM_PROMPT = """Please think as an economic data analyst.
You need to perform financial data-text binding: match the narrative text to the provided tabular data, identify subjects, data names, positions, trends, numerical values, and the exact evidence text.
Return only a valid JSON object with keys "result" and "reason".
"result" must be a list of objects using this schema:
{"ObjectName":[],"DataName":"","Position":[{"Begin":[],"End":[]}],"Trend":"None","Num":[],"Text":""}
Do not add markdown, explanations outside JSON, or unsupported fields."""


def resolve_path(path: Path) -> Path:
    if path.is_absolute():
        return path
    return WORKSPACE_ROOT / path


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"\s+", " ", text).strip()


def compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def quote_bare_percent_values(text: str) -> str:
    return BARE_PERCENT_VALUE_RE.sub(lambda match: f'{match.group(1)}"{match.group(2)}%"', text)


def parse_jsonish(raw: Any, label: str, repairs: list[dict[str, str]] | None = None) -> Any:
    text = normalize_text(raw)
    if not text:
        raise ValueError(f"{label} is empty")
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            return ast.literal_eval(text)
        except (SyntaxError, ValueError) as exc:
            repaired = quote_bare_percent_values(text)
            if repaired != text:
                try:
                    parsed = json.loads(repaired)
                    if repairs is not None:
                        repairs.append({"label": label, "repair": "quoted_bare_percent_value"})
                    return parsed
                except json.JSONDecodeError:
                    try:
                        parsed = ast.literal_eval(repaired)
                        if repairs is not None:
                            repairs.append({"label": label, "repair": "quoted_bare_percent_value"})
                        return parsed
                    except (SyntaxError, ValueError):
                        pass
            raise ValueError(f"{label} is not parseable JSON/Python literal") from exc


def load_convert_module() -> Any:
    module_path = WORKSPACE_ROOT / "data" / "convert_instruction_shots.py"
    spec = importlib.util.spec_from_file_location("convert_instruction_shots", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot import {module_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_workbook_records(path: Path, sheet: str, expected_rows: int | None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        raise ValueError(f"{path} has no sheet named {sheet!r}; sheets={workbook.sheetnames}")
    worksheet = workbook[sheet]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {normalize_text(value): index for index, value in enumerate(header_row) if normalize_text(value)}
    missing = [column for column in REQUIRED_WORKBOOK_COLUMNS if column not in header_map]
    if missing:
        raise ValueError(f"{path}:{sheet} missing columns: {missing}")

    complete_rows: list[dict[str, Any]] = []
    blocked_rows: list[dict[str, Any]] = []
    for excel_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        values = {
            column: row[header_map[column]] if header_map[column] < len(row) else None
            for column in REQUIRED_WORKBOOK_COLUMNS
        }
        missing_values = [column for column, value in values.items() if normalize_text(value) == ""]
        if missing_values:
            blocked_rows.append({"excel_row": excel_row, "missing": missing_values})
            continue
        complete_rows.append({"excel_row": excel_row, **values})

    if blocked_rows:
        raise ValueError(f"Workbook has incomplete required rows: {blocked_rows[:10]}")
    if expected_rows is not None and len(complete_rows) != expected_rows:
        raise ValueError(f"Workbook expected {expected_rows} complete rows, found {len(complete_rows)}")
    return complete_rows


def load_original_example_records(
    path: Path,
    sheet: str | None,
    count: int,
) -> tuple[list[dict[str, str]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = sheet or workbook.sheetnames[0]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{path} has no sheet named {sheet_name!r}; sheets={workbook.sheetnames}")
    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {normalize_text(value): index for index, value in enumerate(header_row) if normalize_text(value)}
    missing = [column for column in ORIGINAL_EXAMPLE_WORKBOOK_COLUMNS if column not in header_map]
    if missing:
        raise ValueError(f"{path}:{sheet_name} missing original example columns: {missing}")

    selected: list[dict[str, str]] = []
    skipped_rows: list[dict[str, Any]] = []
    rows_seen = 0
    for excel_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        rows_seen += 1
        values = {
            column: row[header_map[column]] if header_map[column] < len(row) else None
            for column in ORIGINAL_EXAMPLE_WORKBOOK_COLUMNS
        }
        missing_values = [column for column, value in values.items() if normalize_text(value) == ""]
        if missing_values:
            skipped_rows.append({"excel_row": excel_row, "reason": "missing_values", "columns": missing_values})
            continue
        try:
            parsed_result = parse_jsonish(values["result"], f"{path}:{sheet_name}:row{excel_row}:result")
        except ValueError as exc:
            skipped_rows.append({"excel_row": excel_row, "reason": "unparseable_result", "error": str(exc)})
            continue
        if not isinstance(parsed_result, list):
            skipped_rows.append({"excel_row": excel_row, "reason": "result_is_not_list"})
            continue
        selected.append(
            {
                "excel_row": str(excel_row),
                "data": normalize_text(values["data"]),
                "text": normalize_text(values["text"]),
                "result": compact_json(parsed_result),
                "reason": normalize_text(values["reason"]),
            }
        )
        if len(selected) >= count:
            break

    if len(selected) < count:
        raise ValueError(
            f"{path}:{sheet_name} has only {len(selected)} parseable original examples; required {count}"
        )
    return selected, {
        "path": str(path),
        "sheet": sheet_name,
        "rows_seen_until_selected": rows_seen,
        "rows_selected": len(selected),
        "selected_excel_rows": [int(row["excel_row"]) for row in selected],
        "skipped_rows_before_selection": skipped_rows,
    }


def records_to_table(data_records: Any) -> list[list[str]]:
    if isinstance(data_records, dict):
        data_records = [data_records]
    if not isinstance(data_records, list) or not all(isinstance(row, dict) for row in data_records):
        raise ValueError("data must be a JSON object or list of JSON objects")

    headers: list[str] = []
    for record in data_records:
        for key in record.keys():
            key_text = normalize_text(key)
            if key_text and key_text not in headers:
                headers.append(key_text)
    if not headers:
        raise ValueError("data records do not contain any columns")

    table = [headers]
    for record in data_records:
        table.append([normalize_text(record.get(header, "")) for header in headers])
    return table


def build_table_text(table: list[list[str]]) -> str:
    if len(table) < 2 or len(table[0]) < 2:
        return ""
    headers = table[0]
    row_header = headers[0]
    segments: list[str] = []
    for table_row in table[1:]:
        row_label = table_row[0] if table_row else ""
        if not row_label:
            continue
        for column_index, column_name in enumerate(headers[1:], start=1):
            value = table_row[column_index] if column_index < len(table_row) else ""
            if value == "":
                continue
            subject = row_label if row_header else f"row {len(segments) + 1}"
            segments.append(f"the {column_name} of {subject} is {value}")
    return " ; ".join(segments) + (" ;" if segments else "")


def build_base_rows(workbook_rows: list[dict[str, Any]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    base_rows: list[dict[str, str]] = []
    parse_repairs: list[dict[str, str]] = []
    seen_sources: set[str] = set()
    duplicate_sources: list[str] = []

    for row_index, workbook_row in enumerate(workbook_rows):
        source = normalize_text(workbook_row["Source"])
        if source in seen_sources:
            duplicate_sources.append(source)
        seen_sources.add(source)

        data_records = parse_jsonish(workbook_row["data"], f"{source}.data", parse_repairs)
        binding_result = parse_jsonish(workbook_row["result"], f"{source}.result", parse_repairs)
        if not isinstance(binding_result, list):
            raise ValueError(f"{source}.result must parse to a list")
        table = records_to_table(data_records)
        table_text = build_table_text(table)
        narrative_text = normalize_text(workbook_row["text"])
        reason = normalize_text(workbook_row["reason"])
        rel_fact = {
            "RetFact": narrative_text,
            "Binding": binding_result,
            "Reason": reason,
        }
        base_rows.append(
            {
                "": str(row_index),
                "Sentence": narrative_text,
                "input": "",
                "Question": "What data-text bindings are described in the narrative?",
                "GT_Answer": "",
                "GT_Program": "",
                "Pre_Text": narrative_text,
                "Post_Text": "",
                "Tables": repr(table),
                "Table_Text": table_text,
                "Rel_Fact": compact_json(rel_fact),
                "Source": source,
                "Narrative_Data": compact_json(data_records),
                "Narrative_Text": narrative_text,
                "Binding_Result": compact_json(binding_result),
                "Binding_Reason": reason,
                "Prompt_Mode": "",
                "Generator_Model": "",
            }
        )

    if duplicate_sources:
        raise ValueError(f"Duplicate Source values found: {duplicate_sources[:10]}")
    return base_rows, parse_repairs


def build_original_user_content(data_text: str, narrative_text: str) -> str:
    return f"data:{data_text}\ntext:[{narrative_text}]"


def build_original_example_messages(example_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    messages: list[dict[str, str]] = []
    for example_row in example_rows:
        messages.append(
            {
                "role": "system",
                "name": "example_user",
                "content": build_original_user_content(example_row["data"], example_row["text"]),
            }
        )
        messages.append(
            {
                "role": "system",
                "name": "example_assistant",
                "content": f"result: {example_row['result']}\nreason: {example_row['reason']}",
            }
        )
    return messages


def build_original_messages(convert: Any, row: dict[str, str], example_messages: list[dict[str, str]]) -> list[dict[str, str]]:
    data_text = convert.build_data_records_text(row)
    context = convert.build_context(row)
    return [
        {"role": "system", "content": FINFLIER_SYSTEM_PROMPT},
        *example_messages,
        {"role": "user", "content": build_original_user_content(data_text, context)},
    ]


def build_original_rows(
    convert: Any,
    base_rows: list[dict[str, str]],
    example_rows: list[dict[str, str]],
    model: str,
) -> list[dict[str, str]]:
    example_messages = build_original_example_messages(example_rows)
    output_rows = []
    for row in base_rows:
        output_row = dict(row)
        output_row["Prompt_Mode"] = "original"
        output_row["Generator_Model"] = model
        output_row["input"] = compact_json(build_original_messages(convert, row, example_messages))
        output_rows.append(output_row)
    return output_rows


def build_non_original_rows(
    convert: Any,
    base_rows: list[dict[str, str]],
    template: str,
    rendered_examples: list[str],
    example_vectors: list[dict[str, float]],
    dfs: dict[str, int],
    dynamic_rendered_examples: list[str],
    dynamic_example_vectors: list[dict[str, float]],
    dynamic_dfs: dict[str, int],
    seed: int,
    many_count: int,
    dynamic_count: int,
) -> dict[str, list[dict[str, str]]]:
    outputs = {"zero_shot": [], "many_shot": [], "dynamic_shot": []}
    source_name = "narratives1_rel_fact_instruction.csv"
    document_count = max(len(dynamic_rendered_examples), 1)
    for row_index, row in enumerate(base_rows):
        context = convert.build_context(row)
        question = convert.normalize(row.get("Question"))
        query_text = f"{context} {question}"
        data_text = convert.build_data_records_text(row)

        zero_row = dict(row)
        zero_row["Prompt_Mode"] = "zero_shot"
        zero_row["input"] = convert.build_prompt(template, data_text, context, question, "")
        outputs["zero_shot"].append(zero_row)

        many_row = dict(row)
        many_row["Prompt_Mode"] = "many_shot"
        many_examples = convert.select_many_examples(rendered_examples, seed, source_name, row_index, many_count)
        many_row["input"] = convert.build_prompt(template, data_text, context, question, many_examples)
        outputs["many_shot"].append(many_row)

        dynamic_row = dict(row)
        dynamic_row["Prompt_Mode"] = "dynamic_shot"
        dynamic_examples = convert.select_dynamic_examples(
            query_text,
            dynamic_rendered_examples,
            dynamic_example_vectors,
            dynamic_dfs,
            document_count,
            dynamic_count,
        )
        dynamic_row["input"] = convert.build_prompt(template, data_text, context, question, dynamic_examples)
        outputs["dynamic_shot"].append(dynamic_row)
    return outputs


def extract_json_payload(text: str) -> Any:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?\s*", "", stripped, flags=re.IGNORECASE)
        stripped = re.sub(r"\s*```$", "", stripped)
    stripped = re.sub(r"<think>.*?</think>", "", stripped, flags=re.IGNORECASE | re.DOTALL).strip()
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        first_object = stripped.find("{")
        last_object = stripped.rfind("}")
        first_array = stripped.find("[")
        last_array = stripped.rfind("]")
        candidates: list[str] = []
        if first_object >= 0 and last_object > first_object:
            candidates.append(stripped[first_object : last_object + 1])
        if first_array >= 0 and last_array > first_array:
            candidates.append(stripped[first_array : last_array + 1])
        for candidate in candidates:
            try:
                return json.loads(candidate)
            except json.JSONDecodeError:
                try:
                    return ast.literal_eval(candidate)
                except (SyntaxError, ValueError):
                    continue
    raise ValueError("response does not contain parseable JSON")


def parse_chatmock_response(text: str) -> tuple[bool, str, str, str | None]:
    try:
        payload = extract_json_payload(text)
    except ValueError as exc:
        return False, text, "", str(exc)
    if isinstance(payload, dict):
        if isinstance(payload.get("result"), list):
            return True, compact_json(payload["result"]), normalize_text(payload.get("reason", "")), None
        if isinstance(payload.get("Binding"), list):
            return True, compact_json(payload["Binding"]), normalize_text(payload.get("Reason", "")), None
        return False, compact_json(payload), "", "JSON object has neither result nor Binding list"
    if isinstance(payload, list):
        return True, compact_json(payload), "", None
    return False, compact_json(payload), "", "JSON payload is not an object or list"


def run_chatmock(
    rows: list[dict[str, str]],
    base_url: str,
    api_key: str,
    model: str,
    temperature: float,
    max_tokens: int,
    workers: int,
    timeout: float,
) -> dict[str, Any]:
    from openai import OpenAI

    parse_mismatches: list[dict[str, Any]] = []
    started = time.time()
    worker_count = max(1, min(workers, len(rows) or 1))

    def call_row(row_index: int, row: dict[str, str]) -> tuple[int, str]:
        client = OpenAI(base_url=base_url, api_key=api_key, timeout=timeout)
        messages = json.loads(row["input"])
        try:
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
                max_tokens=max_tokens,
            )
        except Exception as exc:  # noqa: BLE001 - preserve upstream ChatMock error text in report.
            raise RuntimeError(f"ChatMock call failed at row {row_index} Source={row['Source']}: {exc}") from exc
        return row_index, response.choices[0].message.content or ""

    completed = 0
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = [executor.submit(call_row, row_index, row) for row_index, row in enumerate(rows)]
        for future in as_completed(futures):
            row_index, content = future.result()
            row = rows[row_index]
            ok, binding_result, binding_reason, parse_error = parse_chatmock_response(content)
            row["Binding_Result"] = binding_result
            if binding_reason:
                row["Binding_Reason"] = binding_reason
            if not ok:
                parse_mismatches.append(
                    {
                        "row_index": row_index,
                        "Source": row["Source"],
                        "error": parse_error,
                        "response_preview": content[:500],
                    }
                )
            completed += 1
            if completed % 10 == 0 or completed == len(rows):
                print(f"chatmock_completed={completed}/{len(rows)}", flush=True)

    return {
        "executed": True,
        "base_url": base_url,
        "model": model,
        "temperature": temperature,
        "max_tokens": max_tokens,
        "workers": worker_count,
        "timeout": timeout,
        "rows_requested": len(rows),
        "rows_completed": len(rows),
        "parse_ok": len(rows) - len(parse_mismatches),
        "parse_mismatch_count": len(parse_mismatches),
        "parse_mismatches": sorted(parse_mismatches, key=lambda item: item["row_index"]),
        "elapsed_seconds": round(time.time() - started, 3),
    }


def write_csv_atomic(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.tmp")
    with tmp_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in FIELDNAMES})
    tmp_path.replace(path)


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def count_examples(input_text: str) -> int:
    match = re.search(r"### Example\s*\n(.*?)\n###", input_text, flags=re.DOTALL)
    if not match:
        return 0
    block = match.group(1).strip()
    if not block:
        return 0
    return len(re.split(r"\n---\n", block))


def validate_generated_rows(rows_by_mode: dict[str, list[dict[str, str]]], expected_rows: int) -> dict[str, Any]:
    validation: dict[str, Any] = {}
    for mode, rows in rows_by_mode.items():
        if len(rows) != expected_rows:
            raise AssertionError(f"{mode} expected {expected_rows} rows, found {len(rows)}")
        sources = [row.get("Source", "") for row in rows]
        if len(set(sources)) != len(sources):
            raise AssertionError(f"{mode} contains duplicated Source values")

        for row_index, row in enumerate(rows):
            missing_columns = [column for column in FIELDNAMES if column not in row]
            if missing_columns:
                raise AssertionError(f"{mode} row {row_index} missing columns: {missing_columns}")
            rel_fact = json.loads(row["Rel_Fact"])
            for key in ("RetFact", "Binding", "Reason"):
                if key not in rel_fact:
                    raise AssertionError(f"{mode} row {row_index} Rel_Fact missing {key}")

        if mode == "original":
            messages = json.loads(rows[0]["input"])
            names = [message.get("name") for message in messages]
            if "example_user" not in names or "example_assistant" not in names:
                raise AssertionError("original input does not include example_user/example_assistant messages")
            if messages[-1].get("role") != "user":
                raise AssertionError("original input final message is not user")
            validation[mode] = {
                "rows": len(rows),
                "messages_in_sample": len(messages),
                "has_example_user": True,
                "has_example_assistant": True,
            }
            continue

        for placeholder in PLACEHOLDERS:
            if any(placeholder in row["input"] for row in rows):
                raise AssertionError(f"{mode} input still contains placeholder {placeholder}")
        validation[mode] = {"rows": len(rows), "sample_example_count": count_examples(rows[0]["input"])}
    return validation


def write_report(path: Path, report: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_prompt_modes(raw: str) -> tuple[str, ...]:
    modes = tuple(part.strip() for part in raw.split(",") if part.strip())
    invalid_modes = [mode for mode in modes if mode not in PROMPT_MODES]
    if invalid_modes:
        raise ValueError(f"Unsupported prompt modes: {invalid_modes}; supported={PROMPT_MODES}")
    return modes


def build_outputs(args: argparse.Namespace) -> dict[str, Any]:
    source_xlsx = resolve_path(args.source_xlsx)
    output_root = resolve_path(args.output_root)
    prompt_path = resolve_path(args.prompt)
    examples_path = resolve_path(args.examples)
    original_examples_path = resolve_path(args.original_examples_xlsx)
    dynamic_examples_path = resolve_path(args.dynamic_examples_xlsx)
    report_path = resolve_path(args.report)
    requested_modes = parse_prompt_modes(args.prompt_modes)

    convert = load_convert_module()
    workbook_rows = load_workbook_records(source_xlsx, args.sheet, args.expected_rows)
    base_rows, parse_repairs = build_base_rows(workbook_rows)
    template = convert.extract_instruction_template(prompt_path)
    example_rows, rendered_examples, example_queries = convert.load_examples(examples_path)
    dynamic_example_rows, dynamic_rendered_examples, dynamic_example_queries, dynamic_example_report = convert.load_finflier_xlsx_examples(
        dynamic_examples_path,
        args.dynamic_examples_sheet or None,
    )
    original_example_rows: list[dict[str, str]] = []
    original_example_report: dict[str, Any] = {"skipped": "original_not_requested"}
    if "original" in requested_modes:
        original_example_rows, original_example_report = load_original_example_records(
            original_examples_path,
            args.original_examples_sheet or None,
            args.original_examples,
        )
    dfs = convert.build_document_frequencies(example_queries)
    example_vectors = [convert.vectorize(query, dfs, max(len(example_queries), 1)) for query in example_queries]
    dynamic_dfs = convert.build_document_frequencies(dynamic_example_queries)
    dynamic_example_vectors = [
        convert.vectorize(query, dynamic_dfs, max(len(dynamic_example_queries), 1))
        for query in dynamic_example_queries
    ]

    original_rows: list[dict[str, str]] = []
    chatmock_report: dict[str, Any]
    if "original" in requested_modes:
        original_rows = build_original_rows(convert, base_rows, original_example_rows, args.chatmock_model)
    if "original" in requested_modes and args.execute_original_chatmock:
        chatmock_report = run_chatmock(
            original_rows,
            args.chatmock_base_url,
            args.chatmock_api_key,
            args.chatmock_model,
            args.temperature,
            args.max_tokens,
            args.chatmock_workers,
            args.chatmock_timeout,
        )
    else:
        chatmock_report = {
            "executed": False,
            "blocker": "blocked_chatmock_not_executed" if "original" in requested_modes else "skipped_original_not_requested",
            "model": args.chatmock_model,
        }

    non_original = build_non_original_rows(
        convert,
        base_rows,
        template,
        rendered_examples,
        example_vectors,
        dfs,
        dynamic_rendered_examples,
        dynamic_example_vectors,
        dynamic_dfs,
        args.seed,
        args.many_count,
        args.dynamic_count,
    )
    rows_by_mode = {mode: rows for mode, rows in non_original.items() if mode in requested_modes}
    if "original" in requested_modes:
        rows_by_mode["original"] = original_rows
    validation = validate_generated_rows(rows_by_mode, len(base_rows))

    output_paths = {
        "original": output_root / "finqa_original" / "narratives1_rel_fact_instruction.csv",
        "zero_shot": output_root / "finqa_zero_shot" / "narratives1_rel_fact_instruction.csv",
        "many_shot": output_root / "finqa_many_shot" / "narratives1_rel_fact_instruction.csv",
        "dynamic_shot": output_root / "finqa_dynamic_shot" / "narratives1_rel_fact_instruction.csv",
    }
    selected_output_paths = {mode: path for mode, path in output_paths.items() if mode in requested_modes}
    for mode, path in selected_output_paths.items():
        write_csv_atomic(path, rows_by_mode[mode])

    disk_validation = {}
    for mode, path in selected_output_paths.items():
        disk_rows = read_csv_rows(path)
        if len(disk_rows) != len(base_rows):
            raise AssertionError(f"{path} expected {len(base_rows)} rows, found {len(disk_rows)} after write")
        disk_validation[mode] = {"path": str(path), "rows": len(disk_rows), "columns": len(disk_rows[0]) if disk_rows else 0}

    status = "completed"
    blockers: list[str] = []
    if chatmock_report.get("parse_mismatch_count", 0):
        status = "completed_with_chatmock_parse_warnings"
        blockers.append("chatmock_parse_mismatch_rows")

    report = {
        "time": datetime.now(timezone.utc).isoformat(),
        "status": status,
        "source_xlsx": str(source_xlsx),
        "sheet": args.sheet,
        "source_rows": len(base_rows),
        "required_columns": list(REQUIRED_WORKBOOK_COLUMNS),
        "requested_prompt_modes": list(requested_modes),
        "examples_csv_for_non_original_modes": str(examples_path),
        "dynamic_examples_source": dynamic_example_report,
        "original_examples_source": original_example_report,
        "source_parse_repairs": parse_repairs,
        "outputs": {mode: str(path) for mode, path in selected_output_paths.items()},
        "validation": validation,
        "disk_validation": disk_validation,
        "chatmock": chatmock_report,
        "seed": args.seed,
        "many_count": args.many_count,
        "dynamic_count": args.dynamic_count,
        "original_examples": args.original_examples,
        "blockers": blockers,
    }
    write_report(report_path, report)
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Build narratives1 prompt-mode instruction CSVs.")
    parser.add_argument("--source-xlsx", type=Path, default=Path("data/src/narratives/narratives1.xlsx"))
    parser.add_argument("--sheet", default="label")
    parser.add_argument("--output-root", type=Path, default=Path("data"))
    parser.add_argument("--prompt", type=Path, default=Path("src/doc/new_prompt.txt"))
    parser.add_argument("--examples", type=Path, default=Path("data/src/full_example.csv"))
    parser.add_argument("--original-examples-xlsx", type=Path, default=Path("data/src/FinFlier/1_full_used_data.xlsx"))
    parser.add_argument("--original-examples-sheet", default="")
    parser.add_argument("--dynamic-examples-xlsx", type=Path, default=Path("data/src/FinFlier/1_full_used_data.xlsx"))
    parser.add_argument("--dynamic-examples-sheet", default="")
    parser.add_argument("--prompt-modes", default=",".join(PROMPT_MODES))
    parser.add_argument("--report", type=Path, default=Path("data/finqa_narratives1_generation_report.json"))
    parser.add_argument("--expected-rows", type=int, default=85)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--many-count", type=int, default=26)
    parser.add_argument("--dynamic-count", type=int, default=10)
    parser.add_argument("--original-examples", type=int, default=3)
    parser.add_argument("--execute-original-chatmock", action="store_true")
    parser.add_argument("--chatmock-base-url", default=os.environ.get("CHATMOCK_BASE_URL", "${SERVICE_ENDPOINT}"))
    parser.add_argument("--chatmock-api-key", default=os.environ.get("CHATMOCK_API_KEY", "key"))
    parser.add_argument("--chatmock-model", default=os.environ.get("CHATMOCK_GPT5_5_MODEL", "gpt-5.5"))
    parser.add_argument("--temperature", type=float, default=0.2)
    parser.add_argument("--max-tokens", type=int, default=4000)
    parser.add_argument("--chatmock-workers", type=int, default=1)
    parser.add_argument("--chatmock-timeout", type=float, default=300.0)
    args = parser.parse_args()

    report_path = resolve_path(args.report)
    try:
        report = build_outputs(args)
    except Exception as exc:  # noqa: BLE001 - write a sidecar blocker before surfacing the failure.
        blocker_report = {
            "time": datetime.now(timezone.utc).isoformat(),
            "status": "blocked_runtime_error",
            "source_xlsx": str(resolve_path(args.source_xlsx)),
            "sheet": args.sheet,
            "chatmock": {
                "executed": bool(args.execute_original_chatmock),
                "base_url": args.chatmock_base_url,
                "model": args.chatmock_model,
                "workers": args.chatmock_workers,
                "timeout": args.chatmock_timeout,
            },
            "blockers": [str(exc)],
        }
        write_report(report_path, blocker_report)
        raise

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
