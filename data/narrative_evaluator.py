#!/usr/bin/env python3
"""Evaluate FinFlier-style text-data binding outputs.

The main Table I setting compares identified subject, trend, and numerical
vocabularies with labeled vocabularies and reports Precision, Recall, and F1.
This script supports two local gold sources:

1. FinFlier example sheets whose ``result`` cells contain binding objects.
2. The corpus-statistics label workbook, interpreted as per-case vocabulary
   presence labels because it stores category codes rather than verbatim text.
"""

from __future__ import annotations

import argparse
import ast
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import openpyxl


VOCAB_TYPES = ("subject", "trend", "numerical")
NONE_VALUES = {"", "none", "null", "nan", "n/a", "na", "n", "no"}
NUMBER_RE = re.compile(r"[-+]?\d[\d,]*(?:\.\d+)?%?")


@dataclass
class BindingLabels:
    subject: set[str] = field(default_factory=set)
    trend: set[str] = field(default_factory=set)
    numerical: set[str] = field(default_factory=set)

    def as_dict(self) -> dict[str, set[str]]:
        return {
            "subject": self.subject,
            "trend": self.trend,
            "numerical": self.numerical,
        }

    def add_presence(self, vocab_type: str) -> None:
        self.as_dict()[vocab_type].add("__present__")


def normalize_text(value: Any) -> str:
    text = str(value).strip().lower()
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = re.sub(r"\s+", " ", text)
    text = text.strip(" \t\r\n\"'`.,;:()[]{}")
    return text


def normalize_number(value: Any) -> str:
    text = normalize_text(value)
    match = NUMBER_RE.search(text)
    if not match:
        return text
    number = match.group(0).replace(",", "")
    if number.endswith("%"):
        number = number[:-1]
    if "." in number:
        number = number.rstrip("0").rstrip(".")
    return number


def is_none_like(value: Any) -> bool:
    return normalize_text(value) in NONE_VALUES


def split_csv_like(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[,;/|]", value) if part.strip()]


def add_value(labels: BindingLabels, vocab_type: str, value: Any) -> None:
    if value is None or is_none_like(value):
        return

    values: Iterable[Any]
    if isinstance(value, (list, tuple, set)):
        values = value
    else:
        values = split_csv_like(str(value))

    target = labels.as_dict()[vocab_type]
    for item in values:
        if item is None or is_none_like(item):
            continue
        if vocab_type == "numerical":
            normalized = normalize_number(item)
        else:
            normalized = normalize_text(item)
        if normalized and not is_none_like(normalized):
            target.add(normalized)


def coerce_jsonish(value: str) -> Any | None:
    text = value.strip()
    if not text:
        return None
    text = (
        text.replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u2018", "'")
        .replace("\u2019", "'")
    )
    for loader in (json.loads, ast.literal_eval):
        try:
            return loader(text)
        except Exception:
            pass
    return None


def walk_objects(value: Any) -> Iterable[dict[str, Any]]:
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from walk_objects(child)
    elif isinstance(value, list):
        for item in value:
            yield from walk_objects(item)


def labels_from_structured_binding(value: Any) -> BindingLabels:
    labels = BindingLabels()
    for obj in walk_objects(value):
        if isinstance(obj.get("Type"), str) and "Text" in obj:
            item_type = normalize_text(obj["Type"])
            if item_type == "objectname":
                add_value(labels, "subject", obj["Text"])
            elif item_type == "trend":
                add_value(labels, "trend", obj["Text"])
            elif item_type == "num":
                add_value(labels, "numerical", obj["Text"])
        for key, item in obj.items():
            key_norm = normalize_text(key).replace("_", "")
            if key_norm in {"objectname", "subject", "highlight"}:
                add_value(labels, "subject", item)
            elif key_norm in {"trend", "trendword"}:
                add_value(labels, "trend", item)
            elif key_norm in {"num", "number", "numerical", "label"}:
                add_value(labels, "numerical", item)
    return labels


def regex_extract_field(text: str, field_name: str) -> list[str]:
    quoted = r'"([^"]+)"|\'([^\']+)\'|([^,\]\}\n]+)'
    array_pattern = re.compile(
        rf'["\']?{re.escape(field_name)}["\']?\s*:\s*\[(.*?)\]',
        re.IGNORECASE | re.DOTALL,
    )
    scalar_pattern = re.compile(
        rf'["\']?{re.escape(field_name)}["\']?\s*:\s*(?:{quoted})',
        re.IGNORECASE | re.DOTALL,
    )

    values: list[str] = []
    for match in array_pattern.finditer(text):
        values.extend(
            item[0] or item[1] or item[2]
            for item in re.findall(quoted, match.group(1))
        )
    for match in scalar_pattern.finditer(text):
        values.append(match.group(1) or match.group(2) or match.group(3))
    return values


def labels_from_binding_cell(value: Any) -> BindingLabels:
    if value is None:
        return BindingLabels()

    parsed = coerce_jsonish(str(value))
    labels = labels_from_structured_binding(parsed) if parsed is not None else BindingLabels()
    if any(labels.as_dict()[name] for name in VOCAB_TYPES):
        return labels

    text = str(value)
    for item in regex_extract_field(text, "ObjectName"):
        add_value(labels, "subject", item)
    for item in regex_extract_field(text, "Subject"):
        add_value(labels, "subject", item)
    for item in regex_extract_field(text, "Trend"):
        add_value(labels, "trend", item)
    for item in regex_extract_field(text, "Num"):
        add_value(labels, "numerical", item)
    for item in regex_extract_field(text, "Numerical"):
        add_value(labels, "numerical", item)
    return labels


def merged_header(ws: openpyxl.worksheet.worksheet.Worksheet, col: int) -> str:
    parts = []
    for row in (1, 2):
        value = ws.cell(row, col).value
        if value is not None:
            parts.append(str(value))
    return " ".join(parts)


def find_column(ws: openpyxl.worksheet.worksheet.Worksheet, preferred: str | None) -> int:
    headers = {normalize_text(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    if preferred:
        key = normalize_text(preferred)
        if key in headers:
            return headers[key]
        raise ValueError(f"Column {preferred!r} not found in sheet {ws.title!r}")
    for candidate in ("result", "binding", "output", "prediction", "final"):
        if candidate in headers:
            return headers[candidate]
    raise ValueError(f"No binding/result column found in sheet {ws.title!r}")


def load_binding_xlsx(path: Path, column: str | None = None, sheet: str | None = None) -> list[BindingLabels]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    target_col = find_column(ws, column)
    labels = []
    for row in range(2, ws.max_row + 1):
        labels.append(labels_from_binding_cell(ws.cell(row, target_col).value))
    return labels


def load_prediction_json(path: Path) -> list[BindingLabels]:
    raw = path.read_text(encoding="utf-8")
    rows: list[Any]
    try:
        parsed = json.loads(raw)
        rows = parsed if isinstance(parsed, list) else [parsed]
    except json.JSONDecodeError:
        rows = [json.loads(line) for line in raw.splitlines() if line.strip()]

    labels = []
    for row in rows:
        if isinstance(row, dict):
            for key in ("binding", "final", "result", "output", "prediction"):
                if key in row and row[key] not in (None, ""):
                    payload = row[key]
                    if key == "final" and isinstance(payload, list):
                        labels.extend(labels_from_binding_cell(item) for item in payload)
                    else:
                        labels.append(labels_from_binding_cell(payload))
                    break
            else:
                labels.append(labels_from_binding_cell(row))
        else:
            labels.append(labels_from_binding_cell(row))
    return labels


def load_finance_presence_xlsx(path: Path) -> tuple[list[BindingLabels], dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    labels: list[BindingLabels] = []
    sheet_counts: dict[str, int] = {}

    for ws in wb.worksheets:
        count = 0
        headers = {col: normalize_text(merged_header(ws, col)) for col in range(1, ws.max_column + 1)}
        for row in range(3, ws.max_row + 1):
            current = BindingLabels()
            for col, header in headers.items():
                value = ws.cell(row, col).value
                if value is None or is_none_like(value):
                    continue
                if "subject" in header:
                    current.add_presence("subject")
                elif "numerical" in header:
                    current.add_presence("numerical")
                elif "trend" in header:
                    current.add_presence("trend")
            labels.append(current)
            count += 1
        sheet_counts[ws.title] = count

    non_empty = sum(
        1
        for label in labels
        if label.subject or label.trend or label.numerical
    )
    return labels, {
        "path": str(path),
        "sheets": sheet_counts,
        "rows": len(labels),
        "rows_with_any_label": non_empty,
    }


def safe_divide(numerator: int, denominator: int) -> float:
    return numerator / denominator if denominator else 0.0


def to_presence_rows(rows: list[BindingLabels]) -> list[BindingLabels]:
    presence_rows = []
    for row in rows:
        current = BindingLabels()
        for vocab_type in VOCAB_TYPES:
            if row.as_dict()[vocab_type]:
                current.add_presence(vocab_type)
        presence_rows.append(current)
    return presence_rows


def score_rows(gold_rows: list[BindingLabels], pred_rows: list[BindingLabels]) -> dict[str, Any]:
    totals = {name: {"tp": 0, "fp": 0, "fn": 0} for name in (*VOCAB_TYPES, "overall")}
    max_rows = max(len(gold_rows), len(pred_rows))

    for index in range(max_rows):
        gold = gold_rows[index] if index < len(gold_rows) else BindingLabels()
        pred = pred_rows[index] if index < len(pred_rows) else BindingLabels()
        for vocab_type in VOCAB_TYPES:
            gold_set = gold.as_dict()[vocab_type]
            pred_set = pred.as_dict()[vocab_type]
            tp = len(gold_set & pred_set)
            fp = len(pred_set - gold_set)
            fn = len(gold_set - pred_set)
            for key, value in (("tp", tp), ("fp", fp), ("fn", fn)):
                totals[vocab_type][key] += value
                totals["overall"][key] += value

    metrics: dict[str, Any] = {}
    for name, counts in totals.items():
        precision = safe_divide(counts["tp"], counts["tp"] + counts["fp"])
        recall = safe_divide(counts["tp"], counts["tp"] + counts["fn"])
        f1 = safe_divide(2 * precision * recall, precision + recall)
        metrics[name] = {
            **counts,
            "precision": round(precision, 6),
            "recall": round(recall, 6),
            "f1": round(f1, 6),
        }
    return {
        "rows": {
            "gold": len(gold_rows),
            "prediction": len(pred_rows),
            "evaluated": max_rows,
        },
        "metrics": metrics,
    }


def print_table(result: dict[str, Any]) -> None:
    print("type,tp,fp,fn,precision,recall,f1")
    for name in (*VOCAB_TYPES, "overall"):
        row = result["metrics"][name]
        print(
            f"{name},{row['tp']},{row['fp']},{row['fn']},"
            f"{row['precision']:.6f},{row['recall']:.6f},{row['f1']:.6f}"
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--gold-xlsx",
        type=Path,
        default=Path("data/src/FinFlier/1_full_used_data.xlsx"),
        help="Gold workbook with binding items in the result column.",
    )
    parser.add_argument("--gold-column", default="result", help="Gold binding column name.")
    parser.add_argument("--gold-sheet", default=None, help="Optional gold sheet name.")
    parser.add_argument("--pred-xlsx", type=Path, default=None, help="Prediction workbook.")
    parser.add_argument("--pred-column", default=None, help="Prediction binding column name.")
    parser.add_argument("--pred-sheet", default=None, help="Optional prediction sheet name.")
    parser.add_argument("--pred-json", type=Path, default=None, help="Prediction JSON or JSONL file.")
    parser.add_argument(
        "--finance-label-xlsx",
        type=Path,
        default=None,
        help="Finance corpus label workbook for presence-label evaluation or summary.",
    )
    parser.add_argument(
        "--gold-source",
        choices=("result", "finance-presence"),
        default="result",
        help="Use result-cell exact labels or Finance workbook presence labels as gold.",
    )
    parser.add_argument("--json", action="store_true", help="Print full JSON result.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    finance_summary = None
    if args.gold_source == "finance-presence":
        if not args.finance_label_xlsx:
            raise SystemExit("--gold-source finance-presence requires --finance-label-xlsx")
        gold_rows, finance_summary = load_finance_presence_xlsx(args.finance_label_xlsx)
    else:
        gold_rows = load_binding_xlsx(args.gold_xlsx, args.gold_column, args.gold_sheet)
        if args.finance_label_xlsx:
            _, finance_summary = load_finance_presence_xlsx(args.finance_label_xlsx)

    if args.pred_json and args.pred_xlsx:
        raise SystemExit("Use only one of --pred-json or --pred-xlsx")
    if args.pred_json:
        pred_rows = load_prediction_json(args.pred_json)
    elif args.pred_xlsx:
        pred_rows = load_binding_xlsx(args.pred_xlsx, args.pred_column, args.pred_sheet)
    else:
        pred_rows = gold_rows

    if args.gold_source == "finance-presence":
        pred_rows = to_presence_rows(pred_rows)

    result = score_rows(gold_rows, pred_rows)
    if finance_summary:
        result["finance_label_summary"] = finance_summary
    result["mode"] = {
        "gold_source": args.gold_source,
        "self_test": not args.pred_json and not args.pred_xlsx,
    }

    if args.json:
        print(json.dumps(result, indent=2, sort_keys=True))
    else:
        print_table(result)
        if finance_summary:
            print("finance_label_rows,", finance_summary["rows"])
            print("finance_label_rows_with_any_label,", finance_summary["rows_with_any_label"])


if __name__ == "__main__":
    main()
