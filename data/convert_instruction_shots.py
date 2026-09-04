#!/usr/bin/env python3
from __future__ import annotations

import argparse
import ast
import csv
import json
import math
import random
import re
import shutil
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

csv.field_size_limit(10**9)

PROMPT_TYPES = ("original", "zero_shot", "many_shot", "dynamic_shot")
SOURCE_FILES = (
    "finqa_train_rel_fact_instruction.csv",
    "finqa_dev_rel_fact_instruction.csv",
    "finqa_test_rel_fact_instruction.csv",
)
REQUIRED_COLUMNS = (
    "Sentence",
    "input",
    "Question",
    "Pre_Text",
    "Post_Text",
    "Table_Text",
    "Rel_Fact",
)
TOKEN_RE = re.compile(r"[A-Za-z0-9]+(?:[._/%-][A-Za-z0-9]+)*")
NUMBER_RE = re.compile(r"[-+]?\d+(?:,\d{3})*(?:\.\d+)?")
DATA_PLACEHOLDERS = (
    "{as the row[`data`] in `1_full_used_data.csv`}",
    "{FinFlier-like `data`}",
)
JSON_QUOTE_ARTIFACTS = ('""RetFact""', '""Binding""', '""Reason""')
FINFLIER_EXAMPLE_COLUMNS = ("data", "text", "result", "reason")


def normalize(text: object) -> str:
    text = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"\s+", " ", text).strip()


def normalize_optional_text(text: object) -> str:
    normalized = normalize(text)
    if normalized.lower() in {"none", "nan", "null"}:
        return ""
    return normalized


def sanitize_cell(value: object) -> str:
    if value is None:
        return ""
    return normalize(str(value)).replace('"', "'")


def coerce_cell_value(value: object) -> object:
    if not isinstance(value, str):
        return value
    stripped = normalize(value)
    if re.fullmatch(r"[-+]?\d+", stripped):
        try:
            return int(stripped)
        except ValueError:
            return stripped
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", stripped):
        try:
            return float(stripped)
        except ValueError:
            return stripped
    return stripped.replace('"', "'")


def format_data_value(value: object) -> str:
    if isinstance(value, str):
        return "'" + value.replace("\\", "\\\\").replace("'", "\\'") + "'"
    if value is None:
        return "None"
    return str(value)


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        missing = [column for column in REQUIRED_COLUMNS if column not in fieldnames]
        if missing:
            raise ValueError(f"{path} is missing required columns: {missing}")
        return fieldnames, list(reader)




def parse_jsonish(raw: object, label: str) -> Any:
    text = normalize(raw)
    if not text:
        raise ValueError(f"{label} is empty")
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        try:
            return ast.literal_eval(text)
        except (SyntaxError, ValueError) as exc:
            raise ValueError(f"{label} is not parseable JSON/Python literal") from exc


def render_finflier_xlsx_example(row: dict[str, str]) -> str:
    payload = {
        "RetFact": row["text"],
        "Binding": json.loads(row["result"]),
        "Reason": row["reason"],
    }
    context = f"{row['data']}\n{row['text']}; question: What data-text bindings are described in the narrative?"
    return "\n".join(["Context:", context, render_compact_result(payload)])


def load_finflier_xlsx_examples(path: Path, sheet: str | None = None) -> tuple[list[dict[str, str]], list[str], list[str], dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = sheet or workbook.sheetnames[0]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{path} has no sheet named {sheet_name!r}; sheets={workbook.sheetnames}")
    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {normalize(value): index for index, value in enumerate(header_row) if normalize(value)}
    missing = [column for column in FINFLIER_EXAMPLE_COLUMNS if column not in header_map]
    if missing:
        raise ValueError(f"{path}:{sheet_name} missing FinFlier example columns: {missing}")

    rows: list[dict[str, str]] = []
    skipped_rows: list[dict[str, object]] = []
    for excel_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        values = {
            column: row[header_map[column]] if header_map[column] < len(row) else None
            for column in FINFLIER_EXAMPLE_COLUMNS
        }
        missing_values = [column for column, value in values.items() if normalize(value) == ""]
        if missing_values:
            skipped_rows.append({"excel_row": excel_row, "reason": "missing_values", "columns": missing_values})
            continue
        try:
            parsed_result = parse_jsonish(values["result"], f"{path}:{sheet_name}:row{excel_row}:result")
        except ValueError as exc:
            skipped_rows.append({"excel_row": excel_row, "reason": "unparseable_result", "error": str(exc)})
            continue
        if not isinstance(parsed_result, list):
            skipped_rows.append({"excel_row": excel_row, "reason": "result_is_not_list"})
            continue
        rows.append(
            {
                "excel_row": str(excel_row),
                "data": normalize(values["data"]),
                "text": normalize(values["text"]),
                "result": compact_json(parsed_result),
                "reason": normalize(values["reason"]),
            }
        )

    if not rows:
        raise ValueError(f"{path}:{sheet_name} has no parseable FinFlier examples")
    rendered = [render_finflier_xlsx_example(row) for row in rows]
    queries = [f"{row['text']} What data-text bindings are described in the narrative?" for row in rows]
    report = {
        "path": str(path),
        "sheet": sheet_name,
        "rows_selected": len(rows),
        "selected_excel_rows": [int(row["excel_row"]) for row in rows],
        "skipped_rows": skipped_rows,
    }
    return rows, rendered, queries, report


def contains_data_label(text: str) -> bool:
    return bool(re.search(r"(?im)^\s*(?:data|tabular\s+data)\s*:", text))


def contains_json_quote_artifact(text: str) -> bool:
    return any(artifact in text for artifact in JSON_QUOTE_ARTIFACTS)


def write_csv_atomic(path: Path, fieldnames: list[str], rows: Iterable[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.tmp")
    with tmp_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    tmp_path.replace(path)


def extract_instruction_template(prompt_path: Path) -> str:
    prompt_text = prompt_path.read_text(encoding="utf-8-sig")
    match = re.search(r"```[ \t]*instruction[ \t]*\n(.*?)\n```", prompt_text, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        match = re.search(r"```[^\n]*\n(.*?)\n```", prompt_text, flags=re.DOTALL)
    if not match:
        raise ValueError(f"No fenced instruction block found in {prompt_path}")
    template = match.group(1).replace("\r\n", "\n").replace("\r", "\n").strip()
    for placeholder in ("{Sentences}", "`Sentences`", "{Question}", "{examples_by_prompt_type}"):
        if placeholder in template:
            break
    else:
        raise ValueError("Prompt template does not contain expected context/example placeholders")
    return template + "\n"


def build_context(row: dict[str, str]) -> str:
    parts = [
        normalize_optional_text(row.get("Pre_Text")),
        normalize_optional_text(row.get("Table_Text")),
        normalize_optional_text(row.get("Post_Text")),
    ]
    context = " ".join(part for part in parts if part)
    return (context or normalize_optional_text(row.get("Sentence"))).rstrip(" ;")


def build_data_records_text(row: dict[str, str]) -> str:
    table = parse_table(row.get("Tables", "[]"))
    if table:
        headers = [
            sanitize_cell(header) or ("Position" if index == 0 else f"Column {index}")
            for index, header in enumerate(table[0])
        ]
        records = []
        for table_row in table[1:]:
            if not isinstance(table_row, list):
                continue
            pairs = []
            for index, header in enumerate(headers):
                value = coerce_cell_value(table_row[index]) if index < len(table_row) else ""
                pairs.append(f"{format_data_value(header)}:{format_data_value(value)}")
            records.append("{" + ",".join(pairs) + "}")
        return "[" + ",".join(records) + "]"
    table_text = normalize(row.get("Table_Text"))
    return table_text.replace('"', "'") if table_text else ""


def build_context_block(row: dict[str, str], context: str) -> str:
    question = normalize(row.get("Question"))
    data_text = build_data_records_text(row)
    context_line = f"{context}; question:{question}"
    if not data_text:
        return context_line
    return f"{data_text}\n{context_line}"


def build_prompt(template: str, data_text: str, context: str, question: str, examples_text: str) -> str:
    rendered = template
    for placeholder in DATA_PLACEHOLDERS:
        rendered = rendered.replace(placeholder, data_text)
    rendered = rendered.replace("{Sentences}", context).replace("`Sentences`", context).replace("{Question}", question)
    if examples_text.strip():
        rendered = rendered.replace("{examples_by_prompt_type}", examples_text.strip())
    else:
        rendered = re.sub(
            r"\n?### Example\s*\n\{examples_by_prompt_type\}\s*\n###\s*",
            "\n",
            rendered,
            flags=re.IGNORECASE,
        ).replace("{examples_by_prompt_type}", "")
    return rendered.strip() + "\n"


def tokenize(text: str) -> list[str]:
    return [token.lower() for token in TOKEN_RE.findall(text)]


def build_document_frequencies(texts: list[str]) -> dict[str, int]:
    dfs: dict[str, int] = {}
    for text in texts:
        for token in set(tokenize(text)):
            dfs[token] = dfs.get(token, 0) + 1
    return dfs


def vectorize(text: str, dfs: dict[str, int], document_count: int) -> dict[str, float]:
    counts = Counter(tokenize(text))
    vector: dict[str, float] = {}
    for token, count in counts.items():
        idf = math.log((document_count + 1) / (dfs.get(token, 0) + 1)) + 1.0
        vector[token] = (1.0 + math.log(count)) * idf
    return vector


def cosine(left: dict[str, float], right: dict[str, float]) -> float:
    if not left or not right:
        return 0.0
    if len(left) > len(right):
        left, right = right, left
    dot = sum(value * right.get(token, 0.0) for token, value in left.items())
    left_norm = math.sqrt(sum(value * value for value in left.values()))
    right_norm = math.sqrt(sum(value * value for value in right.values()))
    if left_norm == 0.0 or right_norm == 0.0:
        return 0.0
    return dot / (left_norm * right_norm)


def parse_table(raw_table: str) -> list[list[object]]:
    try:
        table = ast.literal_eval(str(raw_table or "[]"))
    except (SyntaxError, ValueError):
        return []
    if not isinstance(table, list):
        return []
    return [row for row in table if isinstance(row, list)]


def extract_numbers(text: str) -> list[float]:
    numbers: list[float] = []
    for match in NUMBER_RE.findall(text):
        try:
            numbers.append(float(match.replace(",", "")))
        except ValueError:
            continue
    return numbers


def detect_trend(text: str) -> str:
    lowered = text.lower()
    for keyword in (
        "declined",
        "decreased",
        "dropped",
        "fell",
        "increased",
        "rose",
        "grew",
        "growth",
        "higher",
        "lower",
        "up",
        "down",
        "upgrade trend",
    ):
        if keyword in lowered:
            return keyword
    return "None"


def split_relevant_facts(rel_fact: str) -> list[str]:
    facts = []
    for segment in str(rel_fact or "").replace("\n", " ").split(";"):
        cleaned = normalize(segment).strip(" .")
        if cleaned:
            facts.append(cleaned)
    return facts


def format_retfact(rel_fact: str) -> str:
    facts = split_relevant_facts(rel_fact)
    if not facts:
        return ""
    return "; ".join(facts) + "."


def locate_binding(fact: str, question: str, table: list[list[object]]) -> tuple[list[str], str, list[dict[str, list[int]]]]:
    if not table or not table[0]:
        return [], "", []
    headers = [normalize(value) for value in table[0]]
    haystack = f"{fact} {question}".lower()

    row_label = ""
    row_position = -1
    for row_index, table_row in enumerate(table[1:], start=1):
        if not table_row:
            continue
        candidate = normalize(table_row[0])
        if candidate and candidate.lower() in haystack:
            row_label = candidate
            row_position = row_index - 1
            break

    data_name = ""
    col_position = -1
    for col_index, header in enumerate(headers[1:], start=1):
        if header and header.lower() in haystack:
            data_name = header
            col_position = col_index
            break
    if not data_name and len(headers) > 1:
        data_name = headers[1]
        col_position = 1

    position = []
    if row_position >= 0 and col_position >= 0:
        position = [{"Begin": [row_position, col_position], "End": [row_position, col_position]}]
    return ([row_label] if row_label else []), data_name, position


def build_binding_records(rel_fact: str, question: str, table: list[list[object]]) -> list[dict[str, object]]:
    records = []
    for fact in split_relevant_facts(rel_fact):
        object_names, data_name, position = locate_binding(fact, question, table)
        records.append(
            {
                "ObjectName": object_names,
                "DataName": data_name,
                "Position": position,
                "Trend": detect_trend(fact),
                "Num": extract_numbers(fact),
                "Text": fact,
            }
        )
    return records


def build_result_payload(row: dict[str, str]) -> dict[str, object]:
    rel_fact = format_retfact(row.get("Rel_Fact", ""))
    question = normalize(row.get("Question"))
    table = parse_table(row.get("Tables", "[]"))
    return {
        "RetFact": rel_fact,
        "Binding": build_binding_records(rel_fact, question, table),
        "Reason": "The retrieved facts are generated from the source context/table fields and aligned with the question.",
    }


def compact_json(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def render_compact_result(payload: dict[str, object]) -> str:
    bindings = payload.get("Binding")
    binding_items = bindings if isinstance(bindings, list) else []
    lines = [f"Result: {{\"RetFact\":{compact_json(payload.get('RetFact', ''))},"]
    if binding_items:
        for index, binding in enumerate(binding_items):
            prefix = "\"Binding\":[" if index == 0 else ""
            suffix = "]," if index == len(binding_items) - 1 else ","
            lines.append(f"{prefix}{compact_json(binding)}{suffix}")
    else:
        lines.append("\"Binding\":[],")
    lines.append(f"\"Reason\":{compact_json(payload.get('Reason', ''))}}}")
    return "\n".join(lines)


def render_example(row: dict[str, str]) -> str:
    context = build_context(row)
    payload = build_result_payload(row)
    return "\n".join(["Context:", build_context_block(row, context), render_compact_result(payload)])


def load_examples(path: Path) -> tuple[list[dict[str, str]], list[str], list[str]]:
    _, rows = read_csv_rows(path)
    rendered = [render_example(row) for row in rows]
    queries = [f"{build_context(row)} {normalize(row.get('Question'))}" for row in rows]
    return rows, rendered, queries


def select_many_examples(rendered_examples: list[str], seed: int, source_name: str, row_index: int, count: int) -> str:
    indices = list(range(len(rendered_examples)))
    rng = random.Random(f"{seed}:{source_name}:{row_index}")
    rng.shuffle(indices)
    selected = [rendered_examples[index] for index in indices[: min(count, len(indices))]]
    return render_examples_block(selected)


def select_dynamic_examples(
    query_text: str,
    rendered_examples: list[str],
    example_vectors: list[dict[str, float]],
    dfs: dict[str, int],
    document_count: int,
    count: int,
) -> str:
    query_vector = vectorize(query_text, dfs, document_count)
    scored = [(cosine(query_vector, vector), index) for index, vector in enumerate(example_vectors)]
    scored.sort(key=lambda item: (-item[0], item[1]))
    selected_indices = [index for _, index in scored[: min(count, len(scored))]]
    selected = [rendered_examples[index] for index in reversed(selected_indices)]
    return render_examples_block(selected)


def render_examples_block(examples: list[str]) -> str:
    if not examples:
        return ""
    return "### Example\n" + "\n---\n".join(examples) + "\n###"


def convert_prompt_row(
    row: dict[str, str],
    template: str,
    rendered_examples: list[str],
    example_vectors: list[dict[str, float]],
    dfs: dict[str, int],
    source_name: str,
    row_index: int,
    seed: int,
    many_count: int,
    dynamic_count: int,
    dynamic_rendered_examples: list[str] | None = None,
    dynamic_example_vectors: list[dict[str, float]] | None = None,
    dynamic_dfs: dict[str, int] | None = None,
) -> dict[str, dict[str, str] | str]:
    context = build_context(row)
    question = normalize(row.get("Question"))
    query_text = f"{context} {question}"
    dynamic_rendered_examples = dynamic_rendered_examples if dynamic_rendered_examples is not None else rendered_examples
    dynamic_example_vectors = dynamic_example_vectors if dynamic_example_vectors is not None else example_vectors
    dynamic_dfs = dynamic_dfs if dynamic_dfs is not None else dfs
    data_text = build_data_records_text(row)
    context_block = build_context_block(row, context)

    zero_row = dict(row)
    zero_row["Sentence"] = context
    zero_row["input"] = build_prompt(template, data_text, context, question, "")

    many_row = dict(row)
    many_row["Sentence"] = context
    many_examples = select_many_examples(rendered_examples, seed, source_name, row_index, many_count)
    many_row["input"] = build_prompt(template, data_text, context, question, many_examples)

    dynamic_row = dict(row)
    dynamic_row["Sentence"] = context
    dynamic_examples = select_dynamic_examples(
        query_text,
        dynamic_rendered_examples,
        dynamic_example_vectors,
        dynamic_dfs,
        max(len(dynamic_rendered_examples), 1),
        dynamic_count,
    )
    dynamic_row["input"] = build_prompt(template, data_text, context, question, dynamic_examples)

    return {
        "context": context,
        "data": data_text,
        "context_block": context_block,
        "zero_shot": zero_row,
        "many_shot": many_row,
        "dynamic_shot": dynamic_row,
    }


def convert_rows(
    rows: list[dict[str, str]],
    template: str,
    rendered_examples: list[str],
    example_vectors: list[dict[str, float]],
    dfs: dict[str, int],
    source_name: str,
    seed: int,
    many_count: int,
    dynamic_count: int,
) -> dict[str, list[dict[str, str]]]:
    outputs = {"zero_shot": [], "many_shot": [], "dynamic_shot": []}
    for row_index, row in enumerate(rows):
        converted = convert_prompt_row(
            row,
            template,
            rendered_examples,
            example_vectors,
            dfs,
            source_name,
            row_index,
            seed,
            many_count,
            dynamic_count,
        )
        outputs["zero_shot"].append(converted["zero_shot"])
        outputs["many_shot"].append(converted["many_shot"])
        outputs["dynamic_shot"].append(converted["dynamic_shot"])
    return outputs


def process_source_file(
    source_path: Path,
    output_root: Path,
    template: str,
    rendered_examples: list[str],
    example_vectors: list[dict[str, float]],
    dfs: dict[str, int],
    dynamic_rendered_examples: list[str],
    dynamic_example_vectors: list[dict[str, float]],
    dynamic_dfs: dict[str, int],
    args: argparse.Namespace,
) -> dict[str, object]:
    fieldnames, source_rows = read_csv_rows(source_path)
    output_paths = {
        "original": output_root / "finqa_original" / source_path.name,
        "zero_shot": output_root / "finqa_zero_shot" / source_path.name,
        "many_shot": output_root / "finqa_many_shot" / source_path.name,
        "dynamic_shot": output_root / "finqa_dynamic_shot" / source_path.name,
    }

    requested_prompt_types = tuple(args.prompt_types)
    preview: dict[str, object] = {}
    if not args.dry_run:
        if "original" in requested_prompt_types:
            output_paths["original"].parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(source_path, output_paths["original"])
        for prompt_type in ("zero_shot", "many_shot", "dynamic_shot"):
            if prompt_type in requested_prompt_types:
                output_paths[prompt_type].parent.mkdir(parents=True, exist_ok=True)

    writer_handles = {}
    writer_tmp_paths = {}
    try:
        if not args.dry_run:
            for prompt_type in ("zero_shot", "many_shot", "dynamic_shot"):
                if prompt_type not in requested_prompt_types:
                    continue
                tmp_path = output_paths[prompt_type].with_name(f".{output_paths[prompt_type].name}.tmp")
                handle = tmp_path.open("w", encoding="utf-8", newline="")
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                writer_handles[prompt_type] = (handle, writer)
                writer_tmp_paths[prompt_type] = tmp_path

        rows_to_process = source_rows[:1] if args.dry_run else source_rows
        for row_index, row in enumerate(rows_to_process):
            converted = convert_prompt_row(
                row,
                template,
                rendered_examples,
                example_vectors,
                dfs,
                source_path.name,
                row_index,
                args.seed,
                args.many_count,
                args.dynamic_count,
                dynamic_rendered_examples,
                dynamic_example_vectors,
                dynamic_dfs,
            )
            zero_row = converted["zero_shot"]
            many_row = converted["many_shot"]
            dynamic_row = converted["dynamic_shot"]

            if row_index == 0:
                preview["source_input_len"] = len(row.get("input", ""))
                for prompt_type, converted_row in (
                    ("zero_shot", zero_row),
                    ("many_shot", many_row),
                    ("dynamic_shot", dynamic_row),
                ):
                    preview[f"{prompt_type}_input_len"] = len(converted_row.get("input", ""))
                    preview[f"{prompt_type}_input_preview"] = converted_row.get("input", "")[:240]

            if not args.dry_run:
                if "zero_shot" in writer_handles:
                    writer_handles["zero_shot"][1].writerow(zero_row)
                if "many_shot" in writer_handles:
                    writer_handles["many_shot"][1].writerow(many_row)
                if "dynamic_shot" in writer_handles:
                    writer_handles["dynamic_shot"][1].writerow(dynamic_row)
    finally:
        for handle, _writer in writer_handles.values():
            handle.close()

    if not args.dry_run:
        for prompt_type, tmp_path in writer_tmp_paths.items():
            tmp_path.replace(output_paths[prompt_type])

    return {
        "source": str(source_path),
        "rows": len(source_rows),
        "outputs": {key: str(path) for key, path in output_paths.items() if key in requested_prompt_types},
        "preview": preview,
        "fieldnames": fieldnames,
    }


def write_few_file(output_root: Path, train_zero_path: Path, dry_run: bool) -> str | None:
    if not train_zero_path.exists() and not dry_run:
        return None
    few_path = output_root / "testing" / "finqa_10_rel_fact_instruction.csv"
    if dry_run:
        return str(few_path)
    few_path.parent.mkdir(parents=True, exist_ok=True)
    with train_zero_path.open("r", encoding="utf-8-sig", newline="") as source_handle:
        reader = csv.DictReader(source_handle)
        fieldnames = list(reader.fieldnames or [])
        tmp_path = few_path.with_name(f".{few_path.name}.tmp")
        with tmp_path.open("w", encoding="utf-8", newline="") as target_handle:
            writer = csv.DictWriter(target_handle, fieldnames=fieldnames)
            writer.writeheader()
            for row_index, row in enumerate(reader):
                if row_index >= 10:
                    break
                writer.writerow(row)
        tmp_path.replace(few_path)
    return str(few_path)


def validate_outputs(output_root: Path, expected_counts: dict[str, int], prompt_types: tuple[str, ...] = PROMPT_TYPES) -> dict[str, object]:
    report: dict[str, object] = {}
    for prompt_type in prompt_types:
        prompt_dir = output_root / f"finqa_{prompt_type}"
        prompt_report = {}
        for file_name, expected_count in expected_counts.items():
            output_file = prompt_dir / file_name
            with output_file.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                fieldnames = list(reader.fieldnames or [])
                row_count = 0
                sample_input = ""
                for row in reader:
                    if row_count == 0:
                        sample_input = row.get("input", "")
                    if prompt_type != "original" and row.get("Sentence", "") != build_context(row):
                        raise AssertionError(f"{output_file} row {row_count + 1} Sentence is not generated from context fields")
                    row_count += 1
            if row_count != expected_count:
                raise AssertionError(f"{output_file} expected {expected_count} rows, found {row_count}")
            if prompt_type != "original":
                context_block = extract_prompt_context_block(sample_input)
                for placeholder in (
                    "{FinFlier-like `data`}",
                    "{as the row[`data`] in `1_full_used_data.csv`}",
                    "{Sentences}",
                    "{Question}",
                    "{examples_by_prompt_type}",
                    "`Sentences`",
                    "`prompt-type`",
                ):
                    if placeholder in sample_input:
                        raise AssertionError(f"{output_file} still contains placeholder {placeholder}")
                if contains_data_label(sample_input):
                    raise AssertionError(f"{output_file} contains a data label in the rendered context")
                if contains_json_quote_artifact(sample_input):
                    raise AssertionError(f"{output_file} contains parsed JSON quote artifacts")
                if contains_data_label(context_block):
                    raise AssertionError(f"{output_file} Context block contains a data label")
                if '""""' in context_block:
                    raise AssertionError(f"{output_file} Context block contains CSV quote artifacts")
            prompt_report[file_name] = {"rows": row_count, "columns": len(fieldnames)}
        report[prompt_type] = prompt_report
    few_path = output_root / "testing" / "finqa_10_rel_fact_instruction.csv"
    if few_path.exists():
        with few_path.open("r", encoding="utf-8-sig", newline="") as handle:
            few_count = sum(1 for _ in csv.DictReader(handle))
        if few_count != 10:
            raise AssertionError(f"{few_path} expected 10 rows, found {few_count}")
        report["testing"] = {"finqa_10_rel_fact_instruction.csv": {"rows": few_count}}
    return report


def extract_prompt_context_block(input_text: str) -> str:
    match = re.search(r"## Context\s*\n(.*?)(?:\n\n## Output Format|\n\n### Example|\Z)", input_text, flags=re.DOTALL)
    return match.group(1).strip() if match else ""


def preview_single_row(
    source_path: Path,
    row_index: int,
    template: str,
    rendered_examples: list[str],
    example_vectors: list[dict[str, float]],
    dfs: dict[str, int],
    dynamic_rendered_examples: list[str],
    dynamic_example_vectors: list[dict[str, float]],
    dynamic_dfs: dict[str, int],
    args: argparse.Namespace,
) -> dict[str, object]:
    fieldnames, source_rows = read_csv_rows(source_path)
    if row_index < 0 or row_index >= len(source_rows):
        raise IndexError(f"preview row index {row_index} is out of range for {source_path} ({len(source_rows)} rows)")
    preview_count = max(args.preview_example_count, 0)
    converted = convert_prompt_row(
        source_rows[row_index],
        template,
        rendered_examples,
        example_vectors,
        dfs,
        source_path.name,
        row_index,
        args.seed,
        preview_count,
        preview_count,
        dynamic_rendered_examples,
        dynamic_example_vectors,
        dynamic_dfs,
    )
    example_block = render_examples_block(rendered_examples[: min(preview_count, len(rendered_examples))])
    return {
        "source": str(source_path),
        "row_index": row_index,
        "fieldnames": fieldnames,
        "Sentence": converted["context"],
        "context": converted["context"],
        "data_records": converted["data"],
        "context_block": converted["context_block"],
        "zero_shot_input": converted["zero_shot"]["input"],
        "many_shot_input": converted["many_shot"]["input"],
        "dynamic_shot_input": converted["dynamic_shot"]["input"],
        "compact_example_block": example_block,
        "checks": {
            "contains_data_prefix": any(
                contains_data_label(text)
                for text in (
                    converted["zero_shot"]["input"],
                    converted["many_shot"]["input"],
                    converted["dynamic_shot"]["input"],
                    example_block,
                )
            ),
            "contains_sentences_placeholder": any(
                "`Sentences`" in text or "{Sentences}" in text
                for text in (
                    converted["zero_shot"]["input"],
                    converted["many_shot"]["input"],
                    converted["dynamic_shot"]["input"],
                )
            ),
            "contains_prompt_type_placeholder": any(
                "`prompt-type`" in text
                for text in (
                    converted["zero_shot"]["input"],
                    converted["many_shot"]["input"],
                    converted["dynamic_shot"]["input"],
                )
            ),
            "contains_context_quote_artifact": '""""' in str(converted["context_block"]),
            "uses_data_records": str(converted["context_block"]).startswith("[{"),
        },
    }


def print_preview_text(preview: dict[str, object]) -> None:
    print("=== Sentence ===")
    print(preview["Sentence"])
    print("=== Context Block ===")
    print(preview["context_block"])
    print("=== Zero-Shot Input ===")
    print(str(preview["zero_shot_input"]).rstrip())
    print("=== Compact Example Block ===")
    print(str(preview["compact_example_block"]).rstrip())


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate FinQA prompt-type instruction CSV directories.")
    parser.add_argument("--source-root", type=Path, default=Path("data/src/Data_Retriever_Module"))
    parser.add_argument("--output-root", type=Path, default=Path("data"))
    parser.add_argument("--prompt", type=Path, default=Path("new_prompt.txt"))
    parser.add_argument("--examples", type=Path, default=Path("data/src/full_example.csv"))
    parser.add_argument("--dynamic-examples-xlsx", type=Path, default=Path("data/src/FinFlier/1_full_used_data.xlsx"))
    parser.add_argument("--dynamic-examples-sheet", default="")
    parser.add_argument("--prompt-types", default=",".join(PROMPT_TYPES), help="Comma-separated output prompt types to write.")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--many-count", type=int, default=26)
    parser.add_argument("--dynamic-count", type=int, default=10)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-few", action="store_true")
    parser.add_argument("--preview-only", action="store_true", help="Print one generated sample without writing output files.")
    parser.add_argument("--preview-source", type=Path, default=None, help="CSV to use for --preview-only.")
    parser.add_argument("--preview-row", type=int, default=0)
    parser.add_argument("--preview-example-count", type=int, default=1)
    parser.add_argument("--preview-text", action="store_true")
    args = parser.parse_args()
    args.prompt_types = tuple(part.strip() for part in args.prompt_types.split(",") if part.strip())
    invalid_prompt_types = [prompt_type for prompt_type in args.prompt_types if prompt_type not in PROMPT_TYPES]
    if invalid_prompt_types:
        raise ValueError(f"Unsupported prompt types: {invalid_prompt_types}; supported={PROMPT_TYPES}")

    if not args.examples.exists():
        fallback_examples = Path("data/src/full_example.csv")
        if fallback_examples.exists():
            args.examples = fallback_examples
        else:
            raise FileNotFoundError(f"Missing examples CSV: {args.examples}")

    template = extract_instruction_template(args.prompt)
    _, rendered_examples, example_queries = load_examples(args.examples)
    dfs = build_document_frequencies(example_queries)
    example_vectors = [vectorize(query, dfs, max(len(example_queries), 1)) for query in example_queries]
    dynamic_example_rows, dynamic_rendered_examples, dynamic_example_queries, dynamic_examples_report = load_finflier_xlsx_examples(
        args.dynamic_examples_xlsx,
        args.dynamic_examples_sheet or None,
    )
    dynamic_dfs = build_document_frequencies(dynamic_example_queries)
    dynamic_example_vectors = [
        vectorize(query, dynamic_dfs, max(len(dynamic_example_queries), 1))
        for query in dynamic_example_queries
    ]

    if args.preview_only:
        source_path = args.preview_source or args.source_root / "finqa_test_rel_fact_instruction.csv"
        if not source_path.exists():
            raise FileNotFoundError(f"Missing preview source CSV: {source_path}")
        preview = preview_single_row(
            source_path,
            args.preview_row,
            template,
            rendered_examples,
            example_vectors,
            dfs,
            dynamic_rendered_examples,
            dynamic_example_vectors,
            dynamic_dfs,
            args,
        )
        if args.preview_text:
            print_preview_text(preview)
            return
        print(
            json.dumps(
                preview,
                ensure_ascii=False,
                indent=2,
            )
        )
        return

    source_files = [args.source_root / file_name for file_name in SOURCE_FILES]
    missing_sources = [str(path) for path in source_files if not path.exists()]
    if missing_sources:
        raise FileNotFoundError(f"Missing source CSV files: {missing_sources}")

    reports = []
    expected_counts: dict[str, int] = {}
    for source_path in source_files:
        _, source_rows = read_csv_rows(source_path)
        expected_counts[source_path.name] = len(source_rows)
        reports.append(
            process_source_file(
                source_path,
                args.output_root,
                template,
                rendered_examples,
                example_vectors,
                dfs,
                dynamic_rendered_examples,
                dynamic_example_vectors,
                dynamic_dfs,
                args,
            )
        )

    few_output = None
    if not args.skip_few:
        few_output = write_few_file(
            args.output_root,
            args.output_root / "finqa_zero_shot" / "finqa_train_rel_fact_instruction.csv",
            args.dry_run,
        )

    validation = None
    preview_path = args.output_root / "generation_preview.json"
    if not args.dry_run:
        validation = validate_outputs(args.output_root, expected_counts, args.prompt_types)
        preview_path.write_text(json.dumps(reports, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(
        json.dumps(
            {
                "dry_run": args.dry_run,
                "source_root": str(args.source_root),
                "output_root": str(args.output_root),
                "prompt": str(args.prompt),
                "examples": str(args.examples),
                "dynamic_examples_source": dynamic_examples_report,
                "prompt_types": list(args.prompt_types),
                "many_count": args.many_count,
                "dynamic_count": args.dynamic_count,
                "source_files": {path.name: expected_counts[path.name] for path in source_files},
                "reports": reports,
                "few_output": few_output,
                "preview_path": None if args.dry_run else str(preview_path),
                "validation": validation,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
